"""Excel exporter for complimentary ticket invitations."""

import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.conf import settings
from django.db.models import QuerySet

from apps.events.models import ComplimentaryTicketInvitation

logger = logging.getLogger(__name__)


def export_to_excel(invitations: QuerySet) -> BytesIO:
    """
    Export invitations to Excel file.
    
    Args:
        invitations: QuerySet of ComplimentaryTicketInvitation instances
        
    Returns:
        BytesIO object with Excel file content
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Cortesías"
    
    # Setup headers
    _setup_excel_headers(worksheet)
    
    # Add data rows
    frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:8080')
    _add_excel_data_rows(worksheet, invitations, frontend_url)
    
    # Auto-adjust column widths
    _adjust_excel_column_widths(worksheet)
    
    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    logger.info(f"Exported {invitations.count()} invitations to Excel")
    
    return output


def _setup_excel_headers(worksheet):
    """Setup Excel worksheet headers with styling."""
    headers = [
        'Nombre',
        'Apellido',
        'Email',
        'Link Público',
        'Estado',
        'Fecha Canje',
        'Tickets Permitidos'
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _add_excel_data_rows(worksheet, invitations, frontend_url: str):
    """Add data rows to Excel worksheet."""
    status_display = {
        'pending': 'Pendiente',
        'redeemed': 'Canjeada',
        'cancelled': 'Cancelada'
    }
    
    for row_idx, invitation in enumerate(invitations, start=2):
        # Name
        worksheet.cell(row=row_idx, column=1, value=invitation.first_name or '')
        
        # Last name
        worksheet.cell(row=row_idx, column=2, value=invitation.last_name or '')
        
        # Email
        worksheet.cell(row=row_idx, column=3, value=invitation.email or '')
        
        # Public link
        public_url = invitation.generate_public_url()
        cell = worksheet.cell(row=row_idx, column=4, value=public_url)
        cell.hyperlink = public_url
        cell.font = Font(color="0563C1", underline="single")
        
        # Status
        worksheet.cell(row=row_idx, column=5, value=status_display.get(invitation.status, invitation.status))
        
        # Redeemed date
        if invitation.redeemed_at:
            worksheet.cell(row=row_idx, column=6, value=invitation.redeemed_at.strftime('%Y-%m-%d %H:%M:%S'))
        else:
            worksheet.cell(row=row_idx, column=6, value='')
        
        # Max tickets
        worksheet.cell(row=row_idx, column=7, value=invitation.max_tickets)


def _adjust_excel_column_widths(worksheet):
    """Adjust Excel column widths for better readability."""
    column_widths = {
        'A': 20,  # Nombre
        'B': 20,  # Apellido
        'C': 30,  # Email
        'D': 50,  # Link Público
        'E': 15,  # Estado
        'F': 20,  # Fecha Canje
        'G': 18   # Tickets Permitidos
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

