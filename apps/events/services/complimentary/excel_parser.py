"""Excel file parser for complimentary ticket invitations."""

import logging
from typing import Dict, List, Tuple
from openpyxl import load_workbook

from .column_detector import detect_columns, normalize_column_name

logger = logging.getLogger(__name__)


def parse_excel_file(file_obj) -> Tuple[List[Dict], List[str]]:
    """
    Parse Excel file and extract invitation data.
    
    Args:
        file_obj: File object (Excel .xlsx)
        
    Returns:
        Tuple of (list of data dicts, list of error messages)
    """
    entries = []
    errors = []
    
    try:
        # Load workbook
        workbook = load_workbook(file_obj, data_only=True)
        
        # Use first sheet
        if not workbook.sheetnames:
            errors.append("Excel file has no sheets")
            return entries, errors
        
        worksheet = workbook[workbook.sheetnames[0]]
        
        # Find header row (first non-empty row)
        header_row_idx = None
        for row_idx, row in enumerate(worksheet.iter_rows(max_row=10), start=1):
            if any(cell.value for cell in row):
                header_row_idx = row_idx
                break
        
        if header_row_idx is None:
            errors.append("Could not find header row in Excel file")
            workbook.close()
            return entries, errors
        
        # Extract headers
        header_row = worksheet[header_row_idx]
        headers = [str(cell.value).strip() if cell.value else '' for cell in header_row]
        
        # Auto-detect column mapping
        column_mapping = detect_columns(headers)
        
        # Parse data rows
        entries = _parse_excel_rows(worksheet, header_row_idx, column_mapping)
        
        workbook.close()
        logger.info(f"Parsed {len(entries)} entries from Excel file")
        
    except Exception as e:
        error_msg = f"Error parsing Excel file: {str(e)}"
        logger.error(error_msg, exc_info=True)
        errors.append(error_msg)
    
    return entries, errors


def _parse_excel_rows(worksheet, header_row_idx: int, column_mapping: Dict) -> List[Dict]:
    """Parse data rows from Excel worksheet."""
    entries = []
    data_start_row = header_row_idx + 1
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=data_start_row), start=data_start_row):
        # Skip empty rows
        if not any(cell.value for cell in row):
            continue
        
        entry = _extract_row_data(row, column_mapping)
        
        # Only add if has at least first_name or email
        if entry.get('first_name') or entry.get('email'):
            entries.append(entry)
    
    return entries


def _extract_row_data(row, column_mapping: Dict) -> Dict:
    """Extract data from a single row based on column mapping."""
    entry = {}
    row_values = [cell.value for cell in row]
    
    # Extract first_name
    if column_mapping['first_name'] is not None:
        idx = column_mapping['first_name']
        if idx < len(row_values):
            value = row_values[idx]
            entry['first_name'] = str(value).strip() if value else ''
    
    # Extract last_name
    if column_mapping['last_name'] is not None:
        idx = column_mapping['last_name']
        if idx < len(row_values):
            value = row_values[idx]
            entry['last_name'] = str(value).strip() if value else ''
    
    # Extract email
    if column_mapping['email'] is not None:
        idx = column_mapping['email']
        if idx < len(row_values):
            value = row_values[idx]
            email = str(value).strip() if value else ''
            # Handle multiple emails separated by semicolon
            if ';' in email:
                email = email.split(';')[0].strip()  # Take first email
            entry['email'] = email
    
    return entry

