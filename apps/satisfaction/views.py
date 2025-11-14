"""
Views for Satisfaction Survey System
"""

import logging
from django.db.models import Avg, Count, Q, Min, Max
from django.db import models
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    SatisfactionSurvey,
    SatisfactionQuestion,
    SatisfactionQuestionOption,
    SatisfactionResponse,
    SatisfactionAnswer
)
from .serializers import (
    SatisfactionSurveySerializer,
    SatisfactionSurveyListSerializer,
    SatisfactionQuestionSerializer,
    SatisfactionResponseSerializer,
    SatisfactionResponseCreateSerializer,
    SatisfactionStatisticsSerializer
)

logger = logging.getLogger(__name__)


class SatisfactionSurveyViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de formularios de satisfacción (SuperAdmin).
    """
    queryset = SatisfactionSurvey.objects.all()
    serializer_class = SatisfactionSurveySerializer
    permission_classes = [AllowAny]  # TODO: Cambiar a IsAdminUser en producción
    lookup_field = 'slug'
    
    def get_serializer_class(self):
        if self.action == 'list':
            return SatisfactionSurveyListSerializer
        return SatisfactionSurveySerializer
    
    def get_queryset(self):
        queryset = SatisfactionSurvey.objects.select_related('event', 'organizer').prefetch_related('questions')
        
        # Filtros opcionales
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        event_id = self.request.query_params.get('event')
        if event_id:
            queryset = queryset.filter(event_id=event_id)
        
        organizer_id = self.request.query_params.get('organizer')
        if organizer_id:
            queryset = queryset.filter(organizer_id=organizer_id)
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, slug=None):
        """Obtener estadísticas del formulario."""
        survey = self.get_object()
        
        # Calcular estadísticas
        total_responses = survey.responses.count()
        completion_rate = survey.completion_rate
        
        # Promedios de ratings por pregunta
        average_ratings = {}
        question_stats = []
        
        for question in survey.questions.all():
            if question.question_type in ['rating', 'rating_5', 'rating_10']:
                avg = question.answers.filter(numeric_value__isnull=False).aggregate(
                    avg=Avg('numeric_value')
                )['avg']
                count = question.answers.filter(numeric_value__isnull=False).count()
                
                if avg is not None:
                    # Determinar el máximo de rating según el tipo de pregunta
                    max_rating = question.max_rating or (10 if question.question_type == 'rating_10' else 5)
                    
                    average_ratings[question.id] = {
                        'question': question.question_text,
                        'average': round(avg, 2),
                        'count': count,
                        'min': question.answers.filter(numeric_value__isnull=False).aggregate(
                            min=models.Min('numeric_value')
                        )['min'],
                        'max': question.answers.filter(numeric_value__isnull=False).aggregate(
                            max=models.Max('numeric_value')
                        )['max'],
                        'max_rating': max_rating,  # Agregar el máximo de rating
                    }
            
            # Estadísticas generales por pregunta
            stats = {
                'question_id': question.id,
                'question_text': question.question_text,
                'question_type': question.question_type,
                'total_answers': question.answers.count(),
            }
            
            if question.question_type == 'multiple_choice':
                option_counts = question.options.annotate(
                    count=Count('satisfactionanswer')
                ).values('option_text', 'count')
                stats['option_counts'] = list(option_counts)
            
            question_stats.append(stats)
        
        data = {
            'total_responses': total_responses,
            'completion_rate': completion_rate,
            'average_ratings': average_ratings,
            'question_statistics': question_stats
        }
        
        serializer = SatisfactionStatisticsSerializer(data)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def export_excel(self, request, slug=None):
        """Exportar respuestas a Excel."""
        survey = self.get_object()
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Respuestas"
        
        # Estilos
        header_fill = PatternFill(start_color="572150", end_color="572150", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Encabezados
        headers = ['ID', 'Fecha', 'Email', 'Nombre']
        questions = survey.questions.all().order_by('order')
        for question in questions:
            headers.append(question.question_text)
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Escribir datos
        responses = survey.responses.select_related('ticket').prefetch_related(
            'answers__question', 'answers__option'
        ).order_by('-submitted_at')
        
        for row_idx, response in enumerate(responses, 2):
            ws.cell(row=row_idx, column=1).value = str(response.id)
            ws.cell(row=row_idx, column=2).value = response.submitted_at.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_idx, column=3).value = response.email or ''
            ws.cell(row=row_idx, column=4).value = response.name or ''
            
            # Respuestas por pregunta
            answers_dict = {answer.question_id: answer for answer in response.answers.all()}
            for col_idx, question in enumerate(questions, 5):
                answer = answers_dict.get(question.id)
                if answer:
                    if answer.numeric_value is not None:
                        value = answer.numeric_value
                    elif answer.text_value:
                        value = answer.text_value
                    elif answer.option:
                        value = answer.option.option_text
                    else:
                        value = ''
                else:
                    value = ''
                ws.cell(row=row_idx, column=col_idx).value = value
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"satisfaction_{survey.slug}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def public_survey_view(request, slug):
    """
    Vista pública para obtener y responder formularios de satisfacción.
    GET: Obtener formulario activo
    POST: Enviar respuesta
    """
    try:
        survey = SatisfactionSurvey.objects.prefetch_related(
            'questions__options'
        ).get(slug=slug)
    except SatisfactionSurvey.DoesNotExist:
        return Response(
            {'error': 'Formulario no encontrado'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    if request.method == 'GET':
        # Validar que esté activo
        if not survey.is_active:
            return Response(
                {'error': 'Este formulario no está disponible actualmente'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = SatisfactionSurveySerializer(survey)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        # Validar que esté activo
        if not survey.is_active:
            return Response(
                {'error': 'Este formulario no está disponible actualmente'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Agregar metadata
        data = request.data.copy()
        data['survey'] = survey.id
        if 'ip_address' not in data:
            data['ip_address'] = get_client_ip(request)
        if 'user_agent' not in data:
            data['user_agent'] = request.META.get('HTTP_USER_AGENT', '')
        
        serializer = SatisfactionResponseCreateSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {'success': True, 'message': 'Respuesta enviada exitosamente'},
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def get_client_ip(request):
    """Obtener IP del cliente."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

