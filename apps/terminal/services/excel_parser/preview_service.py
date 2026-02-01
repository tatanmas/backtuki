"""Service for previewing Excel uploads without saving to database."""

import logging
from datetime import date
from typing import Dict, List
from tempfile import NamedTemporaryFile
import os

from apps.terminal.models import TerminalTrip, TerminalRoute, TerminalCompany
from apps.terminal.services.company_service import get_or_create_company
from .file_parser import parse_excel_file

logger = logging.getLogger(__name__)


def preview_excel_upload(
    file_obj,
    upload_type: str,
    date_range_start: date,
    date_range_end: date
) -> Dict:
    """
    Preview Excel file and return what would be created/updated without saving.
    
    Args:
        file_obj: Uploaded file object
        upload_type: 'departures' or 'arrivals'
        date_range_start: Start date of schedule range
        date_range_end: End date of schedule range
    
    Returns:
        Dictionary with:
            - column_mapping: Dict showing Excel columns -> system fields
            - trips_preview: List of trip data that would be created/updated
            - existing_trips: List of trips that already exist (would be updated)
            - new_trips: List of trips that would be created
            - processed_sheets: List of sheet names
            - errors: List of error messages
            - summary: Dict with counts
    """
    # Save uploaded file temporarily
    with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        for chunk in file_obj.chunks():
            tmp_file.write(chunk)
        tmp_file_path = tmp_file.name
    
    try:
        logger.info("=" * 80)
        logger.info(f"🚀 [preview_excel_upload] STARTING PREVIEW")
        logger.info(f"📂 File: {file_obj.name} ({file_obj.size} bytes)")
        logger.info(f"📋 Upload type: {upload_type}")
        logger.info(f"📅 Date range: {date_range_start} to {date_range_end}")
        logger.info("=" * 80)
        
        # Parse Excel file (without saving)
        logger.info("📖 [preview_excel_upload] Parsing Excel file...")
        parse_result = parse_excel_file(tmp_file_path, upload_type, date_range_start, date_range_end)
        trips_data = parse_result['trips']
        processed_sheets = parse_result['processed_sheets']
        errors = parse_result['errors']
        
        logger.info(f"✅ [preview_excel_upload] Parsed {len(trips_data)} trips from {len(processed_sheets)} sheets")
        logger.info(f"📄 [preview_excel_upload] Processed sheets: {processed_sheets}")
        if errors:
            logger.warning(f"❌ [preview_excel_upload] Found {len(errors)} errors during parsing")
            for error in errors[:5]:  # Log first 5 errors
                logger.warning(f"   - {error}")
        
        # Get column mapping from first sheet (if available)
        logger.info("🗺️  [preview_excel_upload] Extracting column mapping...")
        column_mapping = _get_column_mapping(tmp_file_path, upload_type)
        logger.info(f"✅ [preview_excel_upload] Column mapping extracted: {len(column_mapping)} columns")
        logger.info(f"🗺️  [preview_excel_upload] Mapping details: {column_mapping}")
        
        # Check which trips exist and which are new
        existing_trips = []
        new_trips = []
        
        for trip_data in trips_data:
            try:
                # Get or create company (just for lookup, don't save)
                company = get_or_create_company(trip_data['operator'])
                
                # Get or create route (just for lookup, don't save)
                route, _ = TerminalRoute.objects.get_or_create(
                    origin=trip_data['origin'],
                    destination=trip_data['destination'],
                    defaults={'duration': None, 'distance': None}
                )
                
                # Check if trip exists
                lookup = {
                    'company': company,
                    'route': route,
                    'date': trip_data['date'],
                    'trip_type': trip_data['trip_type'],
                }
                
                if trip_data['trip_type'] == 'departure':
                    lookup['departure_time'] = trip_data['departure_time']
                    lookup['arrival_time'] = None
                else:
                    lookup['arrival_time'] = trip_data['arrival_time']
                    lookup['departure_time'] = None
                
                existing_trip = TerminalTrip.objects.filter(**lookup).first()
                
                trip_preview = {
                    'operator': trip_data['operator'],
                    'origin': trip_data['origin'],
                    'destination': trip_data['destination'],
                    'date': trip_data['date'].isoformat(),
                    'tripType': trip_data['trip_type'],  # camelCase for frontend
                    'departureTime': trip_data.get('departure_time').strftime('%H:%M') if trip_data.get('departure_time') else None,
                    'arrivalTime': trip_data.get('arrival_time').strftime('%H:%M') if trip_data.get('arrival_time') else None,
                    'platform': trip_data.get('platform'),
                    'licensePlate': trip_data.get('license_plate'),  # camelCase
                    'observations': trip_data.get('observations'),
                    'companyId': str(company.id) if company.id else None,  # camelCase
                    'routeId': str(route.id) if route.id else None,  # camelCase
                    'willCreate': existing_trip is None,  # camelCase
                    'willUpdate': existing_trip is not None,  # camelCase
                    'existingTripId': str(existing_trip.id) if existing_trip else None,  # camelCase
                    'existingStatus': existing_trip.status if existing_trip else None,
                    'existingIsActive': existing_trip.is_active if existing_trip else None,  # camelCase
                }
                
                if existing_trip:
                    existing_trips.append(trip_preview)
                else:
                    new_trips.append(trip_preview)
                    
            except Exception as e:
                error_msg = f"Error previewing trip: {e}"
                logger.error(error_msg, exc_info=True)
                errors.append(error_msg)
        
        # Summary
        summary = {
            'total_trips': len(trips_data),
            'new_trips': len(new_trips),
            'existing_trips': len(existing_trips),
            'processed_sheets': len(processed_sheets),
            'errors_count': len(errors),
        }
        
        # Return all trips (not limited) for preview
        return {
            'column_mapping': column_mapping,
            'trips_preview': trips_data,  # All trips for preview
            'existing_trips': existing_trips,
            'new_trips': new_trips,
            'processed_sheets': processed_sheets,
            'errors': errors,
            'summary': summary,
        }
        
    finally:
        # Clean up temp file
        if os.path.exists(tmp_file_path):
            os.unlink(tmp_file_path)


def _get_column_mapping(file_path: str, upload_type: str) -> Dict:
    """Extract column mapping from Excel file with improved robustness."""
    from openpyxl import load_workbook
    from .header_finder import find_header_row, find_header_row_alternative, normalize_header
    from .header_mapping import get_expected_headers
    
    logger.info(f"🔍 Starting column mapping extraction for upload_type: {upload_type}")
    
    try:
        workbook = load_workbook(file_path, data_only=True)
        expected_headers = get_expected_headers(upload_type)
        logger.info(f"📋 Expected headers for {upload_type}: {list(expected_headers.keys())}")
        
        mapping = {}
        all_excel_columns = []  # Track all columns found for debugging
        
        # Find first sheet with date
        logger.info(f"📄 Available sheets: {workbook.sheetnames}")
        for sheet_name in workbook.sheetnames:
            if any(day in sheet_name.lower() for day in ['lunes', 'martes', 'miercoles', 'miércoles', 'jueves', 'viernes', 'sabado', 'sábado', 'domingo']):
                logger.info(f"🔎 Processing sheet: {sheet_name}")
                worksheet = workbook[sheet_name]
                
                # Try primary method first
                header_row_idx = find_header_row(worksheet)
                logger.info(f"📍 Primary method found header row at index: {header_row_idx}")
                if header_row_idx is None:
                    # Try alternative method
                    header_row_idx = find_header_row_alternative(worksheet)
                    logger.info(f"📍 Alternative method found header row at index: {header_row_idx}")
                
                if header_row_idx is not None:
                    logger.info(f"✅ Using header row at index: {header_row_idx}")
                    header_row = worksheet[header_row_idx]
                    
                    # Extract all column headers first
                    for col_idx, cell in enumerate(header_row):
                        if cell.value:
                            excel_col = str(cell.value).strip()
                            if excel_col:  # Only non-empty columns
                                all_excel_columns.append(excel_col)
                                normalized = normalize_header(excel_col)
                                
                                logger.debug(f"Processing column: '{excel_col}' -> normalized: '{normalized}'")
                                
                                # Check if normalized matches any expected header exactly
                                if normalized in expected_headers:
                                    system_field = expected_headers[normalized]
                                    if system_field:  # Skip None mappings
                                        mapping[excel_col] = system_field
                                        logger.info(f"✓ Mapped '{excel_col}' -> '{system_field}' (exact match)")
                                else:
                                    # Try partial matching for more flexibility
                                    matched = False
                                    for expected_key, system_field in expected_headers.items():
                                        if not system_field:
                                            continue
                                        # Check if normalized contains expected_key or vice versa
                                        if expected_key in normalized or normalized in expected_key:
                                            # Additional check: ensure it's a meaningful match (not just a substring)
                                            if len(expected_key) >= 3 and len(normalized) >= 3:
                                                mapping[excel_col] = system_field
                                                logger.info(f"✓ Mapped '{excel_col}' -> '{system_field}' (partial match: '{expected_key}' in '{normalized}')")
                                                matched = True
                                                break
                                    if not matched:
                                        logger.warning(f"✗ No mapping found for column: '{excel_col}' (normalized: '{normalized}')")
                                        logger.debug(f"  Available expected headers: {list(expected_headers.keys())}")
                    
                    logger.info(f"✅ Found {len(mapping)} mapped columns out of {len(all_excel_columns)} total columns")
                    logger.info(f"📊 Excel columns found: {all_excel_columns}")
                    logger.info(f"🗺️  Final mapping: {mapping}")
                    
                    if mapping:  # If we found at least one mapping, we're good
                        logger.info(f"✅ Successfully extracted mapping from sheet: {sheet_name}")
                        break
                    else:
                        logger.warning(f"⚠️  No mappings found in sheet: {sheet_name}, trying next sheet...")
        
        workbook.close()
        return mapping
        
    except Exception as e:
        logger.error(f"Error extracting column mapping: {e}", exc_info=True)
        return {}

