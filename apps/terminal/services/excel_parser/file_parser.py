"""Parse complete Excel files."""

import logging
from datetime import date
from typing import Dict
from openpyxl import load_workbook

from apps.terminal.models import TerminalTrip, TerminalRoute
from apps.terminal.services.company_service import get_or_create_company
from apps.terminal.services.trip_service import create_or_update_trip
from .sheet_parser import parse_sheet

logger = logging.getLogger(__name__)


def parse_excel_file(
    file_path: str,
    upload_type: str,
    date_range_start: date,
    date_range_end: date
) -> Dict:
    """
    Parse Excel file and return trips data.
    
    Args:
        file_path: Path to Excel file
        upload_type: 'departures' or 'arrivals'
        date_range_start: Start date of schedule range
        date_range_end: End date of schedule range
    
    Returns:
        Dictionary with:
            - trips: List of trip data dictionaries
            - processed_sheets: List of sheet names processed
            - errors: List of error messages
    """
    all_trips = []
    processed_sheets = []
    all_errors = []
    
    try:
        workbook = load_workbook(file_path, data_only=True)
        
        # Process each sheet
        for sheet_name in workbook.sheetnames:
            # Skip sheets that don't look like date sheets
            # (e.g., "CONTROL SALIDAS" or other metadata sheets)
            if not any(day in sheet_name.lower() for day in ['lunes', 'martes', 'miercoles', 'miércoles', 'jueves', 'viernes', 'sabado', 'sábado', 'domingo']):
                logger.debug(f"Skipping sheet (doesn't look like date sheet): {sheet_name}")
                continue
            
            worksheet = workbook[sheet_name]
            trips, errors = parse_sheet(worksheet, sheet_name, upload_type, date_range_start, date_range_end)
            
            all_trips.extend(trips)
            processed_sheets.append(sheet_name)
            all_errors.extend(errors)
        
        workbook.close()
        
    except Exception as e:
        error_msg = f"Error reading Excel file: {e}"
        logger.error(error_msg, exc_info=True)
        all_errors.append(error_msg)
    
    return {
        'trips': all_trips,
        'processed_sheets': processed_sheets,
        'errors': all_errors
    }


def process_excel_trips(
    file_path: str,
    upload_type: str,
    date_range_start: date,
    date_range_end: date,
    uploaded_by=None
) -> Dict:
    """
    Parse Excel file and create/update trips in database.
    
    Args:
        file_path: Path to Excel file
        upload_type: 'departures' or 'arrivals'
        date_range_start: Start date of schedule range
        date_range_end: End date of schedule range
        uploaded_by: User who uploaded the file (optional)
    
    Returns:
        Dictionary with processing results:
            - trips_created: Number of trips created
            - trips_updated: Number of trips updated
            - errors: List of error messages
    """
    from apps.terminal.models import TerminalRoute
    from django.db import transaction
    
    trips_created = 0
    trips_updated = 0
    errors = []
    created_trips_list = []  # List of created trips for response
    updated_trips_list = []  # List of updated trips for response
    
    # Parse Excel file
    parse_result = parse_excel_file(file_path, upload_type, date_range_start, date_range_end)
    trips_data = parse_result['trips']
    processed_sheets = parse_result['processed_sheets']
    errors.extend(parse_result['errors'])
    
    # Process each trip
    with transaction.atomic():
        for trip_data in trips_data:
            try:
                # Get or create company
                company = get_or_create_company(trip_data['operator'])
                
                # Get or create route
                route, created = TerminalRoute.objects.get_or_create(
                    origin=trip_data['origin'],
                    destination=trip_data['destination'],
                    defaults={
                        'duration': None,
                        'distance': None,
                    }
                )
                
                # Check if trip exists
                lookup = {
                    'company': company,
                    'route': route,
                    'date': trip_data['date'],
                    'trip_type': trip_data['trip_type'],
                }
                
                if trip_data['trip_type'] == 'departure':
                    lookup['departure_time'] = trip_data['departure_time']
                    lookup['arrival_time'] = None
                else:
                    lookup['arrival_time'] = trip_data['arrival_time']
                    lookup['departure_time'] = None
                
                existing_trip = None
                try:
                    existing_trip = TerminalTrip.objects.get(**lookup)
                except TerminalTrip.DoesNotExist:
                    pass
                except TerminalTrip.MultipleObjectsReturned:
                    # If multiple, take the first one
                    existing_trip = TerminalTrip.objects.filter(**lookup).first()
                
                # Prepare trip data for create_or_update
                trip_create_data = {
                    'company': company,
                    'route': route,
                    'trip_type': trip_data['trip_type'],
                    'date': trip_data['date'],
                    'departure_time': trip_data.get('departure_time'),
                    'arrival_time': trip_data.get('arrival_time'),
                    'platform': trip_data.get('platform'),
                    'license_plate': trip_data.get('license_plate'),
                    'observations': trip_data.get('observations'),
                    'price': None,  # Not in Excel v1
                    'currency': 'CLP',
                }
                
                trip = create_or_update_trip(trip_create_data)
                
                # Prepare trip info for response (similar to preview)
                trip_info = {
                    'operator': company.name,
                    'origin': route.origin,
                    'destination': route.destination,
                    'date': trip.date.isoformat(),
                    'tripType': trip.trip_type,
                    'departureTime': trip.departure_time.strftime('%H:%M') if trip.departure_time else None,
                    'arrivalTime': trip.arrival_time.strftime('%H:%M') if trip.arrival_time else None,
                    'platform': trip.platform,
                    'licensePlate': trip.license_plate,
                    'observations': trip.observations,
                    'companyId': str(company.id),
                    'routeId': str(route.id),
                    'tripId': str(trip.id),
                }
                
                if existing_trip:
                    trips_updated += 1
                    updated_trips_list.append(trip_info)
                else:
                    trips_created += 1
                    created_trips_list.append(trip_info)
                
            except Exception as e:
                error_msg = f"Error processing trip: {e}"
                logger.error(error_msg, exc_info=True)
                errors.append(error_msg)
    
        # After processing all trips, extract and create destinations from routes
        try:
            from apps.terminal.services.destination_service import create_or_update_destinations_from_routes
            dest_result = create_or_update_destinations_from_routes()
            logger.info(f"Created/updated {dest_result['created']} destinations, updated {dest_result['updated']} destinations from routes")
        except Exception as e:
            error_msg = f"Error creating destinations from routes: {e}"
            logger.error(error_msg, exc_info=True)
            errors.append(error_msg)
    
    return {
        'trips_created': trips_created,
        'trips_updated': trips_updated,
        'created_trips': created_trips_list,  # List of created trips
        'updated_trips': updated_trips_list,  # List of updated trips
        'errors': errors,
        'processed_sheets': processed_sheets
    }

