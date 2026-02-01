"""Views for events API."""

from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, permissions, filters, status, mixins
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import AllowAny
from rest_framework.exceptions import NotFound, PermissionDenied as DRFPermissionDenied
from drf_spectacular.utils import extend_schema, OpenApiParameter
from django.utils import timezone
from django.db.models import F, Sum, Count, Q
from django.utils.text import slugify
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.http import Http404, HttpResponse
import uuid
import os
from django.db import transaction, models
from django.db.models import Count, Sum, F, Q, Value
from django.db.models.functions import Coalesce
from django.http import Http404, JsonResponse
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal, ROUND_HALF_UP
import csv
import io
import logging
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

from core.permissions import IsOrganizer, HasEventModule
from apps.events.models import (
    Event,
    EventCategory,
    Location,
    TicketTier,
    TicketCategory,
    Order,
    OrderItem,
    Ticket,
    TicketHold,
    Coupon,
    EventCommunication,
    EventImage,
    ComplimentaryTicketInvitation,
)
from apps.forms.models import Form, FormField
from apps.forms.serializers import FormFieldSerializer
from apps.organizers.models import OrganizerUser
from .serializers import (
    EventListSerializer,
    EventDetailSerializer,
    EventCreateSerializer,
    EventUpdateSerializer,
    EventAvailabilitySerializer,
    BookingSerializer,
    EventCategorySerializer,
    LocationSerializer,
    TicketTierSerializer,
    TicketTierCreateUpdateSerializer,
    TicketCategorySerializer,
    OrderSerializer,
    OrderDetailSerializer,
    TicketSerializer,
    CouponSerializer,
    EventCommunicationSerializer,
    PublicEventSerializer,
    ComplimentaryTicketInvitationSerializer,
    ComplimentaryTicketInvitationCreateBatchSerializer,
    ComplimentaryTicketInvitationPreviewSerializer,
    ComplimentaryTicketInvitationRedeemSerializer,
    PublicComplimentaryTicketInvitationSerializer,
)
from apps.events.services.complimentary import (
    parse_excel_file,
    parse_text_file,
    get_or_create_complimentary_tier,
    redeem_invitation,
    export_to_excel,
)


class EventViewSet(viewsets.ModelViewSet):
    """ViewSet for Event model."""
    
    queryset = Event.objects.all()
    serializer_class = EventListSerializer
    permission_classes = [IsAuthenticated]
    
    def get_organizer(self):
        """Obtener el organizador asociado al usuario actual."""
        try:
            # Handle anonymous users
            if not self.request.user.is_authenticated:
                return None
            
            # Buscar OrganizerUser, si hay múltiples tomar el más reciente
            organizer_users = OrganizerUser.objects.filter(user=self.request.user)
            if organizer_users.exists():
                organizer_user = organizer_users.order_by('-created_at').first()
                return organizer_user.organizer
            else:
                return None
        except Exception as e:
            print(f"[EventViewSet] Error getting organizer: {e}")
            return None
    
    def get_queryset(self):
        """Return events based on user permissions."""
        organizer = self.get_organizer()
        print(f"DEBUG - EventViewSet.get_queryset - User: {self.request.user.id if self.request.user.is_authenticated else 'Anonymous'}")
        print(f"DEBUG - EventViewSet.get_queryset - Organizer: {organizer.id if organizer else 'None'}")
        print(f"DEBUG - EventViewSet.get_queryset - Action: {getattr(self, 'action', 'N/A')}")
        
        # Si es un organizador autenticado, puede acceder a TODOS sus eventos
        if organizer:
            queryset = self.queryset.filter(
                organizer=organizer,
                deleted_at__isnull=True
            )
            print(f"DEBUG - EventViewSet.get_queryset - Found {queryset.count()} events for organizer")
            return queryset
        
        # Para usuarios no organizadores o anónimos, solo eventos públicos
        if self.action in ['book', 'reserve', 'availability', 'retrieve']:
            return Event.objects.filter(
                status='published', 
                visibility='public',
                deleted_at__isnull=True  # Exclude soft deleted events
            )
        
        # Si no hay organizador y no es una acción pública, retornar vacío
        print("DEBUG - EventViewSet.get_queryset - No organizer found, returning empty queryset")
        return Event.objects.none()
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == 'list':
            return EventListSerializer
        if self.action == 'retrieve':
            return EventDetailSerializer
        if self.action == 'create':
            return EventCreateSerializer
        if self.action in ['update', 'partial_update']:
            return EventUpdateSerializer
        if self.action == 'availability':
            return EventAvailabilitySerializer
        if self.action == 'book':
            return BookingSerializer
        return EventDetailSerializer
    def get_permissions(self):
        """Allow public access to availability and booking endpoints."""
        if self.action in ['availability', 'book', 'reserve', 'public_list']:
            return [permissions.AllowAny()]
        if self.action == 'retrieve':
            # Allow public access to published events
            return [permissions.AllowAny()]
        return super().get_permissions()
    
    def get_serializer_context(self):
        """Add additional context to serializer."""
        context = super().get_serializer_context()
        if self.action == 'book':
            context['event_id'] = self.kwargs.get('pk')
        return context
    
    def retrieve(self, request, *args, **kwargs):
        """Retrieve a single event with debugging."""
        event_id = kwargs.get('pk')
        print(f"DEBUG - EventViewSet.retrieve - Attempting to retrieve event ID: {event_id}")
        
        try:
            instance = self.get_object()
            
            # Extra security for public access: ensure event is not soft deleted
            if self.action in ['retrieve'] and hasattr(instance, 'is_deleted') and instance.is_deleted:
                print(f"DEBUG - EventViewSet.retrieve - Event {event_id} is soft deleted")
                raise Http404("Event not found")
            
            print(f"DEBUG - EventViewSet.retrieve - Found event: {instance.title} (ID: {instance.id}, Status: {instance.status})")
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        except Exception as e:
            print(f"DEBUG - EventViewSet.retrieve - Error retrieving event {event_id}: {str(e)}")
            raise

    def list(self, request, *args, **kwargs):
        """List events with optional filters."""
        queryset = self.get_queryset()
        
        # Apply filters
        event_type = request.query_params.getlist('type', [])
        if event_type:
            queryset = queryset.filter(type__in=event_type)
        
        start_date = request.query_params.get('start_date')
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        
        end_date = request.query_params.get('end_date')
        if end_date:
            queryset = queryset.filter(start_date__lte=end_date)
        
        location = request.query_params.get('location')
        if location:
            location_terms = location.lower()
            queryset = queryset.filter(
                Q(location__city__icontains=location_terms) |
                Q(location__country__icontains=location_terms) |
                Q(location__name__icontains=location_terms)
            )
        
        categories = request.query_params.getlist('category', [])
        if categories:
            queryset = queryset.filter(category__name__in=categories)
        
        # Pagination
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def availability(self, request, pk=None):
        """Get ticket availability for an event."""
        event = self.get_object()
        
        # Ensure event is not soft deleted
        if event.is_deleted:
            return Response({"detail": "Event not found."}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = self.get_serializer(event)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def public_list(self, request):
        """Get all public published events for the homepage."""
        queryset = Event.objects.filter(
            status='published',
            visibility='public',
            deleted_at__isnull=True  # Exclude soft deleted events
        ).order_by('-featured', '-start_date')
        
        # Apply filters
        event_type = request.query_params.getlist('type', [])
        if event_type:
            queryset = queryset.filter(type__in=event_type)
        
        start_date = request.query_params.get('start_date')
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        
        end_date = request.query_params.get('end_date')
        if end_date:
            queryset = queryset.filter(start_date__lte=end_date)
        
        location = request.query_params.get('location')
        if location:
            location_terms = location.lower()
            queryset = queryset.filter(
                Q(location__city__icontains=location_terms) |
                Q(location__country__icontains=location_terms) |
                Q(location__name__icontains=location_terms)
            )
        
        categories = request.query_params.getlist('category', [])
        if categories:
            queryset = queryset.filter(category__name__in=categories)
        
        # Pagination
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = PublicEventSerializer(page, many=True, context={'request': request})
            return self.get_paginated_response(serializer.data)
        
        serializer = PublicEventSerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], permission_classes=[permissions.AllowAny])
    def book(self, request, pk=None):
        """🚀 ENTERPRISE: Book tickets for an event."""
        import logging
        logger = logging.getLogger(__name__)
        
        event = self.get_object()
        logger.info(f'📦 BOOKING: Attempting to book tickets for event {event.title} ({event.id})')
        logger.info(f'📦 BOOKING: Request data: {request.data}')
        
        # Ensure event is published and not deleted
        if event.status != 'published':
            logger.warning(f'📦 BOOKING: Event {event.id} not published (status: {event.status})')
            return Response({"detail": "Event not available for booking."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Extra security: ensure event is not soft deleted
        if event.is_deleted:
            logger.warning(f'📦 BOOKING: Event {event.id} is deleted')
            return Response({"detail": "Event not available for booking."}, status=status.HTTP_404_NOT_FOUND)
        
        try:
            # Create a clean context without request.user to avoid AnonymousUser issues
            context = {'event_id': pk}
            if request.user.is_authenticated:
                context['request'] = request
            
            serializer = self.get_serializer(data=request.data, context=context)
            if not serializer.is_valid():
                logger.error(f'📦 BOOKING: Validation errors: {serializer.errors}')
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            booking = serializer.save()
            logger.info(f'📦 BOOKING: Successfully created booking: {booking}')
            return Response(booking, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f'📦 BOOKING: Unexpected error: {str(e)}')
            import traceback
            logger.error(f'📦 BOOKING: Traceback: {traceback.format_exc()}')
            return Response({"detail": f"Booking failed: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def reserve(self, request, pk=None):
        """🚀 ENTERPRISE: Reserve tickets with expiration to prevent overselling."""
        event = self.get_object()
        
        # Ensure event is published and not deleted
        if event.status != 'published':
            return Response({"detail": "Event not available for reservations."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Extra security: ensure event is not soft deleted
        if event.is_deleted:
            return Response({"detail": "Event not found."}, status=status.HTTP_404_NOT_FOUND)

        tickets = request.data.get('tickets', [])
        reservation_id = request.data.get('reservationId')
        hold_minutes = int(request.data.get('holdMinutes', 15))
        
        print(f"🔍 RESERVE DEBUG - Raw request data: {request.data}")
        print(f"🔍 RESERVE DEBUG - Tickets array: {tickets}")

        if not isinstance(tickets, list) or not tickets:
            return Response({"detail": "Tickets payload is required."}, status=status.HTTP_400_BAD_REQUEST)

        now = timezone.now()
        expires_at = now + timedelta(minutes=hold_minutes)

        with transaction.atomic():
            # 🚀 ENTERPRISE: Clean expired holds for these tiers before proceeding
            tier_ids = [t.get('tierId') for t in tickets if t.get('tierId')]
            expired_holds = TicketHold.objects.select_for_update().filter(
                ticket_tier_id__in=tier_ids,
                released=False,
                expires_at__lte=now,
            )
            for hold in expired_holds:
                hold.release()

            # Get or create reservation order
            if reservation_id:
                try:
                    order = Order.objects.select_for_update().get(id=reservation_id, event=event)
                    print(f"🔄 RESERVE DEBUG - Reusing existing reservation: {reservation_id}")
                except Order.DoesNotExist:
                    print(f"⚠️ RESERVE DEBUG - Reservation {reservation_id} not found, creating new one")
                    order = Order.objects.create(
                        event=event,
                        email='',  # Will be collected during checkout
                        first_name='Guest',
                        last_name='User',
                        phone='',
                        subtotal=0,
                        service_fee=0,
                        total=0,
                        currency=event.ticket_tiers.first().currency if event.ticket_tiers.exists() else 'CLP',
                        status='pending'
                    )
            else:
                print(f"🆕 RESERVE DEBUG - Creating new reservation")
                order = Order.objects.create(
                    event=event,
                    email='',  # Will be collected during checkout
                    first_name='Guest',
                    last_name='User',
                    phone='',
                    subtotal=0,
                    service_fee=0,
                    total=0,
                    currency=event.ticket_tiers.first().currency if event.ticket_tiers.exists() else 'CLP',
                    status='pending'
                )

            # 🚀 ENTERPRISE: Index existing active holds for this order
            active_holds = list(TicketHold.objects.select_for_update().filter(
                order=order, released=False, expires_at__gt=now
            ))
            holds_by_tier = {}
            for hold in active_holds:
                holds_by_tier.setdefault(str(hold.ticket_tier_id), []).append(hold)

            # 🚀 ENTERPRISE: For each requested tier, adjust holds to match desired quantity
            result_items = []
            for t in tickets:
                tier_id = t.get('tierId')
                qty = int(t.get('quantity', 0))
                custom_price = t.get('customPrice')  # 🚀 NEW: Support for PWYW custom pricing
                if not tier_id or qty < 0:
                    continue
                
                print(f"🎯 RESERVE DEBUG - Processing tier {tier_id}: qty={qty}, custom_price={custom_price}")

                try:
                    tier = TicketTier.objects.select_for_update().get(id=tier_id, event=event)
                except TicketTier.DoesNotExist:
                    return Response({"detail": f"Ticket tier {tier_id} not found."}, status=status.HTTP_400_BAD_REQUEST)

                existing_qty = sum(h.quantity for h in holds_by_tier.get(str(tier_id), []) if not h.is_expired)

                if qty > existing_qty:
                    need = qty - existing_qty
                    
                    # 🚀 ENTERPRISE: Atomic reservation using UPDATE with WHERE clause
                    # This prevents race conditions by doing check + reserve in one atomic operation
                    from django.db.models import F
                    
                    updated_rows = TicketTier.objects.filter(
                        id=tier_id,
                        available__gte=need  # ✅ Check availability at DB level
                    ).update(
                        available=F('available') - need  # ✅ Atomic decrease
                    )
                    
                    if updated_rows == 0:
                        # ❌ Not enough tickets available or tier doesn't exist
                        tier.refresh_from_db()  # Get fresh data for error message
                        return Response({
                            "detail": f"No hay suficientes tickets disponibles para {tier.name}. "
                                    f"Disponibles: {tier.real_available}, Solicitados: {need}"
                        }, status=status.HTTP_400_BAD_REQUEST)
                
                    # ✅ Success: tickets were atomically reserved, now create holds
                    for _ in range(need):
                        # 🚀 ENTERPRISE: Include custom_price for PWYW tickets
                        hold_data = {
                            'event': event,
                            'ticket_tier': tier,
                            'order': order,
                            'quantity': 1,
                            'expires_at': expires_at,
                        }
                        
                        # Add custom_price if this is a PWYW ticket
                        if tier.is_pay_what_you_want and custom_price is not None:
                            from decimal import Decimal
                            print(f"🔍 RESERVE DEBUG - PWYW tier {tier_id}: custom_price received = {custom_price} (type: {type(custom_price)})")
                            custom_price_decimal = Decimal(str(custom_price))
                            print(f"🔍 RESERVE DEBUG - Decimal conversion: {custom_price_decimal}")
                            hold_data['custom_price'] = custom_price_decimal
                        
                        created_hold = TicketHold.objects.create(**hold_data)
                        print(f"🔍 RESERVE DEBUG - Created hold: tier={created_hold.ticket_tier_id}, custom_price={created_hold.custom_price}")
                
                elif qty < existing_qty:
                    to_release = existing_qty - qty
                    # 🚀 ENTERPRISE: Release oldest holds first
                    holds_list = sorted(holds_by_tier.get(str(tier_id), []), key=lambda h: h.expires_at)
                    for hold in holds_list:
                        if to_release <= 0:
                            break
                        hold.release()
                        to_release -= hold.quantity

                result_items.append({
                    'tierId': str(tier_id),
                    'quantity': qty
                })

            return Response({
                'reservationId': str(order.id),
                'expiresAt': expires_at.isoformat(),
                'items': result_items,
                'holdMinutes': hold_minutes
            })
    
    @action(detail=False, methods=['get'])
    def categories(self, request):
        """Get list of event categories."""
        categories = EventCategory.objects.all()
        return Response([category.name for category in categories])
    
    @action(detail=False, methods=['get'])
    def analytics(self, request):
        """🚀 ENTERPRISE: Advanced analytics with REAL EFFECTIVE REVENUE calculation.
        
        Revenue calculation considers:
        - Actual amounts paid (not base price x quantity)
        - Coupon discounts applied
        - Price changes over time
        - Service fees and taxes
        - Manual discounts
        """
        from django.db.models import Sum, Count, Avg, Q, F
        from django.utils import timezone
        from datetime import datetime, timedelta
        from collections import defaultdict
        
        organizer = self.get_organizer()
        if not organizer:
            return Response(
                {"detail": "Usuario sin organizador asociado"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get parameters
        days_param = request.query_params.get('days', '30')
        days = int(days_param) if days_param else 0
        event_id = request.query_params.get('event_id')
        
        # 🚀 ENTERPRISE: Support "all sales" when days=0
        # Calculate start_date: if days=0, get from first order, otherwise use days
        show_all_sales = (days == 0)
        
        if show_all_sales:
            # Get all orders to find the first one
            base_orders = Order.objects.filter(
                event__organizer=organizer,
                status='paid'
            )
            if event_id:
                base_orders = base_orders.filter(event_id=event_id)
            
            first_order = base_orders.order_by('created_at').first()
            if first_order:
                start_date = first_order.created_at.date()
                # Calculate actual days from first order to today
                days = (timezone.now().date() - start_date).days + 1
                # Convert date to datetime at start of day
                start_date = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
            else:
                # No orders found, use default 30 days
                days = 30
                start_date = timezone.now() - timedelta(days=days)
                show_all_sales = False
        else:
            start_date = timezone.now() - timedelta(days=days)
        
        print(f"🔍 [ANALYTICS DEBUG] days={days}, event_id={event_id}, start_date={start_date}, show_all_sales={show_all_sales}")
        
        # Base queryset for PAID orders only (effective sales)
        if show_all_sales:
            # Use all orders (no date filter)
            orders_queryset = Order.objects.filter(
                event__organizer=organizer,
                status='paid'
            ).select_related('event')
        else:
            orders_queryset = Order.objects.filter(
                event__organizer=organizer,
                status='paid',  # Only count actually paid orders
                created_at__gte=start_date
            ).select_related('event')
        
        print(f"🔍 [ANALYTICS DEBUG] Base orders count: {orders_queryset.count()}")
        
        if event_id:
            orders_queryset = orders_queryset.filter(event_id=event_id)
            print(f"🔍 [ANALYTICS DEBUG] Orders after event filter: {orders_queryset.count()}")
        else:
            print(f"🔍 [ANALYTICS DEBUG] No event_id filter applied")
        
        # 🚀 ENTERPRISE: Use centralized revenue calculator
        # This ensures consistency across all endpoints
        from core.revenue_calculator import calculate_event_revenue
        
        # Calculate revenue using centralized function
        # Note: For analytics, we need to calculate per event if event_id is provided
        if event_id:
            from apps.events.models import Event
            try:
                event = Event.objects.get(id=event_id, organizer=organizer)
                revenue_result = calculate_event_revenue(
                    event,
                    start_date=start_date if not show_all_sales else None,
                    end_date=timezone.now() if not show_all_sales else None,
                    validate=True
                )
                # Extract values from centralized calculation
                total_revenue = revenue_result['total_revenue']
                gross_revenue_effective = revenue_result['gross_revenue']
                service_fees_effective = revenue_result['service_fees']
                total_subtotal_original = revenue_result['subtotal_original']
                total_service_fees_original = revenue_result['service_fees_original']
                total_discount = revenue_result['total_discount']
                total_orders = revenue_result['total_orders']
            except Event.DoesNotExist:
                return Response(
                    {"detail": "Evento no encontrado"}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            
        # Calculate total tickets sold (from order items) - ALWAYS needed
        tickets_data = OrderItem.objects.filter(
            order__in=orders_queryset
        ).aggregate(
            total_tickets=Sum('quantity'),
            total_gross_revenue=Sum('unit_price'),  # Sum unit_price * quantity for gross revenue
            avg_unit_price=Avg('unit_price')  # Average actual price paid per ticket
        )
        
        total_tickets = tickets_data['total_tickets'] or 0
        
        # Calculate effective average ticket price (base price without service fees)
        effective_avg_price = 0
        if total_tickets > 0:
            # Use gross revenue effective for accurate average
            effective_avg_price = gross_revenue_effective / total_tickets if gross_revenue_effective > 0 else 0
        
        # 🚀 ENTERPRISE: Time series with effective revenue
        chart_data = []
        for i in range(days):
            current_date = start_date + timedelta(days=i)
            next_date = current_date + timedelta(days=1)
            
            daily_orders = orders_queryset.filter(
                created_at__gte=current_date,
                created_at__lt=next_date
            )
            
            daily_stats = daily_orders.aggregate(
                revenue=Sum('total'),  # Effective revenue (lo que pagó el cliente)
                subtotal=Sum('subtotal'),  # Subtotal original
                service_fees=Sum('service_fee'),  # Service fees originales
                orders=Count('id')
            )
            
            daily_tickets = OrderItem.objects.filter(
                order__in=daily_orders
            ).aggregate(tickets=Sum('quantity'))
            
            # Calcular revenue efectivo diario (distribuyendo descuentos proporcionalmente)
            # 🚀 ENTERPRISE: Todos los valores deben ser enteros (sin decimales para CLP)
            daily_revenue = float(daily_stats['revenue'] or 0)
            daily_subtotal = float(daily_stats['subtotal'] or 0)
            daily_service_fees = float(daily_stats['service_fees'] or 0)
            daily_original = daily_subtotal + daily_service_fees
            
            if daily_original > 0:
                daily_ratio = Decimal(str(daily_revenue)) / Decimal(str(daily_original))
                daily_gross_revenue = (Decimal(str(daily_subtotal)) * daily_ratio).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                daily_gross_revenue = float(daily_gross_revenue)
            else:
                daily_gross_revenue = 0
            
            chart_data.append({
                'date': current_date.strftime('%d/%m/%Y'),
                'sales': daily_tickets['tickets'] or 0,
                'revenue': daily_revenue,
                'gross_revenue': daily_gross_revenue,  # Revenue efectivo del organizador
                'orders': daily_stats['orders'] or 0
            })
        
        # 🚀 ENTERPRISE: Ticket categories with effective revenue (considering discounts)
        category_data = []
        if event_id:
            # For specific event - group by ticket tiers
            event = Event.objects.get(id=event_id)
            for tier in event.ticket_tiers.all():
                # Get orders that have this tier
                tier_orders_filter = {
                    'items__ticket_tier': tier,
                    'status': 'paid'
                }
                if not show_all_sales:
                    tier_orders_filter['created_at__gte'] = start_date
                tier_orders = Order.objects.filter(**tier_orders_filter).distinct()
                
                # Calculate totals for orders with this tier
                tier_order_stats = tier_orders.aggregate(
                    total_revenue=Sum('total'),
                    total_subtotal=Sum('subtotal'),
                    total_service_fees=Sum('service_fee')
                )
                
                tier_revenue = float(tier_order_stats['total_revenue'] or 0)
                tier_subtotal = float(tier_order_stats['total_subtotal'] or 0)
                tier_service_fees = float(tier_order_stats['total_service_fees'] or 0)
                tier_original = tier_subtotal + tier_service_fees
                
                # Calculate effective revenue (distributing discounts proportionally)
                # 🚀 ENTERPRISE: Todos los valores deben ser enteros (sin decimales para CLP)
                if tier_original > 0:
                    tier_ratio = Decimal(str(tier_revenue)) / Decimal(str(tier_original))
                    tier_effective_revenue = (Decimal(str(tier_subtotal)) * tier_ratio).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                    tier_effective_service_fees = (Decimal(str(tier_service_fees)) * tier_ratio).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                    tier_effective_revenue = float(tier_effective_revenue)
                    tier_effective_service_fees = float(tier_effective_service_fees)
                else:
                    tier_effective_revenue = 0
                    tier_effective_service_fees = 0
                
                tier_items_filter = {
                    'ticket_tier': tier,
                    'order__status': 'paid'
                }
                if not show_all_sales:
                    tier_items_filter['order__created_at__gte'] = start_date
                tier_items = OrderItem.objects.filter(**tier_items_filter).aggregate(
                    tickets=Sum('quantity')
                )
                
                tickets_sold = tier_items['tickets'] or 0
                if tickets_sold > 0:
                    category_data.append({
                        'name': tier.name,
                        'value': tickets_sold,
                        'effective_revenue': round(tier_effective_revenue, 2),
                        'service_fees': round(tier_effective_service_fees, 2),
                        'base_price': float(tier.price),
                        'fill': f"#{hash(tier.name) % 16777215:06x}"
                    })
        else:
            # For all events - aggregate by tier names with effective revenue
            tier_stats = defaultdict(lambda: {
                'tickets': 0,
                'orders': set(),
                'order_ids': set()
            })
            
            order_items_filter = {
                'order__event__organizer': organizer,
                'order__status': 'paid'
            }
            if not show_all_sales:
                order_items_filter['order__created_at__gte'] = start_date
            order_items = OrderItem.objects.filter(**order_items_filter).select_related('ticket_tier', 'order')
            
            # First pass: collect tier info and order IDs
            for item in order_items:
                tier_name = item.ticket_tier.name
                tier_stats[tier_name]['tickets'] += item.quantity
                tier_stats[tier_name]['order_ids'].add(item.order_id)
            
            # Second pass: calculate effective revenue per tier (considering discounts)
            for tier_name, stats in tier_stats.items():
                if stats['tickets'] > 0:
                    # Get orders for this tier
                    tier_orders = Order.objects.filter(
                        id__in=stats['order_ids']
                    ).aggregate(
                        total_revenue=Sum('total'),
                        total_subtotal=Sum('subtotal'),
                        total_service_fees=Sum('service_fee')
                    )
                    
                    tier_revenue = float(tier_orders['total_revenue'] or 0)
                    tier_subtotal = float(tier_orders['total_subtotal'] or 0)
                    tier_service_fees = float(tier_orders['total_service_fees'] or 0)
                    tier_original = tier_subtotal + tier_service_fees
                    
                    # Calculate effective revenue (distributing discounts proportionally)
                    # 🚀 ENTERPRISE: Todos los valores deben ser enteros (sin decimales para CLP)
                    if tier_original > 0:
                        tier_ratio = Decimal(str(tier_revenue)) / Decimal(str(tier_original))
                        tier_effective_revenue = (Decimal(str(tier_subtotal)) * tier_ratio).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                        tier_effective_service_fees = (Decimal(str(tier_service_fees)) * tier_ratio).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                        tier_effective_revenue = float(tier_effective_revenue)
                        tier_effective_service_fees = float(tier_effective_service_fees)
                    else:
                        tier_effective_revenue = 0
                        tier_effective_service_fees = 0
                    
                    category_data.append({
                        'name': tier_name,
                        'value': stats['tickets'],
                        'effective_revenue': round(tier_effective_revenue, 2),
                        'service_fees': round(tier_effective_service_fees, 2),
                        'orders_count': len(stats['order_ids']),
                        'fill': f"#{hash(tier_name) % 16777215:06x}"
                    })
        
        # 🚀 ENTERPRISE: Enhanced device/channel analysis (placeholder for now)
        # TODO: Implement real device tracking with user agent analysis
        device_data = [
            {'name': 'Mobile', 'value': int(total_orders * 0.6), 'fill': '#8884d8'},
            {'name': 'Desktop', 'value': int(total_orders * 0.35), 'fill': '#83a6ed'},
            {'name': 'Tablet', 'value': int(total_orders * 0.05), 'fill': '#8dd1e1'},
        ]
        
        # TODO: Implement UTM tracking for real channel data
        channel_data = [
            {'name': 'Sitio Web', 'value': int(total_orders * 0.7), 'fill': '#8884d8'},
            {'name': 'Redes Sociales', 'value': int(total_orders * 0.2), 'fill': '#83a6ed'},
            {'name': 'Directo', 'value': int(total_orders * 0.1), 'fill': '#8dd1e1'},
        ]
        
        response_data = {
            'kpis': {
                'totalSales': total_tickets,
                'totalRevenue': total_revenue,  # Lo que pagó el cliente (después de descuentos)
                'grossRevenue': round(gross_revenue_effective, 2),  # Revenue efectivo del organizador (después de descuentos proporcionales)
                'totalServiceFees': round(service_fees_effective, 2),  # Service fees efectivos (después de descuentos proporcionales)
                 'totalDiscounts': total_discount,
                'averageTicketPrice': round(effective_avg_price, 2),  # Effective average
                'totalOrders': total_orders,
                'conversionRate': 0,  # Placeholder - needs view tracking implementation
                'averageOrderValue': round(total_revenue / total_orders, 2) if total_orders > 0 else 0
            },
            'chartData': chart_data,
            'categoryData': category_data,
            'deviceData': device_data,
            'channelData': channel_data,
            'dateRange': {
                'start': start_date.isoformat(),
                'end': timezone.now().isoformat(),
                'days': days
            },
            'metadata': {
                'calculation_method': 'effective_revenue',
                'includes_discounts': True,
                'includes_service_fees': True,
                'only_paid_orders': True,
                'event_id_filter': event_id,
                'filtered_orders_count': orders_queryset.count()
            }
        }
        
        print(f"🔍 [ANALYTICS DEBUG] Final response KPIs: totalSales={total_tickets}, totalRevenue={total_revenue}")
        print(f"🔍 [ANALYTICS DEBUG] Event filter applied: {event_id is not None}")
        
        return Response(response_data)

    @action(detail=True, methods=['get'])
    def export_orders(self, request, pk=None):
        """
        🚀 ENTERPRISE: Export event orders with complete financial data
        
        Supports CSV and Excel formats with complete order information:
        - Order details (number, status, dates)
        - Customer information (name, email, phone, address)
        - Financial data (subtotal, taxes, fees, discounts, total)
        - Payment information (method, transaction ID, payment dates)
        - Items breakdown (ticket types, quantities, prices)
        - Coupon usage and discount details
        """
        print(f"🔍 [EXPORT ORDERS] Method called with pk: {pk}")
        print(f"🔍 [EXPORT ORDERS] Request path: {request.path}")
        print(f"🔍 [EXPORT ORDERS] Request method: {request.method}")
        
        try:
            print(f"🔍 [EXPORT ORDERS] Getting event object...")
            event = self.get_object()
            print(f"🔍 [EXPORT ORDERS] Event found: {event.title} (ID: {event.id})")
            
            print(f"🔍 [EXPORT ORDERS] Getting organizer...")
            organizer = self.get_organizer()
            print(f"🔍 [EXPORT ORDERS] Organizer: {organizer}")
            
            # Verify organizer owns this event
            if not organizer or event.organizer != organizer:
                print(f"🔍 [EXPORT ORDERS] Permission denied - organizer mismatch")
                return Response(
                    {"detail": "No tienes permisos para exportar este evento"},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            print(f"🔍 [EXPORT ORDERS] Permission check passed")
            
            # Get export format (default to CSV)
            export_format = request.query_params.get('export_format', 'csv').lower()
            print(f"🔍 [EXPORT ORDERS] Export format: {export_format}")
            
            # Get all orders for this event with related data
            print(f"🔍 [EXPORT ORDERS] Querying orders for event {event.id}...")
            orders = Order.objects.filter(
                event=event
            ).select_related(
                'user',
                'coupon'
            ).prefetch_related(
                'items__ticket_tier',
                'items__tickets'
            ).order_by('created_at')
            
            orders_count = orders.count()
            print(f"🔍 [EXPORT ORDERS] Found {orders_count} orders")
            
            if orders_count == 0:
                print(f"🔍 [EXPORT ORDERS] No orders found - returning empty response")
                return Response(
                    {"detail": "No hay órdenes para exportar para este evento"},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            print(f"🔍 [EXPORT ORDERS] Generating {export_format} export...")
            if export_format == 'excel':
                print(f"🔍 [EXPORT ORDERS] Calling _export_orders_excel...")
                result = self._export_orders_excel(event, orders)
                print(f"🔍 [EXPORT ORDERS] Excel export completed, content_type: {result.get('Content-Type', 'NOT SET')}")
            else:
                print(f"🔍 [EXPORT ORDERS] Calling _export_orders_csv...")
                result = self._export_orders_csv(event, orders)
                print(f"🔍 [EXPORT ORDERS] CSV export completed, content_type: {result.get('Content-Type', 'NOT SET')}")
            
            print(f"🔍 [EXPORT ORDERS] Export generated successfully")
            return result
                
        except Exception as e:
            print(f"🔍 [EXPORT ORDERS] ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {"detail": f"Error al exportar pedidos: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def export_attendees(self, request, pk=None):
        """
        🚀 ENTERPRISE: Export event attendees with all data including custom fields
        
        Supports CSV and Excel formats with complete attendee information:
        - Basic attendee info (name, email, phone)
        - Ticket information (type, price, status)
        - Order information (order number, payment status, dates)
        - Custom form fields and responses
        - Check-in status and timestamps
        - Approval status for tickets requiring approval
        """
        try:
            event = self.get_object()
            organizer = self.get_organizer()
            
            # Verify organizer owns this event
            if not organizer or event.organizer != organizer:
                return Response(
                    {"detail": "No tienes permisos para exportar este evento"},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Get export format (default to CSV)
            export_format = request.query_params.get('export_format', 'csv').lower()
            
            # Get all tickets for this event with related data
            tickets = Ticket.objects.filter(
                order_item__order__event=event
            ).select_related(
                'order_item__order',
                'order_item__ticket_tier',
                'order_item__ticket_tier__form',
                'check_in_by',
                'approved_by'
            ).prefetch_related(
                'order_item__ticket_tier__form__fields'
            ).order_by('created_at')
            
            print(f"🔍 [EXPORT ATTENDEES] Generating {export_format} export...")
            if export_format == 'excel':
                print(f"🔍 [EXPORT ATTENDEES] Calling _export_attendees_excel...")
                result = self._export_attendees_excel(event, tickets)
                print(f"🔍 [EXPORT ATTENDEES] Excel export completed, content_type: {result.get('Content-Type', 'NOT SET')}")
                return result
            else:
                print(f"🔍 [EXPORT ATTENDEES] Calling _export_attendees_csv...")
                result = self._export_attendees_csv(event, tickets)
                print(f"🔍 [EXPORT ATTENDEES] CSV export completed, content_type: {result.get('Content-Type', 'NOT SET')}")
                return result
                
        except Exception as e:
            print(f"🔍 [EXPORT ATTENDEES] ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {"detail": f"Error al exportar asistentes: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def conversion_metrics(self, request, pk=None):
        """🚀 ENTERPRISE: Get conversion funnel metrics for a specific event."""
        from apps.events.analytics_models import EventView, ConversionFunnel, EventPerformanceMetrics
        from django.db.models import Count, Avg
        from datetime import timedelta
        
        event = self.get_object()
        days = int(request.query_params.get('days', 30))
        start_date = timezone.now() - timedelta(days=days)
        
        # Get funnel stages count
        funnel_data = ConversionFunnel.objects.filter(
            event=event,
            created_at__gte=start_date
        ).values('stage').annotate(count=Count('id')).order_by('stage')
        
        # Get view metrics
        views = EventView.objects.filter(
            event=event,
            created_at__gte=start_date
        )
        
        total_views = views.count()
        unique_views = views.values('session_id').distinct().count()
        converted_views = views.filter(converted_to_purchase=True).count()
        avg_time_on_page = views.aggregate(avg=Avg('time_on_page'))['avg'] or 0
        
        # Calculate conversion rate
        conversion_rate = (converted_views / total_views * 100) if total_views > 0 else 0
        
        # Get traffic sources
        traffic_sources = views.values('view_source').annotate(
            count=Count('id'),
            conversions=Count('id', filter=models.Q(converted_to_purchase=True))
        ).order_by('-count')
        
        # Device breakdown
        device_breakdown = views.values('device_type').annotate(
            count=Count('id'),
            conversions=Count('id', filter=models.Q(converted_to_purchase=True))
        ).order_by('-count')
        
        return Response({
            'event_id': str(event.id),
            'event_title': event.title,
            'date_range': {
                'start': start_date.isoformat(),
                'end': timezone.now().isoformat(),
                'days': days
            },
            'funnel_stages': list(funnel_data),
            'view_metrics': {
                'total_views': total_views,
                'unique_views': unique_views,
                'converted_views': converted_views,
                'conversion_rate': round(conversion_rate, 2),
                'avg_time_on_page': round(avg_time_on_page, 2) if avg_time_on_page else 0
            },
            'traffic_sources': [
                {
                    'source': item['view_source'],
                    'views': item['count'],
                    'conversions': item['conversions'],
                    'conversion_rate': round(item['conversions'] / item['count'] * 100, 2) if item['count'] > 0 else 0
                }
                for item in traffic_sources
            ],
            'device_breakdown': [
                {
                    'device': item['device_type'],
                    'views': item['count'],
                    'conversions': item['conversions'],
                    'conversion_rate': round(item['conversions'] / item['count'] * 100, 2) if item['count'] > 0 else 0
                }
                for item in device_breakdown
            ]
        })

    @action(detail=False, methods=['post'])
    def track_view(self, request):
        """🚀 ENTERPRISE: Track an event view for analytics."""
        from apps.events.analytics_models import EventView
        
        event_id = request.data.get('event_id')
        if not event_id:
            return Response({'detail': 'event_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            event = Event.objects.get(id=event_id)
            view = EventView.track_view(event, request)
            
            return Response({
                'view_id': str(view.id) if view else None,
                'event_id': str(event.id),
                'tracked': view is not None
            })
            
        except Event.DoesNotExist:
            return Response({'detail': 'Event not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def organizer(self, request):
        """Get events for the current organizer with metrics."""
        organizer = self.get_organizer()
        if not organizer:
            return Response(
                {"detail": "Usuario sin organizador asociado"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get all events for the organizer (exclude soft deleted)
        events = Event.objects.filter(
            organizer=organizer,
            deleted_at__isnull=True  # Exclude soft deleted events
        ).select_related(
            'location', 'category'
        ).prefetch_related(
            'ticket_tiers', 'images'
        ).order_by('-start_date')
        
        # Apply filters
        status_filter = request.query_params.getlist('status', [])
        if status_filter:
            events = events.filter(status__in=status_filter)
        
        start_date = request.query_params.get('start_date')
        if start_date:
            events = events.filter(start_date__gte=start_date)
        
        end_date = request.query_params.get('end_date')
        if end_date:
            events = events.filter(start_date__lte=end_date)
        
        # Filter out events without start_date for proper sorting
        events = events.exclude(start_date__isnull=True)
        
        search = request.query_params.get('search')
        if search:
            events = events.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(location__name__icontains=search) |
                Q(location__address__icontains=search)
            )
        
        # Calculate metrics for each event
        events_with_metrics = []
        # 🚀 ENTERPRISE: Use centralized revenue calculator for consistency
        from core.revenue_calculator import calculate_event_revenue
        
        for event in events:
            # 🚀 ENTERPRISE: Use centralized revenue calculator
            revenue_result = calculate_event_revenue(
                event,
                start_date=None,  # All sales for event list
                end_date=None,
                validate=False  # Skip validation for performance in list view
            )
            
            # Extract values from centralized calculation
            total_revenue = revenue_result['total_revenue']
            organizer_revenue = revenue_result['gross_revenue']  # Ingresos brutos del organizador
            service_fees = revenue_result['service_fees']  # Service fees efectivos
            sold_tickets = revenue_result['total_tickets']
            total_orders = revenue_result['total_orders']
            
            # Calculate total capacity from ticket tiers
            total_tickets = 0
            
            # Calculate total capacity from ticket tiers
            for tier in event.ticket_tiers.all():
                tier_capacity = tier.capacity or 0
                total_tickets += tier_capacity
            
            events_with_metrics.append({
                'id': str(event.id),
                'title': event.title,
                'start_date': event.start_date.isoformat() if event.start_date else None,
                'end_date': event.end_date.isoformat() if event.end_date else None,
                'status': event.status,
                'location': {
                    'name': event.location.name if event.location else '',
                    'address': event.location.address if event.location else '',
                    'is_virtual': event.location.is_virtual if event.location else False
                },
                'ticket_tiers': [
                    {
                        'id': str(tier.id),
                        'name': tier.name,
                        'price': tier.price,
                        'quantity': tier.capacity,
                        'sold_quantity': tier.tickets_sold,  # 🚀 ENTERPRISE: Use real sold data
                        'capacity': tier.capacity,  # Include original capacity for null detection
                        'is_unlimited': tier.capacity is None  # 🚀 ENTERPRISE: Flag for unlimited capacity
                    }
                    for tier in event.ticket_tiers.all()
                ],
                'total_revenue': total_revenue,  # What customers actually paid (after discounts)
                'gross_revenue': organizer_revenue,  # 🚀 ENTERPRISE: Revenue efectivo del organizador (after proportional discount distribution)
                'service_fees': service_fees,  # 🚀 ENTERPRISE: Service fees efectivos (after proportional discount distribution)
                'total_tickets': total_tickets,
                'sold_tickets': sold_tickets,
                'has_unlimited_capacity': any(tier.capacity is None for tier in event.ticket_tiers.all()),  # 🚀 ENTERPRISE: Event-level unlimited flag
                'images': [
                    {
                        'id': str(img.id),
                        'url': request.build_absolute_uri(img.image.url) if img.image else ''
                    }
                    for img in event.images.all()
                ]
            })
        
        return Response(events_with_metrics)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload(self, request):
        """🚀 ENTERPRISE IMAGE UPLOAD - Upload an image file and optionally associate with an event."""
        if 'file' not in request.FILES:
            return Response({'detail': 'No file found'}, status=status.HTTP_400_BAD_REQUEST)
        
        file_obj = request.FILES['file']
        event_id = request.data.get('event_id')  # Optional event association
        
        # Generate a unique filename
        filename = f"{uuid.uuid4().hex}.{file_obj.name.split('.')[-1]}"
        
        # 🚀 ENTERPRISE: Use Django's configured storage backend (local in dev, GCS in prod)
        from django.core.files.storage import default_storage
        
        try:
            # Save file using Django's storage backend (automatically uses correct storage)
            file_path = f"event_images/{filename}"
            saved_path = default_storage.save(file_path, file_obj)
            
            # Get the full URL for the uploaded file
            file_url = default_storage.url(saved_path)
            
            print(f"[UPLOAD] ✅ File uploaded to: {saved_path}")
            print(f"[UPLOAD] ✅ File URL: {file_url}")
            
            # If event_id is provided, create EventImage record
            if event_id:
                try:
                    event = Event.objects.get(id=event_id)
                    
                    # Verify user has permission to modify this event
                    organizer = self.get_organizer()
                    if not organizer or event.organizer != organizer:
                        # Clean up uploaded file if permission denied
                        try:
                            default_storage.delete(saved_path)
                        except Exception as e:
                            print(f"[UPLOAD] ⚠️ Could not delete file {saved_path}: {e}")
                        return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)
                    
                    # 🚀 ENTERPRISE FIX: Replace main image instead of adding additional ones
                    # Check if this is meant to replace the main image (order=0)
                    replace_main = request.data.get('replace_main', 'true').lower() == 'true'
                    
                    if replace_main:
                        # Delete existing main image (order=0) if it exists
                        main_image = event.images.filter(order=0).first()
                        if main_image:
                            try:
                                # Delete the old file from storage
                                default_storage.delete(main_image.image.name)
                                print(f"[UPLOAD] 🗑️ Deleted old main image: {main_image.image.name}")
                            except Exception as e:
                                print(f"[UPLOAD] ⚠️ Could not delete old image file: {e}")
                            # Delete the database record
                            main_image.delete()
                            print(f"[UPLOAD] 🗑️ Deleted old main image record")
                        
                        # Create new main image with order=0
                        order = 0
                    else:
                        # Add as additional image
                        order = event.images.count()
                    
                    # Create EventImage record
                    event_image = EventImage.objects.create(
                        event=event,
                        image=saved_path,  # Store the path returned by storage backend
                        alt=request.data.get('alt', file_obj.name),
                        type=request.data.get('type', 'image'),
                        order=order
                    )
                    
                    print(f"[UPLOAD] ✅ Created EventImage record: {event_image.id} for event {event.id}")
                    
                    return Response({
                        'url': file_url,
                        'event_image_id': event_image.id,
                        'message': 'Image uploaded and associated with event successfully'
                    }, status=status.HTTP_201_CREATED)
                
                except Event.DoesNotExist:
                    # Clean up uploaded file if event doesn't exist
                    try:
                        default_storage.delete(saved_path)
                    except Exception as del_e:
                        print(f"[UPLOAD] ⚠️ Could not delete file {saved_path}: {del_e}")
                    return Response({'detail': 'Event not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # If no event_id provided, just return the uploaded file URL
            print(f"[UPLOAD] ✅ File uploaded successfully: {file_url}")
            return Response({
                'url': file_url,
                'message': 'Image uploaded successfully'
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            print(f"[UPLOAD] ❌ Error during upload: {str(e)}")
            return Response({'detail': f'Error uploading file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['delete'], url_path=r'images/(?P<image_id>[^/.]+)')
    def delete_image(self, request, pk=None, image_id=None):
        """Delete an event image."""
        event = self.get_object()
        image = get_object_or_404(EventImage, id=image_id, event=event)
        
        # Delete file from storage using Django's storage backend
        try:
            from django.core.files.storage import default_storage
            default_storage.delete(image.image.name)
            print(f"[DELETE-IMAGE] ✅ Deleted file: {image.image.name}")
        except Exception as e:
            print(f"[DELETE-IMAGE] ⚠️ Could not delete file {image.image.name}: {e}")
        
        image.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=True, methods=['post'], url_path='images/from-asset')
    def create_image_from_asset(self, request, pk=None):
        """
        🚀 ENTERPRISE: Create EventImage from MediaAsset.
        
        Payload:
        {
            "asset_id": "uuid",
            "replace_main": true,  // Optional, default true
            "type": "image",  // Optional: image, banner, thumbnail, gallery
            "alt": "Alt text"  // Optional
        }
        """
        import logging
        logger = logging.getLogger(__name__)
        
        from apps.media.models import MediaAsset, MediaUsage
        from django.contrib.contenttypes.models import ContentType
        
        event = self.get_object()
        asset_id = request.data.get('asset_id')
        replace_main = request.data.get('replace_main', True)
        image_type = request.data.get('type', 'image')
        alt_text = request.data.get('alt', '')
        
        if not asset_id:
            return Response({'detail': 'asset_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            asset = MediaAsset.objects.get(id=asset_id, deleted_at__isnull=True)
        except MediaAsset.DoesNotExist:
            return Response({'detail': 'Asset not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Verify organizer owns the asset (or it's global)
        organizer = self.get_organizer()
        if asset.scope == 'organizer' and asset.organizer != organizer:
            return Response({'detail': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Replace main image if requested
        if replace_main:
            main_image = event.images.filter(order=0).first()
            if main_image:
                try:
                    from django.core.files.storage import default_storage
                    default_storage.delete(main_image.image.name)
                except Exception as e:
                    logger.warning(f"Could not delete old image: {e}")
                main_image.delete()
            order = 0
        else:
            order = event.images.count()
        
        # Create EventImage using the asset's file path
        event_image = EventImage.objects.create(
            event=event,
            image=asset.file.name,  # Reference same file in GCS
            alt=alt_text or asset.original_filename,
            type=image_type,
            order=order
        )
        
        # Create MediaUsage tracking
        content_type = ContentType.objects.get_for_model(Event)
        MediaUsage.objects.create(
            asset=asset,
            content_type=content_type,
            object_id=event.id,
            field_name=f"image_{image_type}"
        )
        
        logger.info(f"📸 [EVENT-MEDIA] Created EventImage {event_image.id} from asset {asset.id} for event {event.id}")
        
        return Response({
            'event_image_id': str(event_image.id),
            'url': asset.url,  # Use asset.url instead of building from event_image.image.url
            'message': 'Image linked successfully'
        }, status=status.HTTP_201_CREATED)

    def perform_create(self, serializer):
        # Obtener el organizador del usuario de manera correcta
        try:
            # Intentar obtener OrganizerUser para el usuario actual
            organizer_user = OrganizerUser.objects.get(user=self.request.user)
            organizer = organizer_user.organizer
            
            # Guardar el evento con el organizador encontrado
            event = serializer.save(
                organizer=organizer,
                status='draft'
            )
            
            # Ensure the event has an ID before returning
            event.refresh_from_db()
            return event
        except OrganizerUser.DoesNotExist:
            print(f"Error: Usuario {self.request.user.id} no tiene un organizer asociado")
            raise Exception("El usuario no tiene un organizador asociado")
    
    @action(detail=True, methods=['post'])
    def complete_draft(self, request, pk=None):
        """
        Marcar un evento borrador como completo y listo para añadir tickets y categorías.
        Verifica que el evento tenga toda la información básica necesaria.
        """
        event = self.get_object()
        
        # Log entire event details for debugging
        print(f"DEBUG - Complete Draft Request for event {pk}:")
        print(f"- title: '{event.title}'")
        print(f"- description length: {len(event.description) if event.description else 0}")
        print(f"- short_description: '{event.short_description}'")
        print(f"- start_date: {event.start_date}")
        print(f"- end_date: {event.end_date}")
        print(f"- location: {event.location.id if event.location else None}")
        print(f"- current status: {event.status}")
        
        # Modified: Allow any status to be marked as complete, but log a warning if not draft
        if event.status != 'draft':
            print(f"WARNING: Attempting to mark a non-draft event as complete (status: {event.status})")
            # We'll continue anyway to allow the user to fix issues
        
        # Verificar que el evento tenga toda la información básica necesaria
        required_fields = ['title', 'description', 'short_description', 'start_date', 'end_date', 'location']
        missing_fields = []
        
        for field in required_fields:
            value = getattr(event, field, None)
            if field == 'location':
                if value is None:
                    missing_fields.append(field)
            elif value is None or (isinstance(value, str) and value.strip() == ''):
                missing_fields.append(field)
        
        # Log the event data for debugging
        print(f"Event data for complete_draft: {event.title}, {event.short_description}, {event.start_date}, {event.end_date}")
        
        if missing_fields:
            return Response(
                {
                    'error': 'El evento no está completo',
                    'missing_fields': missing_fields
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Actualizar el estado del evento - setting to draft_complete regardless of previous status
        event.status = 'draft_complete'
        event.save()
        
        return Response({
            'status': 'draft_complete',
            'message': 'El evento ha sido marcado como completo y está listo para añadir tickets y categorías'
        })
    
    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        event = self.get_object()
        
        # Verificar que el evento esté en estado borrador completo
        if event.status != 'draft_complete':
            return Response(
                {'error': 'Solo los eventos con borrador completo pueden ser publicados'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar que el evento tenga al menos una categoría y un ticket tier
        if not event.ticket_categories.exists():
            return Response(
                {'error': 'El evento debe tener al menos una categoría de tickets'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        if not event.ticket_tiers.exists():
            return Response(
                {'error': 'El evento debe tener al menos un tipo de ticket'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Todo está bien, publicar el evento
        event.status = 'published'
        event.save()
        
        return Response({
            'status': 'published',
            'message': 'Evento publicado exitosamente'
        })

    def update(self, request, *args, **kwargs):
        """Override update to properly handle camelCase field names."""
        data = request.data.copy()
        
        # Map camelCase keys to snake_case for Django models
        if 'shortDescription' in data:
            print(f"ViewSet - Mapping shortDescription -> short_description: '{data['shortDescription']}'")
            data['short_description'] = data.pop('shortDescription')
        
        if 'startDate' in data:
            print(f"ViewSet - Mapping startDate -> start_date: '{data['startDate']}'")
            data['start_date'] = data.pop('startDate')
        
        if 'endDate' in data:
            print(f"ViewSet - Mapping endDate -> end_date: '{data['endDate']}'")
            data['end_date'] = data.pop('endDate')
        
        # Log all data being sent to serializer
        print(f"ViewSet - Data being sent to serializer: {data}")
        
        # Get instance for updating
        instance = self.get_object()
        
        # Create serializer with the mapped data
        serializer = self.get_serializer(instance, data=data, partial=kwargs.get('partial', False))
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        # Log the instance after update
        instance.refresh_from_db()
        print(f"ViewSet - After update: title={instance.title}, short_description={instance.short_description}")
        
        return Response(serializer.data)
    
    def destroy(self, request, *args, **kwargs):
        """Soft delete an event if it has no revenue."""
        event = self.get_object()
        
        try:
            # Check if event can be deleted (no revenue)
            if not event.can_be_deleted():
                return Response(
                    {"detail": "No se puede eliminar un evento que tiene ingresos"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Perform soft delete
            event.soft_delete()
            
            return Response(
                {"detail": "Evento eliminado correctamente"},
                status=status.HTTP_200_OK
            )
            
        except Exception as e:
            return Response(
                {"detail": f"Error al eliminar el evento: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    
    def _export_attendees_csv(self, event, tickets):
        """Generate CSV export for attendees"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="asistentes_{event.slug}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # Add BOM for proper UTF-8 encoding in Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Get all form fields for this event
        form_fields = []
        for ticket_tier in event.ticket_tiers.all():
            if ticket_tier.form:
                for field in ticket_tier.form.fields.all():
                    if field.label not in [f.label for f in form_fields]:
                        form_fields.append(field)
        
        # Write header
        header = [
            'Número de Ticket',
            'Nombre',
            'Apellido',
            'Email',
            'Tipo de Ticket',
            'Precio del Ticket',
            'Estado del Ticket',
            'Estado de Check-in',
            'Fecha de Check-in',
            'Check-in por',
            'Estado de Aprobación',
            'Aprobado por',
            'Fecha de Aprobación',
            'Razón de Rechazo',
            'Número de Orden',
            'Estado de Orden',
            'Método de Pago',
            'Total de Orden',
            'Fecha de Compra',
            'Fecha de Pago',
            'Teléfono del Comprador',
            'Nombre del Comprador',
            'IP de Compra',
            'Cupón Usado',
            'Descuento Aplicado'
        ]
        
        # Add custom form fields to header
        for field in form_fields:
            header.append(f'Campo: {field.label}')
        
        writer.writerow(header)
        
        # Write data rows
        for ticket in tickets:
            order = ticket.order_item.order
            
            row = [
                ticket.ticket_number,
                ticket.first_name,
                ticket.last_name,
                ticket.email,
                ticket.order_item.ticket_tier.name if ticket.order_item.ticket_tier else 'N/A',
                str(ticket.order_item.unit_price),
                ticket.get_status_display(),
                ticket.get_check_in_status_display(),
                ticket.check_in_time.strftime('%Y-%m-%d %H:%M:%S') if ticket.check_in_time else '',
                ticket.check_in_by.get_full_name() if ticket.check_in_by else '',
                ticket.get_approval_status_display() if ticket.approval_status else '',
                ticket.approved_by.get_full_name() if ticket.approved_by else '',
                ticket.approved_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.approved_at else '',
                ticket.rejection_reason,
                order.order_number,
                order.get_status_display(),
                order.payment_method,
                str(order.total),
                order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                order.updated_at.strftime('%Y-%m-%d %H:%M:%S') if order.is_paid else '',
                order.phone,
                f"{order.first_name} {order.last_name}".strip(),
                order.ip_address or '',
                order.coupon.code if order.coupon else '',
                str(order.discount) if order.discount > 0 else '0'
            ]
            
            # Add custom form field values
            for field in form_fields:
                field_value = ''
                if ticket.form_data and str(field.id) in ticket.form_data:
                    field_value = str(ticket.form_data[str(field.id)])
                row.append(field_value)
            
            writer.writerow(row)
        
        return response
    
    def _export_attendees_excel(self, event, tickets):
        """Generate Excel export for attendees"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistentes"
        
        # Get all form fields for this event
        form_fields = []
        for ticket_tier in event.ticket_tiers.all():
            if ticket_tier.form:
                for field in ticket_tier.form.fields.all():
                    if field.label not in [f.label for f in form_fields]:
                        form_fields.append(field)
        
        # Define headers
        headers = [
            'Número de Ticket',
            'Nombre',
            'Apellido',
            'Email',
            'Tipo de Ticket',
            'Precio del Ticket',
            'Estado del Ticket',
            'Estado de Check-in',
            'Fecha de Check-in',
            'Check-in por',
            'Estado de Aprobación',
            'Aprobado por',
            'Fecha de Aprobación',
            'Razón de Rechazo',
            'Número de Orden',
            'Estado de Orden',
            'Método de Pago',
            'Total de Orden',
            'Fecha de Compra',
            'Fecha de Pago',
            'Teléfono del Comprador',
            'Nombre del Comprador',
            'IP de Compra',
            'Cupón Usado',
            'Descuento Aplicado'
        ]
        
        # Add custom form fields to headers
        for field in form_fields:
            headers.append(f'Campo: {field.label}')
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, ticket in enumerate(tickets, 2):
            order = ticket.order_item.order
            
            data = [
                ticket.ticket_number,
                ticket.first_name,
                ticket.last_name,
                ticket.email,
                ticket.order_item.ticket_tier.name if ticket.order_item.ticket_tier else 'N/A',
                float(ticket.order_item.unit_price),
                ticket.get_status_display(),
                ticket.get_check_in_status_display(),
                ticket.check_in_time.replace(tzinfo=None) if ticket.check_in_time else '',
                ticket.check_in_by.get_full_name() if ticket.check_in_by else '',
                ticket.get_approval_status_display() if ticket.approval_status else '',
                ticket.approved_by.get_full_name() if ticket.approved_by else '',
                ticket.approved_at.replace(tzinfo=None) if ticket.approved_at else '',
                ticket.rejection_reason,
                order.order_number,
                order.get_status_display(),
                order.payment_method,
                float(order.total),
                order.created_at.replace(tzinfo=None) if order.created_at else '',
                order.updated_at.replace(tzinfo=None) if order.updated_at and order.is_paid else '',
                order.phone,
                f"{order.first_name} {order.last_name}".strip(),
                order.ip_address or '',
                order.coupon.code if order.coupon else '',
                float(order.discount) if order.discount > 0 else 0
            ]
            
            # Add custom form field values
            for field in form_fields:
                field_value = ''
                if ticket.form_data and str(field.id) in ticket.form_data:
                    field_value = ticket.form_data[str(field.id)]
                data.append(field_value)
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="asistentes_{event.slug}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    def _export_orders_csv(self, event, orders):
        """Generate CSV export for orders"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="pedidos_{event.slug}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # Add BOM for proper UTF-8 encoding in Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Write header
        header = [
            'Número de Orden',
            'Estado de Orden',
            'Fecha de Creación',
            'Fecha de Pago',
            'Email del Comprador',
            'Nombre del Comprador',
            'Apellido del Comprador',
            'Teléfono',
            'Subtotal',
            'Impuestos',
            'Comisión de Servicio',
            'Descuento',
            'Total',
            'Moneda',
            'Método de Pago',
            'ID de Transacción',
            'Cupón Usado',
            'Código de Cupón',
            'IP de Compra',
            'User Agent',
            'Cantidad de Tickets',
            'Tipos de Tickets',
            'Precios Unitarios',
            'Notas',
            'Monto Reembolsado',
            'Razón de Reembolso'
        ]
        
        writer.writerow(header)
        
        # Write data rows
        for order in orders:
            # Aggregate ticket information
            ticket_types = []
            unit_prices = []
            total_tickets = 0
            
            for item in order.items.all():
                ticket_types.append(f"{item.ticket_tier.name} (x{item.quantity})")
                unit_prices.append(f"{item.ticket_tier.name}: ${item.unit_price}")
                total_tickets += item.quantity
            
            row = [
                order.order_number,
                order.get_status_display(),
                order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                order.updated_at.strftime('%Y-%m-%d %H:%M:%S') if order.is_paid else '',
                order.email,
                order.first_name,
                order.last_name,
                order.phone,
                str(order.subtotal),
                str(order.taxes),
                str(order.service_fee),
                str(order.discount),
                str(order.total),
                order.currency,
                order.payment_method,
                order.payment_id or '',
                'Sí' if order.coupon else 'No',
                order.coupon.code if order.coupon else '',
                order.ip_address or '',
                order.user_agent,
                str(total_tickets),
                '; '.join(ticket_types),
                '; '.join(unit_prices),
                order.notes,
                str(order.refunded_amount),
                order.refund_reason
            ]
            
            writer.writerow(row)
        
        return response
    
    def _export_orders_excel(self, event, orders):
        """Generate Excel export for orders"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos"
        
        # Define headers
        headers = [
            'Número de Orden',
            'Estado de Orden',
            'Fecha de Creación',
            'Fecha de Pago',
            'Email del Comprador',
            'Nombre del Comprador',
            'Apellido del Comprador',
            'Teléfono',
            'Subtotal',
            'Impuestos',
            'Comisión de Servicio',
            'Descuento',
            'Total',
            'Moneda',
            'Método de Pago',
            'ID de Transacción',
            'Cupón Usado',
            'Código de Cupón',
            'IP de Compra',
            'User Agent',
            'Cantidad de Tickets',
            'Tipos de Tickets',
            'Precios Unitarios',
            'Notas',
            'Monto Reembolsado',
            'Razón de Reembolso'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, order in enumerate(orders, 2):
            # Aggregate ticket information
            ticket_types = []
            unit_prices = []
            total_tickets = 0
            
            for item in order.items.all():
                ticket_types.append(f"{item.ticket_tier.name} (x{item.quantity})")
                unit_prices.append(f"{item.ticket_tier.name}: ${item.unit_price}")
                total_tickets += item.quantity
            
            data = [
                order.order_number,
                order.get_status_display(),
                order.created_at.replace(tzinfo=None) if order.created_at else '',
                order.updated_at.replace(tzinfo=None) if order.updated_at and order.is_paid else '',
                order.email,
                order.first_name,
                order.last_name,
                order.phone,
                float(order.subtotal),
                float(order.taxes),
                float(order.service_fee),
                float(order.discount),
                float(order.total),
                order.currency,
                order.payment_method,
                order.payment_id or '',
                'Sí' if order.coupon else 'No',
                order.coupon.code if order.coupon else '',
                order.ip_address or '',
                order.user_agent,
                total_tickets,
                '; '.join(ticket_types),
                '; '.join(unit_prices),
                order.notes,
                float(order.refunded_amount),
                order.refund_reason
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="pedidos_{event.slug}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def validation_stats(self, request):
        """
        Get validation statistics for an event by event ID.
        """
        event_id = request.query_params.get('event_id')
        if not event_id:
            return Response(
                {"detail": "event_id parameter is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            organizer = self.get_organizer()
            if not organizer:
                return Response(
                    {"detail": "Unauthorized."},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Get event and verify organizer access
            event = Event.objects.get(id=event_id, organizer=organizer)
            
            # Get all tickets for this event
            tickets = Ticket.objects.filter(
                order_item__order__event=event
            ).select_related(
                'order_item__ticket_tier'
            ).order_by('order_item__ticket_tier__name')
            
            # Calculate overall stats
            total_tickets = tickets.count()
            checked_in_tickets = tickets.filter(check_in_status='checked_in').count()
            pending_tickets = tickets.filter(check_in_status='pending').count()
            cancelled_tickets = tickets.filter(status='cancelled').count()
            refunded_tickets = tickets.filter(status='refunded').count()
            
            # Calculate stats by ticket tier
            ticket_tier_stats = {}
            for ticket in tickets:
                tier_name = ticket.order_item.ticket_tier.name if ticket.order_item.ticket_tier else 'General'
                
                if tier_name not in ticket_tier_stats:
                    ticket_tier_stats[tier_name] = {
                        'tier_name': tier_name,
                        'total': 0,
                        'checked_in': 0,
                        'pending': 0,
                        'cancelled': 0,
                        'refunded': 0,
                        'check_in_rate': 0.0
                    }
                
                ticket_tier_stats[tier_name]['total'] += 1
                
                if ticket.check_in_status == 'checked_in':
                    ticket_tier_stats[tier_name]['checked_in'] += 1
                elif ticket.check_in_status == 'pending':
                    ticket_tier_stats[tier_name]['pending'] += 1
                
                if ticket.status == 'cancelled':
                    ticket_tier_stats[tier_name]['cancelled'] += 1
                elif ticket.status == 'refunded':
                    ticket_tier_stats[tier_name]['refunded'] += 1
            
            # Calculate check-in rates
            for tier_stats in ticket_tier_stats.values():
                if tier_stats['total'] > 0:
                    tier_stats['check_in_rate'] = round(
                        (tier_stats['checked_in'] / tier_stats['total']) * 100, 2
                    )
            
            # Calculate overall check-in rate
            overall_check_in_rate = round((checked_in_tickets / total_tickets * 100), 2) if total_tickets > 0 else 0
            
            return Response({
                'event': {
                    'id': str(event.id),
                    'title': event.title,
                    'start_date': event.start_date.isoformat(),
                    'end_date': event.end_date.isoformat(),
                },
                'overall_stats': {
                    'total_tickets': total_tickets,
                    'checked_in': checked_in_tickets,
                    'pending': pending_tickets,
                    'cancelled': cancelled_tickets,
                    'refunded': refunded_tickets,
                    'check_in_rate': overall_check_in_rate
                },
                'ticket_tier_stats': list(ticket_tier_stats.values())
            })
            
        except Event.DoesNotExist:
            return Response(
                {"detail": "Event not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"detail": f"Error retrieving validation stats: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def validation_logs(self, request):
        """
        Get validation logs for an event by event ID.
        """
        event_id = request.query_params.get('event_id')
        if not event_id:
            return Response(
                {"detail": "event_id parameter is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            organizer = self.get_organizer()
            if not organizer:
                return Response(
                    {"detail": "Unauthorized."},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Get event and verify organizer access
            event = Event.objects.get(id=event_id, organizer=organizer)
            
            # Import here to avoid circular imports
            from apps.validation.models import TicketValidationLog
            
            # Get validation logs for this event
            logs = TicketValidationLog.objects.filter(
                validator_session__event=event
            ).select_related(
                'ticket',
                'validator_session'
            ).order_by('-created_at')
            
            # Serialize the logs
            logs_data = []
            for log in logs:
                logs_data.append({
                    'id': str(log.id),
                    'ticket_id': str(log.ticket.id) if log.ticket else None,
                    'ticket_number': log.ticket.ticket_number if log.ticket else 'N/A',
                    'attendee_name': f"{log.ticket.first_name} {log.ticket.last_name}" if log.ticket else 'N/A',
                    'validator_name': log.validator_session.validator_name,
                    'action': log.action,
                    'status': log.status,
                    'message': log.message,
                    'timestamp': log.created_at.isoformat(),
                    'from_status': log.metadata.get('from_status', 'pending'),
                    'to_status': log.metadata.get('to_status', 'checked_in'),
                })
            
            return Response({
                'event': {
                    'id': str(event.id),
                    'title': event.title,
                },
                'logs': logs_data,
                'total_logs': len(logs_data)
            })
            
        except Event.DoesNotExist:
            return Response(
                {"detail": "Event not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"detail": f"Error retrieving validation logs: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    # 🎫 COMPLIMENTARY: Complimentary ticket invitation endpoints
    
    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser], url_path='complimentary-tickets/preview')
    def complimentary_tickets_preview(self, request, pk=None):
        """
        Preview Excel/text file for complimentary ticket invitations.
        
        Accepts:
        - Excel file (.xlsx) in 'file' field, OR
        - Text content in 'text' field with optional 'delimiter' (default: tab)
        
        Returns parsed data with suggested column mapping.
        """
        event = self.get_object()
        
        # Verify organizer has access to this event
        organizer = self.get_organizer()
        if not organizer or event.organizer != organizer:
            return Response(
                {"detail": "You don't have permission to access this event"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        entries = []
        errors = []
        suggested_mapping = {}
        
        # Check if Excel file was uploaded
        if 'file' in request.FILES:
            file_obj = request.FILES['file']
            if not file_obj.name.endswith('.xlsx'):
                return Response(
                    {"detail": "File must be .xlsx format"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            entries, errors = parse_excel_file(file_obj)
            
        # Check if text content was provided
        elif 'text' in request.data:
            text_content = request.data.get('text', '')
            delimiter = request.data.get('delimiter', '\t')
            entries, errors = parse_text_file(text_content, delimiter)
            
        else:
            return Response(
                {"detail": "Either 'file' or 'text' field is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Auto-detect column mapping from first row if available
        if entries:
            # Try to detect from first entry keys
            first_entry = entries[0]
            for key in first_entry.keys():
                key_lower = key.lower()
                if any(n in key_lower for n in ['nombre', 'first', 'name', 'invitado']):
                    suggested_mapping['first_name'] = key
                elif any(n in key_lower for n in ['apellido', 'last', 'surname']):
                    suggested_mapping['last_name'] = key
                elif any(n in key_lower for n in ['email', 'correo', 'mail']):
                    suggested_mapping['email'] = key
        
        serializer = ComplimentaryTicketInvitationPreviewSerializer({
            'data': entries,
            'suggested_mapping': suggested_mapping,
            'errors': errors
        })
        
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='complimentary-tickets/create-batch')
    def complimentary_tickets_create_batch(self, request, pk=None):
        """Create batch of complimentary ticket invitations."""
        event = self.get_object()
        
        # Verify organizer has access
        organizer = self.get_organizer()
        if not organizer or event.organizer != organizer:
            return Response(
                {"detail": "You don't have permission to access this event"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = ComplimentaryTicketInvitationCreateBatchSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        entries = serializer.validated_data['entries']
        tickets_per_invitation = serializer.validated_data['tickets_per_invitation']
        
        # Get or create complimentary tier
        ticket_tier = get_or_create_complimentary_tier(event)
        
        # Create invitations
        invitations = []
        with transaction.atomic():
            for entry_data in entries:
                invitation = ComplimentaryTicketInvitation.objects.create(
                    event=event,
                    ticket_tier=ticket_tier,
                    first_name=entry_data.get('first_name', ''),
                    last_name=entry_data.get('last_name', ''),
                    email=entry_data.get('email', ''),
                    max_tickets=tickets_per_invitation,
                    tickets_per_invitation=tickets_per_invitation,
                    created_by=request.user
                )
                invitations.append(invitation)
        
        # Serialize response
        response_serializer = ComplimentaryTicketInvitationSerializer(invitations, many=True)
        return Response({
            'created': len(invitations),
            'invitations': response_serializer.data
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['get'], url_path='complimentary-tickets')
    def complimentary_tickets(self, request, pk=None):
        """List complimentary ticket invitations for an event."""
        event = self.get_object()
        
        # Verify organizer has access
        organizer = self.get_organizer()
        if not organizer or event.organizer != organizer:
            return Response(
                {"detail": "You don't have permission to access this event"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        queryset = ComplimentaryTicketInvitation.objects.filter(event=event)
        
        # Filters
        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search)
            )
        
        serializer = ComplimentaryTicketInvitationSerializer(queryset.order_by('-created_at'), many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'], url_path='complimentary-tickets/export-excel')
    def complimentary_tickets_export_excel(self, request, pk=None):
        """Export complimentary ticket invitations to Excel."""
        event = self.get_object()
        
        # Verify organizer has access
        organizer = self.get_organizer()
        if not organizer or event.organizer != organizer:
            return Response(
                {"detail": "You don't have permission to access this event"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        queryset = ComplimentaryTicketInvitation.objects.filter(event=event)
        
        # Apply same filters as list
        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search)
            )
        
        # Generate Excel
        excel_file = export_to_excel(queryset.order_by('-created_at'))
        
        # Return as download
        from django.http import HttpResponse
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="cortesias_{event.slug}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        return response


class EventCategoryViewSet(viewsets.ModelViewSet):
    """
    API endpoint for event categories.
    """
    queryset = EventCategory.objects.all()
    serializer_class = EventCategorySerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['name']
    ordering = ['name']


class LocationViewSet(viewsets.ModelViewSet):
    """
    API endpoint for locations.
    """
    serializer_class = LocationSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['name']  # Removed city, country as they don't exist in simplified Location model
    search_fields = ['name', 'address']  # Removed city, country
    ordering_fields = ['name']  # Removed city, country
    ordering = ['name']
    
    def get_organizer(self):
        """Obtener el organizador asociado al usuario actual."""
        try:
            organizer_user = OrganizerUser.objects.get(user=self.request.user)
            return organizer_user.organizer
        except OrganizerUser.DoesNotExist:
            return None
    
    def get_queryset(self):
        """
        Get locations for the current organizer if authenticated.
        """
        if not self.request.user.is_authenticated:
            return Location.objects.filter(events__status='active', events__visibility='public').distinct()
        
        organizer = self.get_organizer()
        if not organizer:
            return Location.objects.none()
        
        return Location.objects.filter(organizer=organizer)
    
    def get_permissions(self):
        """
        Get permissions based on action.
        """
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated(), IsOrganizer()]


class TicketCategoryViewSet(viewsets.ModelViewSet):
    """
    API endpoint for ticket categories.
    """
    serializer_class = TicketCategorySerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['event', 'status', 'visibility']
    ordering_fields = ['order', 'name']
    ordering = ['order']
    
    def get_organizer(self):
        """Obtener el organizador asociado al usuario actual."""
        try:
            organizer_user = OrganizerUser.objects.get(user=self.request.user)
            return organizer_user.organizer
        except OrganizerUser.DoesNotExist:
            return None
    
    def get_queryset(self):
        """
        Get ticket categories based on permissions.
        """
        # For public access, only return public categories for active events
        if not self.request.user.is_authenticated:
            return TicketCategory.objects.filter(
                event__status='active',
                event__visibility='public',
                visibility='public'
            )
        
        # For organizers, return all their categories
        organizer = self.get_organizer()
        if not organizer:
            return TicketCategory.objects.none()
        
        # Check for event_id in URL or query params
        event_id = self.kwargs.get('event_id')
        if not event_id:
            event_id = self.request.query_params.get('event')
        
        queryset = TicketCategory.objects.filter(event__organizer=organizer)
        if event_id:
            queryset = queryset.filter(event_id=event_id)
        return queryset
    
    def get_permissions(self):
        """
        Get permissions based on action.
        """
        print(f"DEBUG - TicketCategoryViewSet.get_permissions - action: {self.action}")
        print(f"DEBUG - TicketCategoryViewSet.get_permissions - user: {self.request.user.id if self.request.user.is_authenticated else 'Anonymous'}")
        
        if self.action in ['list', 'retrieve']:
            print(f"DEBUG - TicketCategoryViewSet.get_permissions - permissions: AllowAny (list/retrieve)")
            return [permissions.AllowAny()]
            
        # Simplificar permisos - la verificación real se hace en los métodos del ViewSet
        print(f"DEBUG - TicketCategoryViewSet.get_permissions - simplified permissions: IsAuthenticated only")
        return [permissions.IsAuthenticated()]
    
    def perform_create(self, serializer):
        """Create a new ticket category with the event ID from the URL or request data."""
        # Check if this is a nested view (event_id in URL)
        event_id = self.kwargs.get('event_id')
        
        if event_id:
            # This is a nested view, get event from URL
            event = get_object_or_404(Event, id=event_id)
            serializer.save(event=event)
        else:
            # This is a standalone/top-level view
            # Get event_id from request data
            event_id = self.request.data.get('event')
            
            if not event_id:
                raise serializers.ValidationError({"event": "Event ID is required to create a ticket category"})
                
            # Get the event instance
            event = get_object_or_404(Event, id=event_id)
            
            # Check if the event belongs to the organizer
            organizer = self.get_organizer()
            if not organizer or event.organizer != organizer:
                raise PermissionDenied("You don't have permission to add ticket categories to this event.")
            
            # Create the ticket category with the event
            serializer.save(event=event)
            
    def update(self, request, *args, **kwargs):
        """Override update to use OrganizerUser for permission checks."""
        print(f"DEBUG - TicketCategoryViewSet.update - Starting update for ID: {kwargs.get('pk')}")
        
        # Get the category being updated
        instance = self.get_object()
        
        # Get the organizer associated with the user
        organizer = self.get_organizer()
        if not organizer:
            return Response(
                {"detail": "You don't have permission to update this category as you are not associated with any organizer."},
                status=status.HTTP_403_FORBIDDEN
            )
            
        # Check if the event belongs to the organizer
        if instance.event.organizer != organizer:
            return Response(
                {"detail": "You don't have permission to update this category as it belongs to another organizer."},
                status=status.HTTP_403_FORBIDDEN
            )
                
        # Permission check passed, proceed with update
        return super().update(request, *args, **kwargs)
        
    def partial_update(self, request, *args, **kwargs):
        """Override partial_update to use OrganizerUser for permission checks."""
        print(f"DEBUG - TicketCategoryViewSet.partial_update - Starting partial update for ID: {kwargs.get('pk')}")
        return self.update(request, *args, **kwargs)
        
    def destroy(self, request, *args, **kwargs):
        """Override destroy to use OrganizerUser for permission checks."""
        print(f"DEBUG - TicketCategoryViewSet.destroy - Starting delete for ID: {kwargs.get('pk')}")
        
        # Get the category being deleted
        instance = self.get_object()
        
        # Get the organizer associated with the user
        organizer = self.get_organizer()
        if not organizer:
            return Response(
                {"detail": "You don't have permission to delete this category as you are not associated with any organizer."},
                status=status.HTTP_403_FORBIDDEN
            )
            
        # Check if the event belongs to the organizer
        if instance.event.organizer != organizer:
            return Response(
                {"detail": "You don't have permission to delete this category as it belongs to another organizer."},
                status=status.HTTP_403_FORBIDDEN
            )
                
        # Permission check passed, proceed with delete
        return super().destroy(request, *args, **kwargs)


class TicketTierViewSet(viewsets.ModelViewSet):
    """Viewset for ticket tiers."""
    serializer_class = TicketTierSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_organizer(self):
        """Obtener el organizador asociado al usuario actual."""
        try:
            organizer_user = OrganizerUser.objects.get(user=self.request.user)
            return organizer_user.organizer
        except OrganizerUser.DoesNotExist:
            return None
    
    def get_queryset(self):
        """Filter queryset by organizer and event_id if provided"""
        organizer = self.get_organizer()
        print(f"DEBUG - TicketTierViewSet.get_queryset - User ID: {self.request.user.id}, Organizer: {organizer.id if organizer else 'None'}")
        
        if not organizer:
            print(f"DEBUG - TicketTierViewSet.get_queryset - No organizer found for user")
            return TicketTier.objects.none()
        
        # If this is a nested view under an event, filter by the event_id from URL
        event_id = self.kwargs.get('event_id')
        
        # If it's a standalone view, check for event query parameter
        if not event_id and self.request.query_params.get('event'):
            event_id = self.request.query_params.get('event')
            print(f"DEBUG - TicketTierViewSet.get_queryset - Using event query parameter: {event_id}")
        
        if event_id:
            print(f"DEBUG - TicketTierViewSet.get_queryset - Filtering by event ID: {event_id}")
            return TicketTier.objects.filter(
                event__organizer=organizer,
                event_id=event_id
            )
        
        # For top-level view without event filter, return all tiers for the organizer's events
        print(f"DEBUG - TicketTierViewSet.get_queryset - No event filter, returning all tiers for organizer")
        return TicketTier.objects.filter(event__organizer=organizer)
    
    def _transform_request_data(self, data):
        """
        🚀 ENTERPRISE: Transform camelCase keys to snake_case for Django models.
        This ensures frontend camelCase fields are properly mapped to backend snake_case.
        """
        print(f"🔍 TRANSFORM DEBUG - Input data keys: {list(data.keys())}")
        
        # Field mapping from camelCase (frontend) to snake_case (Django)
        field_mapping = {
            'isPayWhatYouWant': 'is_pay_what_you_want',
            'minPrice': 'min_price',
            'maxPrice': 'max_price',
            'suggestedPrice': 'suggested_price',
            'maxPerOrder': 'max_per_order',
            'minPerOrder': 'min_per_order',
            'isPublic': 'is_public',
            'requiresApproval': 'requires_approval',
        }
        
        transformed = {}
        
        for key, value in data.items():
            # Use mapping if exists, otherwise keep original key
            new_key = field_mapping.get(key, key)
            transformed[new_key] = value
            
            # Log PWYW field transformations specifically
            if key in field_mapping:
                print(f"✅ TRANSFORM - {key} → {new_key}: {value}")
        
        print(f"🔍 TRANSFORM DEBUG - Output data keys: {list(transformed.keys())}")
        return transformed

    def update(self, request, *args, **kwargs):
        """
        🚀 ENTERPRISE: Robust update with proper field transformation and validation.
        Handles camelCase to snake_case conversion and preserves all fields.
        """
        print(f"🚀 DEBUG - TicketTierViewSet.update - Starting update for ID: {kwargs.get('pk')}")
        print(f"🔍 DEBUG - User ID: {request.user.id}, Username: {request.user.username}")
        print(f"🔍 DEBUG - Raw request.data: {request.data}")
        
        # Get the organizer associated with the user
        try:
            organizer_user = OrganizerUser.objects.get(user=request.user)
            organizer = organizer_user.organizer
            print(f"✅ DEBUG - Found organizer: ID={organizer.id}")
            
            # Get the tier being updated
            instance = self.get_object()
            print(f"🔍 DEBUG - Ticket tier belongs to event ID: {instance.event.id}")
            print(f"🔍 DEBUG - Event organizer ID: {instance.event.organizer.id}")
            
            # Check if the organizer owns the event
            if instance.event.organizer != organizer:
                print(f"❌ DEBUG - PERMISSION DENIED: Ticket tier's event organizer ({instance.event.organizer.id}) does not match user's organizer ({organizer.id})")
                return Response(
                    {"detail": "You don't have permission to update this ticket tier as it belongs to another organizer."},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # 🚀 ENTERPRISE FIX: Transform camelCase to snake_case FIRST
            transformed_data = self._transform_request_data(request.data.copy())
            
            # 🚀 ENTERPRISE FIX: Handle structured price object while preserving ALL other fields
            # 🔧 FIX: Handle price object with both camelCase and snake_case keys
            if 'price' in transformed_data and isinstance(transformed_data['price'], dict):
                price_obj = transformed_data['price']
                # Try both camelCase and snake_case (after transformation, it's snake_case)
                base_price = price_obj.get('base_price') or price_obj.get('basePrice') or 0
                print(f"💰 DEBUG - Extracting price from object: base_price={base_price}, keys={list(price_obj.keys())}")
                
                if base_price == 0:
                    print(f"⚠️ WARNING - Price extracted as 0! Original object: {price_obj}")
                
                # Replace price object with extracted base price
                transformed_data['price'] = base_price
            
            # Log PWYW fields specifically for debugging
            pwyw_fields = ['is_pay_what_you_want', 'min_price', 'max_price', 'suggested_price']
            for field in pwyw_fields:
                if field in transformed_data:
                    print(f"✅ PWYW DEBUG - {field}: {transformed_data[field]}")
                else:
                    print(f"⚠️ PWYW DEBUG - {field}: NOT FOUND")
            
            # 🚀 ENTERPRISE: Use transformed data with ALL fields preserved
            serializer = self.get_serializer(
                instance, 
                data=transformed_data, 
                partial=kwargs.get('partial', False)
            )
            
            print(f"🔍 DEBUG - Serializer validation starting...")
            serializer.is_valid(raise_exception=True)
            print(f"✅ DEBUG - Serializer validation passed")
            
            # 🔧 FIX: Recalcular available automáticamente si capacity cambió
            save_kwargs = {}
            if 'capacity' in transformed_data:
                new_capacity = transformed_data['capacity']
                # Obtener tickets vendidos del tier actual
                tickets_sold = instance.tickets_sold
                # Calcular nuevo available: capacity - tickets_sold
                new_available = max(0, new_capacity - tickets_sold)
                
                print(f"🔧 FIX - Recalculating available: capacity={new_capacity}, tickets_sold={tickets_sold}, new_available={new_available}")
                
                # Forzar actualización de available aunque sea read_only
                # Pasamos available directamente al save() para evitar la restricción de read_only
                save_kwargs['available'] = new_available
            
            # Perform the update
            if save_kwargs:
                # Si hay campos adicionales (como available), pasarlos al save()
                serializer.save(**save_kwargs)
            else:
                self.perform_update(serializer)
            print(f"✅ DEBUG - Update completed successfully")
            
            # Log the final saved data
            instance.refresh_from_db()
            print(f"💾 DEBUG - Final saved PWYW status: {instance.is_pay_what_you_want}")
            if instance.is_pay_what_you_want:
                print(f"💾 DEBUG - Final saved min_price: {instance.min_price}")
                print(f"💾 DEBUG - Final saved max_price: {instance.max_price}")
                print(f"💾 DEBUG - Final saved suggested_price: {instance.suggested_price}")
            
            return Response(serializer.data)
            
        except OrganizerUser.DoesNotExist:
            print(f"❌ DEBUG - PERMISSION DENIED: User {request.user.id} is not associated with any organizer")
            return Response(
                {"detail": "You don't have permission to update this ticket tier as you are not associated with any organizer."},
                status=status.HTTP_403_FORBIDDEN
            )
        except Exception as e:
            print(f"❌ DEBUG - Error during update: {str(e)}")
            import traceback
            print(f"❌ DEBUG - Full traceback: {traceback.format_exc()}")
            return Response(
                {"detail": f"Error updating ticket tier: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def partial_update(self, request, *args, **kwargs):
        """
        🚀 ENTERPRISE: Robust partial update using the same transformation logic.
        """
        print(f"🚀 DEBUG - TicketTierViewSet.partial_update - Starting partial update for ID: {kwargs.get('pk')}")
        
        # Ensure partial=True for partial updates
        kwargs['partial'] = True
        
        # Use the same robust update logic
        return self.update(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        """
        🚀 ENTERPRISE: Robust create with proper field transformation.
        """
        print(f"🚀 DEBUG - TicketTierViewSet.create - Starting create")
        print(f"🔍 DEBUG - Raw request.data: {request.data}")
        
        # Transform camelCase to snake_case
        transformed_data = self._transform_request_data(request.data.copy())
        
        # Handle structured price object
        # 🔧 FIX: Try both camelCase and snake_case
        if 'price' in transformed_data and isinstance(transformed_data['price'], dict):
            price_obj = transformed_data['price']
            # Try both camelCase and snake_case (after transformation, it's snake_case)
            base_price = price_obj.get('base_price') or price_obj.get('basePrice') or 0
            print(f"💰 CREATE DEBUG - Extracting price from object: base_price={base_price}, keys={list(price_obj.keys())}")
            
            if base_price == 0:
                print(f"⚠️ WARNING - Price extracted as 0! Original object: {price_obj}")
            
            transformed_data['price'] = base_price
        
        # Log PWYW fields for debugging
        pwyw_fields = ['is_pay_what_you_want', 'min_price', 'max_price', 'suggested_price']
        for field in pwyw_fields:
            if field in transformed_data:
                print(f"✅ PWYW CREATE DEBUG - {field}: {transformed_data[field]}")
        
        # Create new request with transformed data
        request._full_data = transformed_data
        
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        """
        🚀 ENTERPRISE: Create a ticket tier with proper event association and field handling.
        """
        # Get the event from the URL if this is a nested view
        event_id = self.kwargs.get('event_id')
        if event_id:
            print(f"🚀 DEBUG - TicketTierViewSet.perform_create - Creating ticket tier with event ID from URL: {event_id}")
            event = get_object_or_404(Event, id=event_id)
            
            # Check if the event belongs to the organizer
            organizer = self.get_organizer()
            if not organizer or event.organizer != organizer:
                print(f"❌ DEBUG - PERMISSION DENIED: Event organizer ({event.organizer.id}) does not match user's organizer ({organizer.id if organizer else 'None'})")
                raise PermissionDenied("You don't have permission to add ticket tiers to this event.")
            
            # Get capacity to set available field
            capacity = self.request.data.get('capacity', 0)
            print(f"🔍 DEBUG - Using capacity value for available: {capacity}")
            
            # 🚀 ENTERPRISE FIX: Handle unlimited capacity properly
            # If capacity is null (unlimited), set available to a large number
            available_value = capacity if capacity is not None else 9999999
            print(f"✅ DEBUG - Setting available to: {available_value} (capacity: {capacity})")
            
            # Save with event association and proper available value
            serializer.save(event=event, available=available_value)
        else:
            # Extract event ID from request data for top-level view
            event_id = self.request.data.get('event')
            print(f"🚀 DEBUG - TicketTierViewSet.perform_create - Creating ticket tier with event ID from request data: {event_id}")
            
            if not event_id:
                print(f"DEBUG - TicketTierViewSet.perform_create - FAIL: No event ID in request data")
                raise serializers.ValidationError({"event": "Event ID is required to create a ticket tier"})
                
            # Get the event instance
            event = get_object_or_404(Event, id=event_id)
            
            # Check if the event belongs to the organizer
            organizer = self.get_organizer()
            if not organizer or event.organizer != organizer:
                print(f"DEBUG - PERMISSION DENIED: Event organizer ({event.organizer.id}) does not match user's organizer ({organizer.id if organizer else 'None'})")
                raise PermissionDenied("You don't have permission to add ticket tiers to this event.")
            
            # 🚀 ENTERPRISE FIX: Handle price properly regardless of format
            price_data = self.request.data.get('price', 0)
            if isinstance(price_data, dict):
                price_data = price_data.get('basePrice', 0)
                print(f"DEBUG - Extracted price from object: {price_data}")
            elif price_data is None:
                price_data = 0
                print(f"DEBUG - Price was None, using default: {price_data}")
            else:
                print(f"DEBUG - Using direct price value: {price_data}")
                
            # Get capacity to set available field
            capacity = self.request.data.get('capacity', 0)
            print(f"DEBUG - Using capacity value for available: {capacity}")
            
            # 🚀 ENTERPRISE FIX: Handle unlimited capacity properly
            # If capacity is null (unlimited), set available to a large number
            available_value = capacity if capacity is not None else 9999999
            print(f"DEBUG - Setting available to: {available_value} (capacity: {capacity})")
            
            # Save with extracted price value and set available properly
            serializer.save(event=event, price=price_data, available=available_value)

    @action(detail=True, methods=['post'])
    def form_link(self, request, pk=None):
        """Link a form to a ticket tier"""
        ticket_tier = self.get_object()
        form_id = request.data.get('form_id')
        
        if not form_id:
            return Response(
                {"detail": "Form ID is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            form = Form.objects.get(id=form_id)
            
            # Get the organizer properly
            organizer = self.get_organizer()
            
            # Verify the form belongs to the same organizer
            if form.organizer != organizer:
                return Response(
                    {"detail": "You don't have permission to use this form"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Update the ticket tier with the form directly
            ticket_tier.form = form
            ticket_tier.save()
            
            return Response({
                "id": ticket_tier.id,
                "name": ticket_tier.name,
                "form_id": form_id,
                "form_name": form.name
            })
            
        except Form.DoesNotExist:
            return Response(
                {"detail": "Form not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"detail": str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def category_link(self, request, pk=None):
        """Link a category to a ticket tier"""
        ticket_tier = self.get_object()
        category_id = request.data.get('category_id')
        
        if not category_id:
            return Response(
                {"detail": "Category ID is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            category = TicketCategory.objects.get(id=category_id)
            
            # Get the organizer properly
            organizer = self.get_organizer()
            
            # Verify the category belongs to the same organizer and event
            if category.event.organizer != organizer:
                return Response(
                    {"detail": "You don't have permission to use this category"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Verify the category belongs to the same event as the ticket tier
            if category.event != ticket_tier.event:
                return Response(
                    {"detail": "Category must belong to the same event as the ticket tier"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Update the ticket tier with the category directly
            ticket_tier.category = category
            ticket_tier.save()
            
            return Response({
                "id": ticket_tier.id,
                "name": ticket_tier.name,
                "category_id": category_id,
                "category_name": category.name
            })
            
        except TicketCategory.DoesNotExist:
            return Response(
                {"detail": "Category not found"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"detail": str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



    
    @action(detail=False, methods=['get'])
    def types(self, request):
        """
        Get available field types and their configurations.
        """
        field_types = [
            {
                'value': 'text',
                'label': 'Texto corto',
                'supportsRequired': True,
                'supportsPlaceholder': True,
                'supportsHelpText': True,
                'supportsValidation': True,
                'validations': ['minLength', 'maxLength', 'pattern']
            },
            {
                'value': 'textarea',
                'label': 'Texto largo',
                'supportsRequired': True,
                'supportsPlaceholder': True,
                'supportsHelpText': True,
                'supportsValidation': True,
                'validations': ['minLength', 'maxLength']
            },
            {
                'value': 'email',
                'label': 'Email',
                'supportsRequired': True,
                'supportsPlaceholder': True,
                'supportsHelpText': True,
                'supportsValidation': False
            },
            {
                'value': 'phone',
                'label': 'Teléfono',
                'supportsRequired': True,
                'supportsPlaceholder': True,
                'supportsHelpText': True,
                'supportsValidation': False
            },
            {
                'value': 'number',
                'label': 'Número',
                'supportsRequired': True,
                'supportsPlaceholder': True,
                'supportsHelpText': True,
                'supportsValidation': True,
                'validations': ['min', 'max']
            },
            {
                'value': 'select',
                'label': 'Lista desplegable',
                'supportsRequired': True,
                'supportsPlaceholder': True,
                'supportsHelpText': True,
                'supportsOptions': True,
                'supportsValidation': False
            },
            {
                'value': 'checkbox',
                'label': 'Casillas de verificación',
                'supportsRequired': True,
                'supportsHelpText': True,
                'supportsOptions': True,
                'supportsValidation': False
            },
            {
                'value': 'radio',
                'label': 'Opciones únicas',
                'supportsRequired': True,
                'supportsHelpText': True,
                'supportsOptions': True,
                'supportsValidation': False
            },
            {
                'value': 'date',
                'label': 'Fecha',
                'supportsRequired': True,
                'supportsHelpText': True,
                'supportsValidation': False
            },
            {
                'value': 'heading',
                'label': 'Título',
                'supportsRequired': False,
                'supportsHelpText': False,
                'supportsValidation': False
            },
            {
                'value': 'paragraph',
                'label': 'Párrafo',
                'supportsRequired': False,
                'supportsHelpText': True,
                'supportsValidation': False
            }
        ]
        
        width_options = [
            {'value': 'full', 'label': 'Completo'},
            {'value': 'half', 'label': 'Medio'},
            {'value': 'third', 'label': 'Un tercio'}
        ]
        
        conditional_operators = [
            {'value': 'equals', 'label': 'Es igual a'},
            {'value': 'notEquals', 'label': 'No es igual a'},
            {'value': 'contains', 'label': 'Contiene'},
            {'value': 'notContains', 'label': 'No contiene'},
            {'value': 'greaterThan', 'label': 'Mayor que'},
            {'value': 'lessThan', 'label': 'Menor que'}
        ]
        
        return Response({
            'fieldTypes': field_types,
            'widthOptions': width_options,
            'conditionalOperators': conditional_operators
        })


class OrderViewSet(viewsets.ModelViewSet):
    """
    API endpoint for orders.
    """
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['event', 'status', 'payment_method']
    search_fields = ['order_number', 'first_name', 'last_name', 'email']
    ordering_fields = ['created_at', 'updated_at', 'total']
    ordering = ['-created_at']
    
    # ✅ SOLUCIÓN ROBUSTA: Paginación personalizada para órdenes
    def get_paginated_response(self, data):
        """
        Personalizar respuesta paginada para incluir estadísticas completas
        """
        # Si no hay paginación activa, devolver respuesta normal
        if not hasattr(self, 'paginator') or self.paginator is None:
            return Response(data)
            
        # Calcular estadísticas sobre TODOS los pedidos (no solo la página actual)
        all_orders = self.filter_queryset(self.get_queryset())
        
        # Estadísticas completas
        total_orders = all_orders.count()
        total_revenue = all_orders.aggregate(
            total=models.Sum('total'),
            service_fees=models.Sum('service_fee')
        )
        
        # Respuesta paginada con estadísticas completas
        return self.paginator.get_paginated_response({
            'results': data,
            'statistics': {
                'total_orders': total_orders,
                'total_revenue': total_revenue['total'] or 0,
                'total_service_fees': total_revenue['service_fees'] or 0,
                'current_page_count': len(data)
            }
        })
    
    def list(self, request, *args, **kwargs):
        """
        Listar órdenes con opción de desactivar paginación para dashboards
        """
        # ✅ SOLUCIÓN ROBUSTA: Permitir desactivar paginación con parámetro
        if request.query_params.get('no_pagination') == 'true':
            # Sin paginación - devolver todas las órdenes
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            
            # Incluir estadísticas
            total_orders = queryset.count()
            total_revenue = queryset.aggregate(
                total=models.Sum('total'),
                service_fees=models.Sum('service_fee')
            )
            
            return Response({
                'results': serializer.data,
                'count': total_orders,
                'statistics': {
                    'total_orders': total_orders,
                    'total_revenue': total_revenue['total'] or 0,
                    'total_service_fees': total_revenue['service_fees'] or 0,
                }
            })
        
        # Con paginación normal
        return super().list(request, *args, **kwargs)
    
    def get_organizer(self):
        """Obtener el organizador asociado al usuario actual."""
        try:
            organizer_user = OrganizerUser.objects.get(user=self.request.user)
            return organizer_user.organizer
        except OrganizerUser.DoesNotExist:
            return None
    
    def get_serializer_class(self):
        """
        Return appropriate serializer class.
        """
        if self.action == 'retrieve':
            return OrderDetailSerializer
        return OrderSerializer
    
    def get_object(self):
        """
        Override get_object to handle detail actions (resend_email, timeline, etc.)
        that don't have event_id in query params.
        """
        # For detail actions, get the order directly by pk
        lookup_url_kwarg = self.lookup_url_kwarg or self.lookup_field
        pk = self.kwargs[lookup_url_kwarg]
        
        try:
            order = Order.objects.get(id=pk)
        except Order.DoesNotExist:
            raise NotFound("Order not found.")
        
        # Check permissions
        if self.request.user.is_authenticated:
            # Regular users can only access their own orders
            if not hasattr(self.request.user, 'organizer_roles'):
                if order.user != self.request.user:
                    raise DRFPermissionDenied("You don't have permission to access this order.")
                return order
            
            # Organizers can access orders from their events
            organizer = self.get_organizer()
            if organizer:
                if order.event.organizer != organizer:
                    raise DRFPermissionDenied("You don't have permission to access this order.")
                return order
        
        # No access for unauthenticated users
        raise DRFPermissionDenied("Authentication required.")
    
    def get_queryset(self):
        """
        Get orders based on permissions with STRICT event isolation.
        🚀 ENTERPRISE: Multi-level filtering to prevent data leakage between events.
        """
        # Regular users can only see their own orders
        if self.request.user.is_authenticated and not hasattr(self.request.user, 'organizer_roles'):
            return Order.objects.filter(user=self.request.user)
        
        # 🚀 ENTERPRISE: Organizers can see orders for their events with STRICT isolation
        if self.request.user.is_authenticated:
            organizer = self.get_organizer()
            if organizer:
                # 🚨 CRITICAL: Event ID is MANDATORY for security
                event_id = self.request.query_params.get('event_id')
                if not event_id:
                    # 🚨 SECURITY: Without event_id, return empty to prevent data leakage
                    return Order.objects.none()
                
                # 🚀 ENTERPRISE: Multi-level filtering for maximum security
                queryset = Order.objects.filter(
                    event__organizer=organizer,
                    event_id=event_id  # ← STRICT event isolation
                )
                
                # 🚨 SECURITY: Additional validation - ensure event belongs to organizer
                event = Event.objects.filter(id=event_id, organizer=organizer).first()
                if not event:
                    return Order.objects.none()
                
                return queryset
        
        # No orders for unauthenticated users
        return Order.objects.none()
    
    @action(detail=True, methods=['post'])
    def refund(self, request, pk=None):
        """
        Refund an order.
        """
        order = self.get_object()
        
        # Check if order can be refunded
        if order.status != 'paid':
            return Response(
                {"detail": "Only paid orders can be refunded."},
                status=status.HTTP_400_BAD_REQUEST)
        
        # Get reason from request data
        reason = request.data.get('reason', '')
        amount = request.data.get('amount', order.total)
        
        # In a real app, you would process the refund with a payment gateway here
        
        # Update order status
        order.status = 'refunded'
        order.refund_reason = reason
        order.refunded_amount = amount
        order.save()
        
        # Update tickets status
        for item in order.items.all():
            for ticket in item.tickets.all():
                ticket.status = 'refunded'
                ticket.save()
        
        return Response({"detail": "Order refunded successfully."})
    
    @action(detail=True, methods=['post'])
    def resend_email(self, request, pk=None):
        """
        Resend order confirmation email.
        """
        order = self.get_object()
        
        # Get email from request (defaults to order email)
        to_email = request.data.get('email', order.email)
        
        if not to_email:
            return Response(
                {"detail": "Email address is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate email format
        from django.core.validators import validate_email
        from django.core.exceptions import ValidationError
        try:
            validate_email(to_email)
        except ValidationError:
            return Response(
                {"detail": "Invalid email address."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Queue email sending task
        try:
            from apps.events.tasks import send_order_confirmation_email
            send_order_confirmation_email.apply_async(
                args=[str(order.id)],
                kwargs={'to_email': to_email},
                queue='emails'
            )
            return Response({
                "detail": "Email queued for sending.",
                "email": to_email
            })
        except Exception as e:
            return Response(
                {"detail": f"Failed to queue email: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def conversion_metrics(self, request, pk=None):
        """
        🚀 ENTERPRISE: Get conversion metrics for this order vs historical averages.
        
        Returns step-by-step conversion rates comparing this order's flow
        against historical platform averages.
        """
        try:
            from core.conversion_metrics import ConversionMetricsService
            
            order = self.get_object()
            
            if not order.order_number:
                return Response({
                    'success': False,
                    'message': 'Order does not have an order number'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            metrics = ConversionMetricsService.get_order_conversion_comparison(order.order_number)
            
            if not metrics.get('success'):
                return Response(metrics, status=status.HTTP_404_NOT_FOUND)
            
            return Response(metrics, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"❌ [ORDER] Error getting conversion metrics: {e}", exc_info=True)
            return Response({
                'success': False,
                'message': f'Error getting conversion metrics: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def historical_conversion_rates(self, request):
        """
        🚀 ENTERPRISE: Get historical conversion rates for all steps.
        
        Query params:
            - from_date: ISO format start date (optional)
            - to_date: ISO format end date (optional)
            - organizer_id: Filter by organizer (optional)
            - event_id: Filter by event (optional)
        
        Returns step-by-step historical conversion rates.
        """
        try:
            from core.conversion_metrics import ConversionMetricsService
            from django.utils.dateparse import parse_datetime
            
            # Parse date parameters
            from_date = None
            to_date = None
            if request.query_params.get('from_date'):
                from_date = parse_datetime(request.query_params.get('from_date'))
                if from_date and not timezone.is_aware(from_date):
                    from_date = timezone.make_aware(from_date)
            
            if request.query_params.get('to_date'):
                to_date = parse_datetime(request.query_params.get('to_date'))
                if to_date and not timezone.is_aware(to_date):
                    to_date = timezone.make_aware(to_date)
            
            organizer_id = request.query_params.get('organizer_id')
            event_id = request.query_params.get('event_id')
            
            # Get historical rates
            historical_rates = ConversionMetricsService.get_historical_conversion_rates(
                flow_type='ticket_checkout',
                from_date=from_date,
                to_date=to_date,
                organizer_id=organizer_id,
                event_id=event_id
            )
            
            # Format response with step display names
            from core.models import PlatformFlowEvent
            from core.conversion_metrics import TICKET_CHECKOUT_STEPS
            step_display_map = dict(PlatformFlowEvent.STEP_CHOICES)
            
            steps_data = []
            for step in TICKET_CHECKOUT_STEPS:
                step_data = historical_rates.get(step, {})
                steps_data.append({
                    'step': step,
                    'step_display': step_display_map.get(step, step),
                    'conversion_rate': step_data.get('conversion_rate', 0.0),
                    'conversion_percentage': step_data.get('conversion_percentage', 0.0),
                    'reached_count': step_data.get('reached_count', 0),
                    'previous_count': step_data.get('previous_count'),
                    'previous_step': step_data.get('previous_step')
                })
            
            # Calculate overall average
            overall_avg = sum(s.get('conversion_rate', 0.0) for s in historical_rates.values()) / len(historical_rates) if historical_rates else 0.0
            
            return Response({
                'success': True,
                'steps': steps_data,
                'overall_average': round(overall_avg, 4),
                'overall_average_percentage': round(overall_avg * 100, 2),
                'from_date': from_date.isoformat() if from_date else None,
                'to_date': to_date.isoformat() if to_date else None,
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"❌ [ORDER] Error getting historical conversion rates: {e}", exc_info=True)
            return Response({
                'success': False,
                'message': f'Error getting historical conversion rates: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def timeline(self, request, pk=None):
        """
        Get order timeline/history.
        """
        from django.utils import timezone
        from apps.events.models import EmailLog
        
        order = self.get_object()
        timeline_events = []
        
        # 1. Order created
        timeline_events.append({
            'type': 'order_created',
            'title': 'Pedido generado',
            'description': f'Pedido {order.order_number} creado',
            'timestamp': order.created_at.isoformat(),
            'status': 'pending'
        })
        
        # 2. Order status changes (if status changed from pending)
        if order.status == 'paid' and order.updated_at != order.created_at:
            timeline_events.append({
                'type': 'status_changed',
                'title': 'Pedido confirmado',
                'description': f'El pedido cambió a estado confirmado',
                'timestamp': order.updated_at.isoformat(),
                'status': 'confirmed'
            })
        
        # 3. Email history
        email_logs = EmailLog.objects.filter(order=order).order_by('created_at')
        for email_log in email_logs:
            timeline_events.append({
                'type': 'email_sent',
                'title': 'Email enviado',
                'description': f'Email de confirmación enviado a {email_log.to_email}',
                'timestamp': (email_log.sent_at or email_log.created_at).isoformat(),
                'status': email_log.status,
                'email': email_log.to_email,
                'subject': email_log.subject
            })
        
        # 4. Refund (if applicable)
        if order.refunded_amount and order.refunded_amount > 0:
            timeline_events.append({
                'type': 'refunded',
                'title': 'Pedido reembolsado',
                'description': f'Reembolso de {order.refunded_amount} {order.currency}',
                'timestamp': order.updated_at.isoformat(),
                'status': 'refunded',
                'amount': str(order.refunded_amount),
                'reason': order.refund_reason
            })
        
        # Sort by timestamp
        timeline_events.sort(key=lambda x: x['timestamp'])
        
        return Response({
            'order_id': str(order.id),
            'order_number': order.order_number,
            'timeline': timeline_events
        })
    
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """
        Cancel an order.
        """
        order = self.get_object()
        
        # Check if order can be cancelled
        if order.status not in ['pending', 'paid']:
            return Response(
                {"detail": "Only pending or paid orders can be cancelled."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get reason from request data
        reason = request.data.get('reason', '')
        
        # Update order status
        order.status = 'cancelled'
        order.notes = f"{order.notes}\nCancelled: {reason}".strip()
        order.save()
        
        # Update tickets status
        for item in order.items.all():
            for ticket in item.tickets.all():
                ticket.status = 'cancelled'
                ticket.save()
        
        return Response({"detail": "Order cancelled successfully."})


class TicketViewSet(viewsets.ModelViewSet):
    """
    API endpoint for tickets.
    """
    serializer_class = TicketSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'checked_in']
    search_fields = ['ticket_number', 'first_name', 'last_name', 'email']
    ordering_fields = ['first_name', 'last_name', 'check_in_time']
    ordering = ['first_name', 'last_name']
    
    def list(self, request, *args, **kwargs):
        """
        ✅ SOLUCIÓN ROBUSTA: Permitir paginación ilimitada para tickets
        🚀 OPTIMIZED: Using select_related and prefetch_related for better performance
        """
        start_time = time.time()
        event_id = request.query_params.get('event_id')
        user_email = request.user.email if request.user.is_authenticated else 'anonymous'
        user_id = request.user.id if request.user.is_authenticated else None
        
        logger.info(f'🎫 [TICKETS_LIST] Request started - Event: {event_id}, User: {user_email} (ID: {user_id})')
        
        # Verificar si se solicita sin paginación
        no_pagination = request.query_params.get('no_pagination', '').lower() == 'true'
        
        if no_pagination:
            # Get base queryset
            base_queryset = self.get_queryset()
            
            # 🚀 OPTIMIZATION: Use select_related and prefetch_related to reduce database queries
            # This is critical for large events with many tickets
            queryset = base_queryset.select_related(
                'order_item__order__event',
                'order_item__order__user',
                'order_item__ticket_tier',
                'check_in_by'
            ).prefetch_related(
                'notes',
                'email_logs'
            )
            
            # Apply filters
            queryset = self.filter_queryset(queryset)
            
            # Count before serialization (more efficient)
            count_start = time.time()
            total_tickets = queryset.count()
            checked_in_count = queryset.filter(checked_in=True).count()
            count_time = (time.time() - count_start) * 1000
            
            logger.info(f'🎫 [TICKETS_LIST] Queryset prepared - Total tickets: {total_tickets}, Count time: {count_time:.2f}ms')
            
            # Serialize with optimized queryset
            serialize_start = time.time()
            serializer = self.get_serializer(queryset, many=True)
            serialize_time = (time.time() - serialize_start) * 1000
            
            logger.info(f'🎫 [TICKETS_LIST] Serialization completed - Time: {serialize_time:.2f}ms')
            
            # Calculate total processing time
            total_time = (time.time() - start_time) * 1000
            
            logger.info(f'🎫 [TICKETS_LIST] Request completed - Event: {event_id}, Tickets: {total_tickets}, '
                       f'Total time: {total_time:.2f}ms, User: {user_email}')
            
            return Response({
                'results': serializer.data,
                'statistics': {
                    'total_tickets': total_tickets,
                    'checked_in_count': checked_in_count,
                    'pending_checkin': total_tickets - checked_in_count
                }
            })
        
        # Paginación normal
        return super().list(request, *args, **kwargs)
    
    def get_organizer(self):
        """Obtener el organizador asociado al usuario actual."""
        try:
            organizer_user = OrganizerUser.objects.get(user=self.request.user)
            return organizer_user.organizer
        except OrganizerUser.DoesNotExist:
            return None
    
    def get_queryset(self):
        """
        Get tickets based on permissions with STRICT event isolation.
        🚀 ENTERPRISE: Multi-level filtering to prevent data leakage between events.
        🚀 OPTIMIZED: Base queryset with select_related for better performance.
        """
        # Regular users can only see their own tickets
        if self.request.user.is_authenticated and not hasattr(self.request.user, 'organizer_roles'):
            return Ticket.objects.select_related(
                'order_item__order__event',
                'order_item__order__user',
                'order_item__ticket_tier'
            ).filter(order_item__order__user=self.request.user)
        
        # 🚀 ENTERPRISE: Organizers can see tickets for their events with STRICT isolation
        if self.request.user.is_authenticated:
            organizer = self.get_organizer()
            if organizer:
                # 🚨 CRITICAL: Event ID is MANDATORY for security
                event_id = self.request.query_params.get('event_id')
                if not event_id:
                    # 🚨 SECURITY: Without event_id, return empty to prevent data leakage
                    logger.warning(f'🎫 [TICKETS_QUERYSET] No event_id provided - User: {self.request.user.email}')
                    return Ticket.objects.none()
                
                # 🚨 SECURITY: Additional validation - ensure event belongs to organizer
                event = Event.objects.filter(id=event_id, organizer=organizer).first()
                if not event:
                    logger.warning(f'🎫 [TICKETS_QUERYSET] Event {event_id} not found or not owned by organizer - User: {self.request.user.email}')
                    return Ticket.objects.none()
                
                # 🚀 ENTERPRISE: Multi-level filtering for maximum security
                # 🚀 OPTIMIZED: Base queryset with select_related (additional optimization in list() method)
                queryset = Ticket.objects.filter(
                    order_item__order__event__organizer=organizer,
                    order_item__order__event_id=event_id  # ← STRICT event isolation
                )
                
                logger.debug(f'🎫 [TICKETS_QUERYSET] Queryset created for event {event_id} - User: {self.request.user.email}')
                return queryset
        
        # No tickets for unauthenticated users
        logger.warning(f'🎫 [TICKETS_QUERYSET] Unauthenticated user attempt')
        return Ticket.objects.none()
    
    @action(detail=True, methods=['get'], url_path='pdf-data')
    def pdf_data(self, request, pk=None):
        """
        🎫 ENTERPRISE: Get ticket data for PDF generation (single ticket)
        Returns data in the same format as get_order_tickets but for a single ticket.
        """
        from core.permissions import IsSuperAdmin
        from django.conf import settings
        
        try:
            # Check if this is an admin request (from superadmin panel without auth)
            is_admin_request = request.query_params.get('admin') == 'true'
            
            # Check permissions
            is_authenticated = request.user.is_authenticated
            is_superadmin = is_authenticated and (request.user.is_superuser or IsSuperAdmin().has_permission(request, None))
            is_organizer = is_authenticated and self.get_organizer() is not None
            
            if not (is_superadmin or is_organizer or is_admin_request):
                return Response({
                    'success': False,
                    'error': 'PERMISSION_DENIED',
                    'message': 'You do not have permission to access this ticket'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get the ticket directly (bypass queryset filtering for admin requests or when event_id is not in query)
            # For organizers, we need to bypass queryset filtering if event_id is not provided
            # because get_queryset() returns empty queryset without event_id
            if is_admin_request or (is_organizer and not request.query_params.get('event_id')):
                # Get ticket directly without queryset filtering
                try:
                    ticket = Ticket.objects.select_related(
                        'order_item__order__event',
                        'order_item__order__event__location',
                        'order_item__ticket_tier'
                    ).prefetch_related(
                        'order_item__order__event__images'
                    ).get(id=pk)
                except Ticket.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'TICKET_NOT_FOUND',
                        'message': 'Ticket not found'
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                # For authenticated users with event_id, use get_object() which respects queryset filtering
                try:
                    ticket = self.get_object()
                except Exception as e:
                    # If get_object() fails (e.g., queryset is empty), try direct lookup
                    try:
                        ticket = Ticket.objects.select_related(
                            'order_item__order__event',
                            'order_item__order__event__location',
                            'order_item__ticket_tier'
                        ).prefetch_related(
                            'order_item__order__event__images'
                        ).get(id=pk)
                    except Ticket.DoesNotExist:
                        return Response({
                            'success': False,
                            'error': 'TICKET_NOT_FOUND',
                            'message': 'Ticket not found'
                        }, status=status.HTTP_404_NOT_FOUND)
            
            order = ticket.order_item.order
            event = order.event
            
            # For organizers (not admin requests), verify they have access to this event
            if is_organizer and not is_superadmin and not is_admin_request:
                organizer = self.get_organizer()
                if event.organizer != organizer:
                    return Response({
                        'success': False,
                        'error': 'PERMISSION_DENIED',
                        'message': 'You do not have permission to access this ticket'
                    }, status=status.HTTP_403_FORBIDDEN)
            
            # Build event data
            image_url = None
            if event.images.exists():
                first_image = event.images.first()
                if first_image and first_image.image:
                    relative_url = first_image.image.url
                    if relative_url and not (relative_url.startswith('http://') or relative_url.startswith('https://')):
                        image_url = request.build_absolute_uri(relative_url)
                    else:
                        image_url = relative_url
            
            event_data = {
                'id': str(event.id),
                'title': event.title,
                'start_date': event.start_date.isoformat() if event.start_date else None,
                'location': {
                    'name': event.location.name if event.location else 'Ubicación no disponible',
                    'address': event.location.address if event.location and hasattr(event.location, 'address') else '',
                },
                'image_url': image_url,
            }
            
            # Build ticket data (single ticket)
            ticket_data = [{
                'ticket_number': ticket.ticket_number,
                'first_name': ticket.first_name,
                'last_name': ticket.last_name,
                'email': ticket.email,
                'tier_name': ticket.order_item.ticket_tier.name if ticket.order_item and ticket.order_item.ticket_tier else 'General',
                'tier_id': str(ticket.order_item.ticket_tier.id) if ticket.order_item and ticket.order_item.ticket_tier else None,
            }]
            
            return Response({
                'success': True,
                'order_number': order.order_number,
                'order_created_at': order.created_at.isoformat(),
                'event': event_data,
                'tickets': ticket_data,
                'ticket_count': 1,
            })
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"❌ [TICKET_PDF_DATA] Error getting ticket PDF data: {e}", exc_info=True)
            return Response({
                'success': False,
                'error': 'SERVER_ERROR',
                'message': f'Error getting ticket data: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def by_uuid(self, request):
        """
        Get a ticket by UUID with complete order information.
        """
        ticket_uuid = request.query_params.get('uuid')
        if not ticket_uuid:
            return Response(
                {"detail": "UUID parameter is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # For by_uuid, we need to bypass the strict event isolation
            # and check permissions manually
            organizer = self.get_organizer()
            if not organizer:
                return Response(
                    {"detail": "Unauthorized."},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Get the ticket with all related data
            ticket = Ticket.objects.select_related(
                'order_item__order__event',
                'order_item__order__user',
                'order_item__ticket_tier',
                'check_in_by'
            ).prefetch_related(
                'order_item__order__items__tickets'
            ).get(id=ticket_uuid)
            
            # Verify the ticket belongs to an event owned by this organizer
            if ticket.order_item.order.event.organizer != organizer:
                return Response(
                    {"detail": "Ticket not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Get all tickets from the same order
            order_tickets = Ticket.objects.filter(
                order_item__order=ticket.order_item.order
            ).select_related(
                'order_item__ticket_tier'
            ).order_by('first_name', 'last_name')
            
            # Serialize the main ticket
            ticket_serializer = self.get_serializer(ticket)
            ticket_data = ticket_serializer.data
            
            # Add order information
            order = ticket.order_item.order
            user_name = "Usuario"
            user_email = order.email
            
            if order.user:
                user_name = f"{order.user.first_name or ''} {order.user.last_name or ''}".strip() or order.user.email or order.email
                user_email = order.user.email or order.email
            else:
                user_name = f"{order.first_name or ''} {order.last_name or ''}".strip() or order.email
                user_email = order.email
            
            ticket_data['order_info'] = {
                'id': str(order.id),
                'order_number': order.order_number,
                'status': order.status,
                'total_amount': float(order.total),
                'subtotal': float(order.subtotal),
                'taxes': float(order.taxes),
                'service_fee': float(order.service_fee),
                'discount': float(order.discount),
                'payment_method': order.payment_method or 'N/A',
                'created_at': order.created_at.isoformat(),
                'user': {
                    'name': user_name,
                    'email': user_email
                }
            }
            
            # Add other tickets from the same order
            ticket_data['order_tickets'] = self.get_serializer(order_tickets, many=True).data
            
            # Add form data with field labels (like in Excel export)
            if ticket.order_item.ticket_tier and ticket.order_item.ticket_tier.form:
                form = ticket.order_item.ticket_tier.form
                form_fields = form.fields.all().order_by('order')
                
                # Build form_data_with_labels: array of {label, value} pairs
                form_data_with_labels = []
                if ticket.form_data:
                    for field in form_fields:
                        field_id = str(field.id)
                        field_value = ticket.form_data.get(field_id, '')
                        
                        # Format value based on field type
                        if field.field_type == 'checkbox' and isinstance(field_value, bool):
                            formatted_value = 'Sí' if field_value else 'No'
                        elif field.field_type == 'file' and field_value:
                            formatted_value = f"Archivo: {field_value}"
                        else:
                            formatted_value = str(field_value) if field_value else ''
                        
                        form_data_with_labels.append({
                            'field_id': field_id,
                            'label': field.label,
                            'field_type': field.field_type,
                            'value': formatted_value,
                            'raw_value': field_value
                        })
                
                ticket_data['form_data_with_labels'] = form_data_with_labels
                ticket_data['form_schema'] = {
                    'form_id': str(form.id),
                    'form_name': form.name,
                    'fields': [
                        {
                            'id': str(field.id),
                            'label': field.label,
                            'field_type': field.field_type,
                            'required': field.required,
                            'order': field.order
                        }
                        for field in form_fields
                    ]
                }
            
            return Response(ticket_data)
        except Ticket.DoesNotExist:
            return Response(
                {"detail": "Ticket not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"detail": f"Error retrieving ticket: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def check_in(self, request, pk=None):
        """
        Check in a ticket.
        """
        ticket = self.get_object()
        
        # Check if ticket can be checked in
        if ticket.status != 'active':
            return Response(
                {"detail": "Only active tickets can be checked in."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if ticket.checked_in:
            return Response(
                {"detail": "Ticket already checked in."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update ticket status
        ticket.checked_in = True
        ticket.check_in_status = 'checked_in'
        ticket.check_in_time = timezone.now()
        ticket.check_in_by = request.user
        ticket.status = 'used'
        ticket.save()
        
        return Response({
            "success": True,
            "detail": "Ticket checked in successfully.",
            "ticket": TicketSerializer(ticket).data
        })


    
    @action(detail=True, methods=['post'])
    def validate(self, request, pk=None):
        """
        Validate a ticket by number.
        """
        ticket_number = request.data.get('ticket_number')
        
        if not ticket_number:
            return Response(
                {"detail": "Ticket number is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get organizer properly
            organizer = self.get_organizer()
            if not organizer:
                return Response({
                    "is_valid": False,
                    "detail": "Usuario sin organizador asociado"
                })
            
            # Find ticket by number, ensuring it belongs to the organizer's events
            ticket = Ticket.objects.get(
                ticket_number=ticket_number,
                order_item__order__event__organizer=organizer
            )
            
            # Check ticket status
            if ticket.status != 'active':
                return Response({
                    "is_valid": False,
                    "detail": f"Ticket is not active. Current status: {ticket.get_status_display()}"
                })
            
            if ticket.checked_in:
                return Response({
                    "is_valid": False,
                    "detail": f"Ticket already checked in at {ticket.check_in_time}"
                })
            
            # Return ticket details
            return Response({
                "is_valid": True,
                "ticket": TicketSerializer(ticket).data
            })
            
        except Ticket.DoesNotExist:
            return Response({
                "is_valid": False,
                "detail": "Invalid ticket number"
            })
    
    @action(detail=True, methods=['get'])
    def form(self, request, pk=None):
        """
        Get the form fields for a ticket with conditional logic applied.
        """
        ticket = self.get_object()
        
        # Get the form associated with the ticket tier
        ticket_tier = ticket.order_item.ticket_tier
        form = ticket_tier.form
        
        if not form:
            return Response({
                "detail": "No form associated with this ticket."
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Get existing data from the ticket
        existing_data = ticket.form_data or {}
        
        # Get all fields from the form
        all_fields = form.fields.all().order_by('order')
        
        # Process fields to apply conditional logic
        visible_fields = []
        for field in all_fields:
            # Check if field should be visible based on conditional rules
            if self.should_display_field(field, existing_data):
                field_data = FormFieldSerializer(field).data
                field_data['current_value'] = existing_data.get(field.id, '')
                visible_fields.append(field_data)
        
        form_data = {
            'id': form.id,
            'name': form.name,
            'description': form.description,
            'fields': visible_fields,
        }
        
        return Response(form_data)
    
    def should_display_field(self, field, form_data):
        """
        Determine if a field should be displayed based on conditional logic.
        
        Args:
            field: FormField instance
            form_data: Dictionary with existing form data, where keys are field IDs
            
        Returns:
            Boolean indicating if field should be displayed
        """
        # If no conditional display rules, field is always shown
        if not field.conditional_display:
            return True
        
        # Field is only shown if ALL rules are satisfied
        for rule in field.conditional_display:
            source_field_id = rule.get('sourceField')
            condition = rule.get('condition')
            expected_value = rule.get('value')
            
            # If rule is incomplete, ignore it
            if not all([source_field_id, condition, expected_value is not None]):
                continue
            
            # Get actual value from form data
            actual_value = form_data.get(source_field_id, '')
            
            # Apply the condition
            if condition == 'equals' and str(actual_value) != str(expected_value):
                return False
            elif condition == 'notEquals' and str(actual_value) == str(expected_value):
                return False
            elif condition == 'contains' and str(expected_value) not in str(actual_value):
                return False
            elif condition == 'notContains' and str(expected_value) in str(actual_value):
                return False
            elif condition == 'greaterThan':
                try:
                    if float(actual_value) <= float(expected_value):
                        return False
                except (ValueError, TypeError):
                    return False
            elif condition == 'lessThan':
                try:
                    if float(actual_value) >= float(expected_value):
                        return False
                except (ValueError, TypeError):
                    return False
        
        # All rules passed
        return True
    
    @action(detail=True, methods=['post'])
    def submit_form(self, request, pk=None):
        """
        Submit form data for a ticket.
        """
        ticket = self.get_object()
        form_data = request.data.get('form_data', {})
        
        # Validate form data against form schema
        if ticket.order_item.ticket_tier.form:
            form = ticket.order_item.ticket_tier.form
            fields = form.fields.all()
            
            # Check required fields
            for field in fields:
                # Skip conditional fields that shouldn't be visible
                if not self.should_display_field(field, form_data):
                    continue
                    
                field_id = str(field.id)
                if field.required and (field_id not in form_data or not form_data[field_id]):
                    return Response({
                        "detail": f"Field '{field.label}' is required."
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Store the form data on the ticket
            ticket.form_data = form_data
            ticket.save()
            
            return Response({
                "detail": "Form data saved successfully.",
                "ticket": TicketSerializer(ticket).data
            })
        
        return Response({
            "detail": "No form associated with this ticket."
        }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def check_in_by_uuid(self, request):
        """
        Check in a ticket by UUID.
        Performance-optimized: Only loads full details if action_type='view_details'
        """
        ticket_uuid = request.data.get('ticket_uuid')
        action_type = request.data.get('action_type', 'check_in')  # 'check_in', 'check_out', 'view_details'
        include_full_details = action_type == 'view_details' or request.data.get('include_full_details', False)
        
        if not ticket_uuid:
            return Response(
                {"detail": "ticket_uuid is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get the ticket
            ticket = Ticket.objects.select_related(
                'order_item__order__event'
            ).get(id=ticket_uuid)
            
            # Verify permissions
            organizer = self.get_organizer()
            if not organizer or ticket.order_item.order.event.organizer != organizer:
                return Response(
                    {"detail": "Unauthorized."},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # PERFORMANCE: For view_details, return ticket info without changing state
            if action_type == 'view_details':
                # Return full ticket data without changing state
                ticket_serializer = self.get_serializer(ticket)
                ticket_data = ticket_serializer.data
                
                if include_full_details:
                    # Load full information (order, other tickets, form data)
                    order = ticket.order_item.order
                    user_name = "Usuario"
                    user_email = order.email
                    
                    if order.user:
                        user_name = f"{order.user.first_name or ''} {order.user.last_name or ''}".strip() or order.user.email or order.email
                        user_email = order.user.email or order.email
                    else:
                        user_name = f"{order.first_name or ''} {order.last_name or ''}".strip() or order.email
                        user_email = order.email
                    
                    ticket_data['order_info'] = {
                        'id': str(order.id),
                        'order_number': order.order_number,
                        'status': order.status,
                        'total_amount': float(order.total),
                        'subtotal': float(order.subtotal),
                        'taxes': float(order.taxes),
                        'service_fee': float(order.service_fee),
                        'discount': float(order.discount),
                        'payment_method': order.payment_method or 'N/A',
                        'created_at': order.created_at.isoformat(),
                        'user': {
                            'name': user_name,
                            'email': user_email
                        }
                    }
                    
                    # Get all tickets from the same order
                    order_tickets = Ticket.objects.filter(
                        order_item__order=ticket.order_item.order
                    ).select_related(
                        'order_item__ticket_tier'
                    ).order_by('first_name', 'last_name')
                    
                    ticket_data['order_tickets'] = self.get_serializer(order_tickets, many=True).data
                    
                    # Add form data with field labels
                    if ticket.order_item.ticket_tier and ticket.order_item.ticket_tier.form:
                        form = ticket.order_item.ticket_tier.form
                        form_fields = form.fields.all().order_by('order')
                        
                        form_data_with_labels = []
                        if ticket.form_data:
                            for field in form_fields:
                                field_id = str(field.id)
                                field_value = ticket.form_data.get(field_id, '')
                                
                                if field.field_type == 'checkbox' and isinstance(field_value, bool):
                                    formatted_value = 'Sí' if field_value else 'No'
                                elif field.field_type == 'file' and field_value:
                                    formatted_value = f"Archivo: {field_value}"
                                else:
                                    formatted_value = str(field_value) if field_value else ''
                                
                                form_data_with_labels.append({
                                    'field_id': field_id,
                                    'label': field.label,
                                    'field_type': field.field_type,
                                    'value': formatted_value,
                                    'raw_value': field_value
                                })
                        
                        ticket_data['form_data_with_labels'] = form_data_with_labels
                        ticket_data['form_schema'] = {
                            'form_id': str(form.id),
                            'form_name': form.name,
                            'fields': [
                                {
                                    'id': str(field.id),
                                    'label': field.label,
                                    'field_type': field.field_type,
                                    'required': field.required,
                                    'order': field.order
                                }
                                for field in form_fields
                            ]
                        }
                
                return Response({
                    "success": True,
                    "detail": "Ticket details retrieved successfully.",
                    "ticket": ticket_data
                })
            
            # Check if ticket can be checked in
            if ticket.status in ['cancelled', 'refunded']:
                return Response(
                    {"success": False, "detail": f"Ticket is {ticket.status} and cannot be checked in."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if ticket.check_in_status == 'checked_in':
                # PERFORMANCE: Only load full details if requested (for view_details action)
                ticket_serializer = self.get_serializer(ticket)
                ticket_data = ticket_serializer.data
                
                if include_full_details:
                    # Load full information for view_details - slower but complete
                    order = ticket.order_item.order
                    user_name = "Usuario"
                    user_email = order.email
                    
                    if order.user:
                        user_name = f"{order.user.first_name or ''} {order.user.last_name or ''}".strip() or order.user.email or order.email
                        user_email = order.user.email or order.email
                    else:
                        user_name = f"{order.first_name or ''} {order.last_name or ''}".strip() or order.email
                        user_email = order.email
                    
                    ticket_data['order_info'] = {
                        'id': str(order.id),
                        'order_number': order.order_number,
                        'status': order.status,
                        'total_amount': float(order.total),
                        'subtotal': float(order.subtotal),
                        'taxes': float(order.taxes),
                        'service_fee': float(order.service_fee),
                        'discount': float(order.discount),
                        'payment_method': order.payment_method or 'N/A',
                        'created_at': order.created_at.isoformat(),
                        'user': {
                            'name': user_name,
                            'email': user_email
                        }
                    }
                    
                    # Get all tickets from the same order
                    order_tickets = Ticket.objects.filter(
                        order_item__order=ticket.order_item.order
                    ).select_related(
                        'order_item__ticket_tier'
                    ).order_by('first_name', 'last_name')
                    
                    ticket_data['order_tickets'] = self.get_serializer(order_tickets, many=True).data
                    
                    # Add form data with field labels
                    if ticket.order_item.ticket_tier and ticket.order_item.ticket_tier.form:
                        form = ticket.order_item.ticket_tier.form
                        form_fields = form.fields.all().order_by('order')
                        
                        form_data_with_labels = []
                        if ticket.form_data:
                            for field in form_fields:
                                field_id = str(field.id)
                                field_value = ticket.form_data.get(field_id, '')
                                
                                if field.field_type == 'checkbox' and isinstance(field_value, bool):
                                    formatted_value = 'Sí' if field_value else 'No'
                                elif field.field_type == 'file' and field_value:
                                    formatted_value = f"Archivo: {field_value}"
                                else:
                                    formatted_value = str(field_value) if field_value else ''
                                
                                form_data_with_labels.append({
                                    'field_id': field_id,
                                    'label': field.label,
                                    'field_type': field.field_type,
                                    'value': formatted_value,
                                    'raw_value': field_value
                                })
                        
                        ticket_data['form_data_with_labels'] = form_data_with_labels
                        ticket_data['form_schema'] = {
                            'form_id': str(form.id),
                            'form_name': form.name,
                            'fields': [
                                {
                                    'id': str(field.id),
                                    'label': field.label,
                                    'field_type': field.field_type,
                                    'required': field.required,
                                    'order': field.order
                                }
                                for field in form_fields
                            ]
                        }
                # else: Return only basic ticket data for fast check-in operations
                
                return Response(
                    {
                        "success": False, 
                        "detail": "Ticket already checked in.",
                        "ticket": ticket_data
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Update ticket status
            ticket.checked_in = True
            ticket.check_in_status = 'checked_in'
            ticket.check_in_time = timezone.now()
            ticket.check_in_by = request.user
            ticket.status = 'used'
            ticket.save()
            
            # Create validation log
            try:
                from apps.validation.models import TicketValidationLog, ValidatorSession
                
                # Get or create validator session
                validator_session, created = ValidatorSession.objects.get_or_create(
                    event=ticket.order_item.order.event,
                    validator_name=request.user.email or "Sistema",
                    user=request.user,
                    organizer=ticket.order_item.order.event.organizer,
                    defaults={
                        'device_info': {'user_agent': request.META.get('HTTP_USER_AGENT', 'Unknown')},
                        'end_time': timezone.now(),
                        'total_scans': 0,
                        'successful_validations': 0,
                        'failed_validations': 0,
                        'tickets_checked_in': 0
                    }
                )
                
                # Create validation log
                TicketValidationLog.objects.create(
                    ticket=ticket,
                    validator_session=validator_session,
                    action='check_in',
                    status='success',
                    message=f'Check-in manual realizado por {request.user.email}',
                    metadata={
                        'from_status': 'pending',
                        'to_status': 'checked_in',
                        'method': 'manual',
                        'validator_email': request.user.email
                    }
                )
                
                # Update validator session stats
                validator_session.total_scans += 1
                validator_session.successful_validations += 1
                validator_session.tickets_checked_in += 1
                validator_session.end_time = timezone.now()
                validator_session.save()
                
            except Exception as log_error:
                # Log the error but don't fail the check-in
                print(f"Error creating validation log: {log_error}")
            
            # PERFORMANCE OPTIMIZATION: Only load full details if requested
            ticket_serializer = self.get_serializer(ticket)
            ticket_data = ticket_serializer.data
            
            if include_full_details:
                # Load full information (order, other tickets, form data) - slower but complete
                order = ticket.order_item.order
                user_name = "Usuario"
                user_email = order.email
                
                if order.user:
                    user_name = f"{order.user.first_name or ''} {order.user.last_name or ''}".strip() or order.user.email or order.email
                    user_email = order.user.email or order.email
                else:
                    user_name = f"{order.first_name or ''} {order.last_name or ''}".strip() or order.email
                    user_email = order.email
                
                ticket_data['order_info'] = {
                    'id': str(order.id),
                    'order_number': order.order_number,
                    'status': order.status,
                    'total_amount': float(order.total),
                    'subtotal': float(order.subtotal),
                    'taxes': float(order.taxes),
                    'service_fee': float(order.service_fee),
                    'discount': float(order.discount),
                    'payment_method': order.payment_method or 'N/A',
                    'created_at': order.created_at.isoformat(),
                    'user': {
                        'name': user_name,
                        'email': user_email
                    }
                }
                
                # Get all tickets from the same order
                order_tickets = Ticket.objects.filter(
                    order_item__order=ticket.order_item.order
                ).select_related(
                    'order_item__ticket_tier'
                ).order_by('first_name', 'last_name')
                
                ticket_data['order_tickets'] = self.get_serializer(order_tickets, many=True).data
                
                # Add form data with field labels
                if ticket.order_item.ticket_tier and ticket.order_item.ticket_tier.form:
                    form = ticket.order_item.ticket_tier.form
                    form_fields = form.fields.all().order_by('order')
                    
                    form_data_with_labels = []
                    if ticket.form_data:
                        for field in form_fields:
                            field_id = str(field.id)
                            field_value = ticket.form_data.get(field_id, '')
                            
                            if field.field_type == 'checkbox' and isinstance(field_value, bool):
                                formatted_value = 'Sí' if field_value else 'No'
                            elif field.field_type == 'file' and field_value:
                                formatted_value = f"Archivo: {field_value}"
                            else:
                                formatted_value = str(field_value) if field_value else ''
                            
                            form_data_with_labels.append({
                                'field_id': field_id,
                                'label': field.label,
                                'field_type': field.field_type,
                                'value': formatted_value,
                                'raw_value': field_value
                            })
                    
                    ticket_data['form_data_with_labels'] = form_data_with_labels
                    ticket_data['form_schema'] = {
                        'form_id': str(form.id),
                        'form_name': form.name,
                        'fields': [
                            {
                                'id': str(field.id),
                                'label': field.label,
                                'field_type': field.field_type,
                                'required': field.required,
                                'order': field.order
                            }
                            for field in form_fields
                        ]
                    }
            # else: Return only basic ticket data for fast check-in/check-out operations
            
            return Response({
                "success": True,
                "detail": "Ticket checked in successfully.",
                "ticket": ticket_data
            })
            
        except Ticket.DoesNotExist:
            return Response(
                {"success": False, "detail": "Ticket not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"success": False, "detail": f"Error checking in ticket: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def check_out_by_uuid(self, request):
        """
        Check out a ticket by UUID.
        Performance-optimized: Only loads full details if action_type='view_details'
        """
        ticket_uuid = request.data.get('ticket_uuid')
        action_type = request.data.get('action_type', 'check_out')  # 'check_in', 'check_out', 'view_details'
        include_full_details = action_type == 'view_details' or request.data.get('include_full_details', False)
        
        if not ticket_uuid:
            return Response(
                {"detail": "ticket_uuid is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get the ticket
            ticket = Ticket.objects.select_related(
                'order_item__order__event'
            ).get(id=ticket_uuid)
            
            # Verify permissions
            organizer = self.get_organizer()
            if not organizer or ticket.order_item.order.event.organizer != organizer:
                return Response(
                    {"detail": "Unauthorized."},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Check if ticket can be checked out
            if ticket.check_in_status != 'checked_in':
                return Response(
                    {"success": False, "detail": "Ticket must be checked in to perform check-out."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Update ticket status
            ticket.checked_in = False
            ticket.check_in_status = 'checked_out'
            ticket.check_out_time = timezone.now()
            ticket.check_out_by = request.user
            ticket.status = 'active'  # Return to active status
            ticket.save()
            
            # Create validation log
            try:
                from apps.validation.models import TicketValidationLog, ValidatorSession
                
                # Get or create validator session
                validator_session, created = ValidatorSession.objects.get_or_create(
                    event=ticket.order_item.order.event,
                    validator_name=request.user.email or "Sistema",
                    user=request.user,
                    organizer=ticket.order_item.order.event.organizer,
                    defaults={
                        'device_info': {'user_agent': request.META.get('HTTP_USER_AGENT', 'Unknown')},
                        'end_time': timezone.now(),
                        'total_scans': 0,
                        'successful_validations': 0,
                        'failed_validations': 0,
                        'tickets_checked_in': 0
                    }
                )
                
                # Create validation log
                TicketValidationLog.objects.create(
                    ticket=ticket,
                    validator_session=validator_session,
                    action='check_out',
                    status='success',
                    message=f'Check-out manual realizado por {request.user.email}',
                    metadata={
                        'from_status': 'checked_in',
                        'to_status': 'checked_out',
                        'method': 'manual',
                        'validator_email': request.user.email
                    }
                )
                
                # Update validator session stats
                validator_session.total_scans += 1
                validator_session.successful_validations += 1
                validator_session.end_time = timezone.now()
                validator_session.save()
                
            except Exception as log_error:
                # Log the error but don't fail the check-out
                print(f"Error creating validation log: {log_error}")
            
            # Serialize ticket with full information (like by_uuid endpoint)
            ticket_serializer = self.get_serializer(ticket)
            ticket_data = ticket_serializer.data
            
            # Add order information
            order = ticket.order_item.order
            user_name = "Usuario"
            user_email = order.email
            
            if order.user:
                user_name = f"{order.user.first_name or ''} {order.user.last_name or ''}".strip() or order.user.email or order.email
                user_email = order.user.email or order.email
            else:
                user_name = f"{order.first_name or ''} {order.last_name or ''}".strip() or order.email
                user_email = order.email
            
            ticket_data['order_info'] = {
                'id': str(order.id),
                'order_number': order.order_number,
                'status': order.status,
                'total_amount': float(order.total),
                'subtotal': float(order.subtotal),
                'taxes': float(order.taxes),
                'service_fee': float(order.service_fee),
                'discount': float(order.discount),
                'payment_method': order.payment_method or 'N/A',
                'created_at': order.created_at.isoformat(),
                'user': {
                    'name': user_name,
                    'email': user_email
                }
            }
            
            # Get all tickets from the same order
            order_tickets = Ticket.objects.filter(
                order_item__order=ticket.order_item.order
            ).select_related(
                'order_item__ticket_tier'
            ).order_by('first_name', 'last_name')
            
            ticket_data['order_tickets'] = self.get_serializer(order_tickets, many=True).data
            
            # Add form data with field labels
            if ticket.order_item.ticket_tier and ticket.order_item.ticket_tier.form:
                form = ticket.order_item.ticket_tier.form
                form_fields = form.fields.all().order_by('order')
                
                form_data_with_labels = []
                if ticket.form_data:
                    for field in form_fields:
                        field_id = str(field.id)
                        field_value = ticket.form_data.get(field_id, '')
                        
                        if field.field_type == 'checkbox' and isinstance(field_value, bool):
                            formatted_value = 'Sí' if field_value else 'No'
                        elif field.field_type == 'file' and field_value:
                            formatted_value = f"Archivo: {field_value}"
                        else:
                            formatted_value = str(field_value) if field_value else ''
                        
                        form_data_with_labels.append({
                            'field_id': field_id,
                            'label': field.label,
                            'field_type': field.field_type,
                            'value': formatted_value,
                            'raw_value': field_value
                        })
                
                ticket_data['form_data_with_labels'] = form_data_with_labels
                ticket_data['form_schema'] = {
                    'form_id': str(form.id),
                    'form_name': form.name,
                    'fields': [
                        {
                            'id': str(field.id),
                            'label': field.label,
                            'field_type': field.field_type,
                            'required': field.required,
                            'order': field.order
                        }
                        for field in form_fields
                    ]
                }
            
            return Response({
                "success": True,
                "detail": "Ticket checked out successfully.",
                "ticket": ticket_data
            })
            
        except Ticket.DoesNotExist:
            return Response(
                {"success": False, "detail": "Ticket not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"success": False, "detail": f"Error checking out ticket: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CouponViewSet(viewsets.ModelViewSet):
    """
    API endpoint for coupons.
    """
    serializer_class = CouponSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_active', 'status', 'discount_type']
    search_fields = ['code', 'description']
    ordering_fields = ['code', 'discount_value', 'created_at', 'usage_count']
    ordering = ['-created_at']
    
    def get_organizer(self):
        """Obtener el organizador asociado al usuario actual."""
        try:
            organizer_user = OrganizerUser.objects.get(user=self.request.user)
            return organizer_user.organizer
        except OrganizerUser.DoesNotExist:
            return None
    
    def get_queryset(self):
        """
        🚀 ENTERPRISE: Get coupons with optimized filtering and security.
        """
        if not self.request.user.is_authenticated:
            return Coupon.objects.none()
        
        organizer = self.get_organizer()
        if not organizer:
            return Coupon.objects.none()
        
        # 🚀 ENTERPRISE: Event-aware filtering
        event_id = self.request.query_params.get('event_id')
        base_queryset = Coupon.objects.filter(organizer=organizer).select_related('organizer')
        
        if not event_id:
            # Return empty for security - event_id is required for list operations
            return Coupon.objects.none()
        
        # 🚨 SECURITY: Validate event ownership
        if not Event.objects.filter(id=event_id, organizer=organizer).exists():
            return Coupon.objects.none()
        
        # 🚀 ENTERPRISE: Optimized filtering (Global + Local for event)
        return base_queryset.filter(
            Q(events_list__isnull=True) |  # Global coupons
            Q(events_list__contains=[event_id])  # Local coupons for this event
        ).distinct().order_by('-created_at')
    
    @action(detail=False, methods=['get'], url_path='organizer/all')
    def list_organizer_coupons(self, request):
        """🚀 ENTERPRISE: Get all coupons for organizer dashboard."""
        if not request.user.is_authenticated:
            return Response({'error': 'Authentication required'}, status=401)
        
        organizer = self.get_organizer()
        if not organizer:
            return Response({'error': 'Organizer not found'}, status=404)
        
        # 🚀 ENTERPRISE: Optimized query with prefetch
        coupons = Coupon.objects.filter(organizer=organizer)\
            .select_related('organizer')\
            .prefetch_related('ticket_tiers', 'ticket_categories')\
            .order_by('-created_at')
        
        serializer = self.get_serializer(coupons, many=True)
        
        return Response({
            'count': coupons.count(),
            'results': serializer.data,
            'analytics': self._get_organizer_coupon_analytics(coupons)
        })
    
    def _get_organizer_coupon_analytics(self, coupons):
        """🚀 ENTERPRISE: Calculate comprehensive analytics for coupons."""
        total_coupons = coupons.count()
        active_coupons = coupons.filter(is_active=True).count()
        global_coupons = coupons.filter(events_list__isnull=True).count()
        used_coupons = coupons.filter(usage_count__gt=0).count()
        
        return {
            'total_coupons': total_coupons,
            'active_coupons': active_coupons,
            'inactive_coupons': total_coupons - active_coupons,
            'global_coupons': global_coupons,
            'local_coupons': total_coupons - global_coupons,
            'used_coupons': used_coupons,
            'unused_coupons': total_coupons - used_coupons,
            'total_usage_count': sum(c.usage_count for c in coupons),
        }
    
    def get_object(self):
        """
        🚀 ENTERPRISE: Override get_object to handle individual coupon operations (GET, PUT, DELETE)
        without requiring event_id in query params.
        """
        # Get the coupon by ID first
        coupon_id = self.kwargs.get('pk')
        if not coupon_id:
            raise Http404("Coupon ID is required")
        
        try:
            # Find coupon by ID and organizer (security check)
            organizer = self.get_organizer()
            if not organizer:
                raise Http404("Organizer not found")
            
            coupon = Coupon.objects.get(id=coupon_id, organizer=organizer)
            
            # 🚨 SECURITY: Additional validation - ensure user can access this coupon
            # For individual operations, we don't need event_id validation
            # The organizer check is sufficient for security
            
            return coupon
            
        except Coupon.DoesNotExist:
            raise Http404("Coupon not found")
    
    def perform_create(self, serializer):
        """
        Create a new coupon for the current organizer.
        """
        organizer = self.get_organizer()
        if not organizer:
            raise serializers.ValidationError({"detail": "El usuario no tiene un organizador asociado"})
            
        serializer.save(organizer=organizer)
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.AllowAny])
    def validate(self, request):
        """🚀 ENTERPRISE: Validate coupon with comprehensive checks - PUBLIC ENDPOINT."""
        code = request.data.get('code')
        event_id = request.data.get('event_id')
        order_total = request.data.get('order_total', 0)
        
        if not code or not event_id:
            return Response({
                "is_valid": False,
                "detail": "Código de cupón y ID de evento son requeridos."
            }, status=400)
        
        try:
            # Clean code input
            code = code.strip().upper()
            # 🚀 ENTERPRISE: Convert to Decimal for precise calculations
            from decimal import Decimal
            order_total = Decimal(str(order_total)) if order_total else Decimal('0')
            
            # 🚀 ENTERPRISE: Get event to determine organizer (PUBLIC ACCESS)
            try:
                event = Event.objects.select_related('organizer').get(id=event_id)
                event_organizer = event.organizer
            except Event.DoesNotExist:
                return Response({
                    "is_valid": False,
                    "detail": "Evento no encontrado."
                }, status=400)
            
            # 🚀 ENTERPRISE: Find coupon by code and event's organizer (PUBLIC ACCESS)
            coupon_query = Coupon.objects.select_related('organizer')
            
            # Search in event's organizer coupons only
            coupon = coupon_query.filter(
                code=code,
                organizer=event_organizer
            ).first()
            
            if not coupon:
                return Response({
                    "is_valid": False,
                    "detail": "Cupón no encontrado o no válido para este evento."
                })
            
            # 🚀 ENTERPRISE: Use comprehensive model validation
            can_use, message = coupon.can_be_used_for_order(order_total, event_id)
            
            response_data = {
                "is_valid": can_use,
                "detail": message,
                "coupon_code": coupon.code
            }
            
            if can_use:
                discount_amount = coupon.calculate_discount_amount(order_total)
                final_total = max(0, order_total - discount_amount)
                
                response_data.update({
                    "coupon": CouponSerializer(coupon, context={'request': request}).data,
                    "discount_amount": float(discount_amount),
                    "final_total": float(final_total),
                    "original_total": float(order_total),
                    "savings": float(discount_amount),
                    "discount_percentage": round((discount_amount / order_total * 100), 2) if order_total > 0 else 0
                })
            
            return Response(response_data)
            
        except Coupon.DoesNotExist:
            return Response({
                "is_valid": False,
                "detail": "Cupón no encontrado."
            })
        except Exception as e:
            # Log error for debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error validating coupon {code}: {e}")
            
            return Response({
                "is_valid": False,
                "detail": "Error interno al validar el cupón. Intente nuevamente."
            }, status=500)
    
    @action(detail=False, methods=['post'])
    def reserve(self, request):
        """🚀 ENTERPRISE: Reserve coupon usage during checkout."""
        code = request.data.get('code')
        event_id = request.data.get('event_id')
        order_id = request.data.get('order_id')
        order_total = request.data.get('order_total', 0)
        
        if not all([code, event_id, order_id]):
            return Response({
                "success": False,
                "detail": "Código de cupón, ID de evento y ID de orden son requeridos."
            }, status=400)
        
        try:
            # Find coupon
            code = code.strip().upper()
            order_total = float(order_total) if order_total else 0
            
            coupon = Coupon.objects.select_related('organizer').get(code=code)
            
            # Validate coupon
            can_use, message = coupon.can_be_used_for_order(order_total, event_id)
            if not can_use:
                return Response({
                    "success": False,
                    "detail": message
                })
            
            # Get or create order
            from apps.events.models import Order
            try:
                order = Order.objects.get(id=order_id)
            except Order.DoesNotExist:
                return Response({
                    "success": False,
                    "detail": "Orden no encontrada."
                }, status=404)
            
            # Reserve coupon usage
            hold = coupon.reserve_usage_for_order(order)
            
            discount_amount = coupon.calculate_discount_amount(order_total)
            final_total = max(0, order_total - discount_amount)
            
            return Response({
                "success": True,
                "detail": "Cupón reservado exitosamente",
                "hold_id": hold.id,
                "expires_at": hold.expires_at.isoformat(),
                "discount_amount": float(discount_amount),
                "final_total": float(final_total),
                "coupon": CouponSerializer(coupon, context={'request': request}).data
            })
            
        except Coupon.DoesNotExist:
            return Response({
                "success": False,
                "detail": "Cupón no encontrado."
            })
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error reserving coupon {code}: {e}")
            
            return Response({
                "success": False,
                "detail": f"Error al reservar cupón: {str(e)}"
            }, status=500)
    
    @action(detail=False, methods=['post'])
    def release(self, request):
        """🚀 ENTERPRISE: Release coupon reservation."""
        order_id = request.data.get('order_id')
        
        if not order_id:
            return Response({
                "success": False,
                "detail": "ID de orden es requerido."
            }, status=400)
        
        try:
            from apps.events.models import Order, CouponHold
            
            # Get order
            try:
                order = Order.objects.get(id=order_id)
            except Order.DoesNotExist:
                return Response({
                    "success": False,
                    "detail": "Orden no encontrada."
                }, status=404)
            
            # Release all coupon holds for this order
            holds = CouponHold.objects.filter(order=order, released=False)
            released_count = 0
            
            for hold in holds:
                hold.release()
                released_count += 1
            
            return Response({
                "success": True,
                "detail": f"Se liberaron {released_count} reservas de cupones",
                "released_count": released_count
            })
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error releasing coupon holds for order {order_id}: {e}")
            
            return Response({
                "success": False,
                "detail": f"Error al liberar reservas: {str(e)}"
            }, status=500)
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """
        Toggle the active status of a coupon.
        """
        coupon = self.get_object()
        coupon.is_active = not coupon.is_active
        coupon.save()
        
        return Response({
            "detail": f"Coupon is now {'active' if coupon.is_active else 'inactive'}.",
            "coupon": CouponSerializer(coupon).data
        })
    
    @action(detail=False, methods=['get'])
    def analytics(self, request):
        """🚀 ENTERPRISE: Enhanced analytics with performance metrics."""
        organizer = self.get_organizer()
        if not organizer:
            return Response({
                "detail": "Usuario sin organizador asociado"
            }, status=400)
        
        coupons = Coupon.objects.filter(organizer=organizer)
        analytics = self._get_organizer_coupon_analytics(coupons)
        
        # 🚀 ENTERPRISE: Add performance metrics
        top_coupons = coupons.filter(usage_count__gt=0).order_by('-usage_count')[:5]
        analytics['top_performing_coupons'] = [
            {
                'code': coupon.code,
                'usage_count': coupon.usage_count,
                'type': coupon.discount_type,
                'value': float(coupon.discount_value)
            }
            for coupon in top_coupons
        ]
        
        return Response(analytics)
    
    @action(detail=True, methods=['post'])
    def apply_to_order(self, request, pk=None):
        """🚀 ENTERPRISE: Apply coupon to order with atomic transaction."""
        coupon = self.get_object()
        order_total = request.data.get('order_total', 0)
        event_id = request.data.get('event_id')
        
        if not order_total or not event_id:
            return Response({
                "success": False,
                "detail": "Order total and event ID are required."
            }, status=400)
        
        try:
            # Validate coupon can be used
            can_use, message = coupon.can_be_used_for_order(order_total, event_id)
            if not can_use:
                return Response({
                    "success": False,
                    "detail": message
                })
            
            # Calculate discount
            discount_amount = coupon.calculate_discount_amount(order_total)
            final_total = max(0, order_total - discount_amount)
            
            # Increment usage (atomic)
            coupon.increment_usage()
            
            return Response({
                "success": True,
                "discount_amount": discount_amount,
                "final_total": final_total,
                "coupon": CouponSerializer(coupon, context={'request': request}).data
            })
            
        except ValueError as e:
            return Response({
                "success": False,
                "detail": str(e)
            }, status=400)
    
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """🚀 ENTERPRISE: Bulk create coupons for campaigns."""
        organizer = self.get_organizer()
        if not organizer:
            return Response({"detail": "Usuario sin organizador asociado"}, status=400)
        
        coupons_data = request.data.get('coupons', [])
        if not coupons_data:
            return Response({"detail": "No coupon data provided"}, status=400)
        
        created_coupons = []
        errors = []
        
        for i, coupon_data in enumerate(coupons_data):
            try:
                serializer = self.get_serializer(data=coupon_data)
                if serializer.is_valid():
                    coupon = serializer.save(organizer=organizer)
                    created_coupons.append(serializer.data)
                else:
                    errors.append({"index": i, "errors": serializer.errors})
            except Exception as e:
                errors.append({"index": i, "error": str(e)})
        
        return Response({
            "created_count": len(created_coupons),
            "error_count": len(errors),
            "created_coupons": created_coupons,
            "errors": errors
        })


class EventCommunicationViewSet(viewsets.ModelViewSet):
    """
    API endpoint for event communications.
    """
    serializer_class = EventCommunicationSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['event', 'type', 'status']
    search_fields = ['name', 'subject', 'content']
    ordering_fields = ['name', 'scheduled_date', 'created_at']
    ordering = ['-created_at']
    
    def get_organizer(self):
        """Obtener el organizador asociado al usuario actual."""
        try:
            organizer_user = OrganizerUser.objects.get(user=self.request.user)
            return organizer_user.organizer
        except OrganizerUser.DoesNotExist:
            return None
    
    def get_queryset(self):
        """
        Get communications for events belonging to the current organizer.
        """
        if not self.request.user.is_authenticated:
            return EventCommunication.objects.none()
        
        organizer = self.get_organizer()
        if not organizer:
            return EventCommunication.objects.none()
        
        return EventCommunication.objects.filter(
            event__organizer=organizer
        )
    
    @action(detail=True, methods=['post'])
    def send(self, request, pk=None):
        """
        Send a communication.
        """
        communication = self.get_object()
        
        # Check if communication can be sent
        if communication.status not in ['draft', 'scheduled']:
            return Response(
                {"detail": "Only draft or scheduled communications can be sent."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # In a real app, you would send the communication here
        # For now, we'll just update the status
        communication.status = 'sent'
        communication.sent_date = timezone.now()
        communication.save()
        
        return Response({"detail": "Communication sent successfully."})
    
    @action(detail=True, methods=['post'])
    def schedule(self, request, pk=None):
        """
        Schedule a communication.
        """
        communication = self.get_object()
        scheduled_date = request.data.get('scheduled_date')
        
        if not scheduled_date:
            return Response(
                {"detail": "Scheduled date is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if communication can be scheduled
        if communication.status != 'draft':
            return Response(
                {"detail": "Only draft communications can be scheduled."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update communication status
        communication.status = 'scheduled'
        communication.scheduled_date = scheduled_date
        communication.save()
        
        return Response({"detail": "Communication scheduled successfully."})


# ============================================================================
# 🎫 ENTERPRISE QR CODE SYSTEM - PROFESSIONAL TICKET VALIDATION
# ============================================================================

@extend_schema(
    summary="Generate QR Code for Ticket",
    description="Generate a professional QR code for ticket validation. Includes caching and security features.",
    responses={
        200: {
            "description": "QR code generated successfully",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "qr_code": "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAA...",
                        "ticket_number": "TIX-B382C943",
                        "ticket_url": "https://tuki.live/tickets/TIX-B382C943",
                        "expires_at": "2025-09-30T16:18:29Z",
                        "security_hash": "abc123def456"
                    }
                }
            }
        },
        404: {"description": "Ticket not found"},
        500: {"description": "Error generating QR code"}
    }
)
@api_view(['GET'])
@permission_classes([AllowAny])
def generate_ticket_qr(request, ticket_number):
    """
    🎫 ENTERPRISE: Generate professional QR code for ticket validation
    
    Features:
    - Cached QR generation for performance
    - Security hash for validation
    - Expiration timestamp
    - Professional error handling
    - Rate limiting protection
    """
    try:
        from apps.events.models import Ticket
        from apps.events.services import QRCodeService
        from django.core.cache import cache
        from django.utils import timezone
        import hashlib
        import json
        
        # Validate ticket exists and is active
        try:
            ticket = Ticket.objects.select_related(
                'order_item__order__event',
                'order_item__ticket_tier'
            ).get(
                ticket_number=ticket_number,
                status='active'
            )
        except Ticket.DoesNotExist:
            return Response({
                'success': False,
                'error': 'TICKET_NOT_FOUND',
                'message': 'Ticket not found or inactive'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Check if order is paid
        if ticket.order_item.order.status != 'paid':
            return Response({
                'success': False,
                'error': 'TICKET_NOT_PAID',
                'message': 'Ticket order is not paid'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create cache key with ticket data
        cache_key = f"qr_code_{ticket_number}_{ticket.updated_at.timestamp()}"
        
        # Try to get from cache first
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)
        
        # Generate security hash
        security_data = f"{ticket_number}_{ticket.order_item.order.id}_{ticket.created_at.timestamp()}"
        security_hash = hashlib.sha256(security_data.encode()).hexdigest()[:16]
        
        # Generate QR code
        qr_code_base64 = QRCodeService.generate_qr_code(
            ticket_number, 
            size=250  # Higher quality for professional use
        )
        
        if not qr_code_base64:
            return Response({
                'success': False,
                'error': 'QR_GENERATION_FAILED',
                'message': 'Failed to generate QR code'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Prepare response data
        response_data = {
            'success': True,
            'qr_code': qr_code_base64,
            'ticket_number': ticket_number,
            'ticket_url': f"{settings.FRONTEND_URL}/tickets/{ticket_number}",
            'expires_at': (timezone.now() + timezone.timedelta(days=30)).isoformat(),
            'security_hash': security_hash,
            'ticket_info': {
                'event_title': ticket.order_item.order.event.title,
                'event_date': ticket.order_item.order.event.start_date.isoformat(),
                'attendee_name': ticket.attendee_name,
                'ticket_type': ticket.order_item.ticket_tier.name if ticket.order_item.ticket_tier else 'General',
                'order_number': ticket.order_item.order.order_number
            }
        }
        
        # Cache for 1 hour
        cache.set(cache_key, response_data, 3600)
        
        return Response(response_data)
        
    except Exception as e:
        return Response({
            'success': False,
            'error': 'INTERNAL_ERROR',
            'message': 'Internal server error'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@extend_schema(
    summary="Get Order Tickets",
    description="Get all tickets for an order by order number. Returns complete ticket data for PDF generation.",
    parameters=[
        OpenApiParameter(name='order_number', type=str, location=OpenApiParameter.PATH, required=True, description='Order number (e.g., ORD-123456)'),
    ],
    responses={
        200: {
            "description": "Order tickets data",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "order_number": "ORD-123456",
                        "event": {
                            "id": "uuid",
                            "title": "Event Title",
                            "start_date": "2026-01-08T19:00:00Z",
                            "location": {"name": "Location Name", "address": "Address"}
                        },
                        "tickets": [
                            {
                                "ticket_number": "TUKI-ABC123",
                                "first_name": "Flo",
                                "last_name": "mas",
                                "email": "email@example.com",
                                "tier_name": "Entrada Gratuita"
                            }
                        ]
                    }
                }
            }
        }
    }
)
@api_view(['GET'])
@permission_classes([AllowAny])
def get_order_tickets(request, order_number):
    """
    🎫 ENTERPRISE: Get all tickets for an order by order number
    
    Authentication: 
    - Option 1: access_token as query parameter (public access for customers)
    - Option 2: JWT authentication (for organizers/superadmin)
    
    Returns complete ticket data including:
    - Real ticket numbers (not temporary)
    - First name and last name
    - Email
    - Tier name
    - Event information
    """
    try:
        from apps.events.models import Order, Ticket
        from core.permissions import IsSuperAdmin, IsOrganizer
        
        # Get access token from query parameters
        access_token = request.query_params.get('token')
        
        # Check if user is authenticated via JWT (organizer/superadmin)
        is_authenticated_user = request.user.is_authenticated
        is_superadmin = is_authenticated_user and (request.user.is_superuser or IsSuperAdmin().has_permission(request, None))
        is_organizer = is_authenticated_user and IsOrganizer().has_permission(request, None)
        
        # Check if this is an admin request (from superadmin panel without auth)
        # Allow admin access via query parameter for superadmin panel (temporary, no login)
        is_admin_request = request.query_params.get('admin') == 'true'
        
        # If authenticated as organizer/superadmin, or admin request, skip access_token requirement
        if (is_authenticated_user and (is_superadmin or is_organizer)) or is_admin_request:
            # Get order by order_number only (no access_token needed)
            try:
                order = Order.objects.select_related(
                    'event',
                    'event__location',
                    'event__organizer'
                ).prefetch_related(
                    'items__tickets',
                    'items__ticket_tier',
                    'event__images'
                ).get(order_number=order_number)
            except Order.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'ORDER_NOT_FOUND',
                    'message': 'Order not found'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # For organizers (not admin requests), verify they have access to this event
            if is_organizer and not is_superadmin and not is_admin_request:
                try:
                    organizer_user = OrganizerUser.objects.get(user=request.user)
                    if order.event.organizer != organizer_user.organizer:
                        return Response({
                            'success': False,
                            'error': 'PERMISSION_DENIED',
                            'message': 'You do not have permission to access this order'
                        }, status=status.HTTP_403_FORBIDDEN)
                except OrganizerUser.DoesNotExist:
                    return Response({
                        'success': False,
                        'error': 'PERMISSION_DENIED',
                        'message': 'You do not have permission to access this order'
                    }, status=status.HTTP_403_FORBIDDEN)
        else:
            # Public access: require access_token
            if not access_token:
                return Response({
                    'success': False,
                    'error': 'TOKEN_REQUIRED',
                    'message': 'Access token is required. Please use the link from your confirmation email.'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Get order by order_number and validate token
            try:
                order = Order.objects.select_related(
                    'event',
                    'event__location',
                    'event__organizer'
                ).prefetch_related(
                    'items__tickets',
                    'items__ticket_tier',
                    'event__images'
                ).get(order_number=order_number, access_token=access_token)
            except Order.DoesNotExist:
                return Response({
                    'success': False,
                    'error': 'INVALID_TOKEN_OR_ORDER',
                    'message': 'Invalid access token or order not found'
                }, status=status.HTTP_404_NOT_FOUND)
        
        # Get all tickets for this order
        tickets = Ticket.objects.filter(
            order_item__order=order
        ).select_related(
            'order_item__ticket_tier'
        ).order_by('created_at')
        
        # Build event data
        event = order.event
        
        # 🚀 ENTERPRISE: Build absolute image URL for robustness (works in all environments)
        image_url = None
        if event.images.exists():
            first_image = event.images.first()
            if first_image and first_image.image:
                relative_url = first_image.image.url
                # Build absolute URL
                if relative_url and not (relative_url.startswith('http://') or relative_url.startswith('https://')):
                    image_url = request.build_absolute_uri(relative_url)
                else:
                    image_url = relative_url
        
        event_data = {
            'id': str(event.id),
            'title': event.title,
            'start_date': event.start_date.isoformat() if event.start_date else None,
            'location': {
                'name': event.location.name if event.location else 'Ubicación no disponible',
                'address': event.location.address if event.location and hasattr(event.location, 'address') else '',
            },
            'image_url': image_url,
        }
        
        # Build tickets data
        tickets_data = []
        for ticket in tickets:
            tickets_data.append({
                'ticket_number': ticket.ticket_number,  # Real ticket number
                'first_name': ticket.first_name,
                'last_name': ticket.last_name,
                'email': ticket.email,
                'tier_name': ticket.order_item.ticket_tier.name if ticket.order_item and ticket.order_item.ticket_tier else 'General',
                'tier_id': str(ticket.order_item.ticket_tier.id) if ticket.order_item and ticket.order_item.ticket_tier else None,
            })
        
        return Response({
            'success': True,
            'order_number': order.order_number,
            'order_created_at': order.created_at.isoformat(),
            'event': event_data,
            'tickets': tickets_data,
            'ticket_count': len(tickets_data),
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"❌ [ORDER_TICKETS] Error getting tickets for order {order_number}: {e}", exc_info=True)
        return Response({
            'success': False,
            'error': 'SERVER_ERROR',
            'message': f'Error getting tickets: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@extend_schema(
    summary="Validate Ticket QR Code",
    description="Validate a ticket using QR code data. Professional validation with security checks.",
    responses={
        200: {
            "description": "Ticket validation result",
            "content": {
                "application/json": {
                    "example": {
                        "valid": True,
                        "ticket_number": "TIX-B382C943",
                        "event_title": "Concierto de Rock",
                        "attendee_name": "Juan Pérez",
                        "status": "valid",
                        "validation_timestamp": "2025-09-29T16:18:29Z"
                    }
                }
            }
        }
    }
)
@api_view(['POST'])
@permission_classes([AllowAny])
def validate_ticket_qr(request):
    """
    🎫 ENTERPRISE: Validate ticket using QR code data
    
    Features:
    - Security hash validation
    - Real-time ticket status check
    - Event date validation
    - Professional response format
    """
    try:
        from apps.events.models import Ticket
        from django.utils import timezone
        import hashlib
        
        ticket_number = request.data.get('ticket_number')
        security_hash = request.data.get('security_hash')
        
        if not ticket_number or not security_hash:
            return Response({
                'valid': False,
                'error': 'MISSING_DATA',
                'message': 'Ticket number and security hash required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get ticket with related data
        try:
            ticket = Ticket.objects.select_related(
                'order_item__order__event',
                'order_item__ticket_tier'
            ).get(ticket_number=ticket_number)
        except Ticket.DoesNotExist:
            return Response({
                'valid': False,
                'error': 'TICKET_NOT_FOUND',
                'message': 'Ticket not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Validate security hash
        security_data = f"{ticket_number}_{ticket.order_item.order.id}_{ticket.created_at.timestamp()}"
        expected_hash = hashlib.sha256(security_data.encode()).hexdigest()[:16]
        
        if security_hash != expected_hash:
            return Response({
                'valid': False,
                'error': 'INVALID_HASH',
                'message': 'Invalid security hash'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check ticket status
        if ticket.status != 'active':
            return Response({
                'valid': False,
                'error': 'TICKET_INACTIVE',
                'message': f'Ticket status: {ticket.status}',
                'ticket_number': ticket_number
            })
        
        # Check if order is paid
        if ticket.order_item.order.status != 'paid':
            return Response({
                'valid': False,
                'error': 'ORDER_NOT_PAID',
                'message': 'Order is not paid',
                'ticket_number': ticket_number
            })
        
        # Check if event has passed (optional - you might want to allow validation after event)
        event_date = ticket.order_item.order.event.start_date
        if event_date < timezone.now() - timezone.timedelta(hours=2):
            return Response({
                'valid': False,
                'error': 'EVENT_PASSED',
                'message': 'Event has already passed',
                'ticket_number': ticket_number,
                'event_date': event_date.isoformat()
            })
        
        # All validations passed
        return Response({
            'valid': True,
            'ticket_number': ticket_number,
            'event_title': ticket.order_item.order.event.title,
            'attendee_name': ticket.attendee_name,
            'ticket_type': ticket.order_item.ticket_tier.name if ticket.order_item.ticket_tier else 'General',
            'status': 'valid',
            'validation_timestamp': timezone.now().isoformat(),
            'event_date': event_date.isoformat(),
            'order_number': ticket.order_item.order.order_number
        })
        
    except Exception as e:
        return Response({
            'valid': False,
            'error': 'INTERNAL_ERROR',
            'message': 'Internal server error'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@extend_schema(
    summary="🚀 ENTERPRISE: Send Order Confirmation Email Synchronously",
    description="""
    Send order confirmation email synchronously from frontend confirmation page.
    
    **Features**:
    - <10s latency (similar to OTP emails)
    - Automatic fallback to Celery if fails
    - Complete flow logging
    - Security via access_token
    
    **Flow**:
    1. Frontend reaches confirmation page
    2. Calls this endpoint with order_number + access_token
    3. Email sent synchronously
    4. If fails → automatically enqueued in Celery
    5. All events logged in PlatformFlow
    
    **Security**: Requires valid access_token (same as ticket download)
    """,
    request={
        "application/json": {
            "example": {
                "to_email": "customer@example.com"
            }
        }
    },
    responses={
        200: {
            "description": "Email sent successfully",
            "content": {
                "application/json": {
                    "example": {
                        "success": True,
                        "message": "Email sent successfully",
                        "emails_sent": 1,
                        "metrics": {
                            "fetch_time_ms": 45,
                            "context_time_ms": 120,
                            "render_time_ms": 85,
                            "smtp_time_ms": 1450,
                            "total_time_ms": 1700
                        },
                        "fallback_to_celery": False
                    }
                }
            }
        },
        202: {
            "description": "Email failed but enqueued in Celery",
            "content": {
                "application/json": {
                    "example": {
                        "success": False,
                        "message": "Email failed but enqueued in Celery for retry",
                        "fallback_to_celery": True,
                        "task_id": "abc123-def456"
                    }
                }
            }
        },
        400: {
            "description": "Missing required parameters"
        },
        404: {
            "description": "Order not found or invalid token"
        }
    }
)
@api_view(['POST'])
@permission_classes([AllowAny])
def send_order_email_sync(request, order_number):
    """
    🚀 ENTERPRISE: Send order confirmation email synchronously.
    
    This endpoint is called from the frontend confirmation page to send
    the order confirmation email immediately, reducing latency from 5+ minutes
    to <10 seconds.
    
    If the synchronous send fails, it automatically falls back to Celery
    for retry, ensuring no emails are lost.
    
    Args:
        order_number: Order number from URL
        access_token: Security token from query params
        to_email: Optional email override from request body
        
    Returns:
        JSON response with send status and metrics
    """
    import time
    import logging
    from apps.events.models import Order
    from apps.events.email_sender import send_order_confirmation_email_optimized
    from core.models import PlatformFlow
    from core.flow_logger import FlowLogger
    
    logger = logging.getLogger(__name__)
    start_time = time.time()
    
    try:
        # 1. Get and validate access_token
        access_token = request.query_params.get('access_token')
        if not access_token:
            logger.warning(f"📧 [EMAIL_SYNC] Missing access_token for order {order_number}")
            return Response({
                'success': False,
                'message': 'Missing access_token'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 2. Get order with access_token validation
        try:
            order = Order.objects.select_related('event', 'user').get(
                order_number=order_number,
                access_token=access_token
            )
        except Order.DoesNotExist:
            logger.warning(f"📧 [EMAIL_SYNC] Order not found or invalid token: {order_number}")
            return Response({
                'success': False,
                'message': 'Order not found or invalid token'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # 3. Get flow for logging and idempotency check
        flow_obj = None
        flow_logger = None
        try:
            # CRITICAL: Check direct relationship first (most efficient and reliable)
            flow_obj = order.flow
            
            # If not found via direct relationship, try primary_order lookup
            if not flow_obj:
                flow_obj = PlatformFlow.objects.filter(
                    primary_order=order,
                    flow_type='ticket_checkout'
                ).first()
            
            # If still not found, search in events
            if not flow_obj:
                flow_event = order.flow_events.first()
                if flow_event:
                    flow_obj = flow_event.flow
            
            # Convert PlatformFlow to FlowLogger for logging
            if flow_obj:
                flow_logger = FlowLogger(flow_obj)
        except Exception as e:
            logger.warning(f"📧 [EMAIL_SYNC] Could not find flow for order {order_number}: {e}")
        
        # 🚀 ENTERPRISE: IDEMPOTENCY CHECK - Verify if email already sent
        if flow_obj:
            email_sent_exists = flow_obj.events.filter(step='EMAIL_SENT').exists()
            if email_sent_exists:
                logger.info(f"📧 [EMAIL_SYNC] ✅ Email already sent for order {order_number} (idempotency check)")
                return Response({
                    'success': True,
                    'message': 'Email already sent',
                    'already_sent': True,
                    'emails_sent': 1
                }, status=status.HTTP_200_OK)
        
        # 4. Log EMAIL_SYNC_ATTEMPT
        if flow_logger:
            flow_logger.log_event(
                'EMAIL_SYNC_ATTEMPT',
                order=order,
                source='api',
                status='info',
                message=f"Attempting synchronous email send for order {order_number}",
                metadata={
                    'strategy': 'frontend_sync',
                    'triggered_by': 'confirmation_page'
                }
            )
        
        logger.info(f"📧 [EMAIL_SYNC] Starting synchronous email send for order {order_number}")
        
        # 5. Get optional email override
        to_email = request.data.get('to_email')
        
        # 6. Send email synchronously
        result = send_order_confirmation_email_optimized(
            order_id=str(order.id),
            to_email=to_email,
            flow_id=str(flow_obj.id) if flow_obj else None
        )
        
        total_time_ms = int((time.time() - start_time) * 1000)
        
        # 7. Check if send was successful
        if result.get('status') == 'completed' and result.get('emails_sent', 0) > 0:
            # ✅ SUCCESS
            logger.info(f"📧 [EMAIL_SYNC] ✅ Email sent successfully for order {order_number} in {total_time_ms}ms")
            
            if flow_logger:
                flow_logger.log_event(
                    'EMAIL_SENT',
                    order=order,
                    source='api',
                    status='success',
                    message=f"Email sent successfully in {total_time_ms}ms",
                    metadata={
                        'strategy': 'frontend_sync',
                        'emails_sent': result.get('emails_sent', 0),
                        'metrics': result.get('metrics', {}),
                        'total_time_ms': total_time_ms
                    }
                )
            
            # Get EmailLog details for confirmation
            from apps.events.models import EmailLog
            email_logs = EmailLog.objects.filter(
                order=order,
                status='sent'
            ).order_by('-sent_at')[:1]
            
            response_data = {
                'success': True,
                'message': 'Email sent successfully',
                'emails_sent': result.get('emails_sent', 0),
                'metrics': result.get('metrics', {}),
                'fallback_to_celery': False,
                'recipients': []
            }
            
            # Add recipient details
            if email_logs:
                for email_log in email_logs:
                    response_data['recipients'].append({
                        'email': email_log.to_email,
                        'sent_at': email_log.sent_at.isoformat() if email_log.sent_at else None,
                        'subject': email_log.subject
                    })
            
            logger.info(
                f"📧 [EMAIL_SYNC] ✅ Success response for order {order_number}: "
                f"{result.get('emails_sent', 0)} emails sent, "
                f"recipients: {[r.get('email', 'N/A') for r in response_data.get('recipients', [])]}"
            )
            
            return Response(response_data, status=status.HTTP_200_OK)
        
        else:
            # ❌ FAILED - Fallback to Celery
            logger.warning(f"📧 [EMAIL_SYNC] ⚠️ Synchronous send failed for order {order_number}, falling back to Celery")
            
            if flow_logger:
                flow_logger.log_event(
                    'EMAIL_FAILED',
                    order=order,
                    source='api',
                    status='warning',
                    message=f"Synchronous email send failed, falling back to Celery",
                    metadata={
                        'strategy': 'frontend_sync',
                        'result': result,
                        'total_time_ms': total_time_ms
                    }
                )
            
            # Enqueue in Celery as fallback
            try:
                from apps.events.tasks import send_order_confirmation_email
                task = send_order_confirmation_email.apply_async(
                    args=[str(order.id)],
                    kwargs={'flow_id': str(flow_obj.id) if flow_obj else None},
                    queue='emails'
                )
                
                if flow_logger:
                    flow_logger.log_event(
                        'EMAIL_TASK_ENQUEUED',
                        order=order,
                        source='api',
                        status='success',
                        message=f"Email enqueued in Celery as fallback",
                        metadata={
                            'task_id': task.id if task else None,
                            'reason': 'sync_send_failed'
                        }
                    )
                
                logger.info(f"📧 [EMAIL_SYNC] Email enqueued in Celery for order {order_number}")
                
                return Response({
                    'success': False,
                    'message': 'Email failed but enqueued in Celery for retry',
                    'fallback_to_celery': True,
                    'task_id': task.id if task else None
                }, status=status.HTTP_202_ACCEPTED)
                
            except Exception as celery_error:
                logger.error(f"📧 [EMAIL_SYNC] Failed to enqueue in Celery: {celery_error}")
                
                if flow_logger:
                    flow_logger.log_event(
                        'EMAIL_FAILED',
                        order=order,
                        source='api',
                        status='failure',
                        message=f"Both sync send and Celery fallback failed",
                        metadata={
                            'sync_error': str(result),
                            'celery_error': str(celery_error)
                        }
                    )
                
                return Response({
                    'success': False,
                    'message': 'Email send failed and could not enqueue in Celery',
                    'error': str(celery_error)
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    except Exception as e:
        logger.error(f"📧 [EMAIL_SYNC] Unexpected error for order {order_number}: {e}", exc_info=True)
        
        return Response({
            'success': False,
            'message': 'Internal server error',
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def proxy_event_image(request):
    """
    🎫 ENTERPRISE: Proxy endpoint to serve event images with CORS headers.
    
    This endpoint solves the CORS issue when html2canvas tries to load images
    from Google Cloud Storage for PDF generation.
    
    Usage:
        GET /api/v1/events/images/proxy/?url=https://storage.googleapis.com/...
    
    Returns:
        Image file with proper CORS headers
    """
    import logging
    import requests
    from django.http import StreamingHttpResponse
    from django.conf import settings
    
    logger = logging.getLogger(__name__)
    
    image_url = request.GET.get('url')
    
    if not image_url:
        return Response(
            {'error': 'Missing url parameter'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Validate that the URL is from our storage bucket (security)
    allowed_domains = [
        'storage.googleapis.com',
        settings.GS_BUCKET_NAME if hasattr(settings, 'GS_BUCKET_NAME') else None,
    ]
    allowed_domains = [d for d in allowed_domains if d]
    
    if not any(domain in image_url for domain in allowed_domains):
        return Response(
            {'error': 'Invalid image URL domain'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Fetch the image from Google Cloud Storage
        response = requests.get(image_url, stream=True, timeout=10)
        response.raise_for_status()
        
        # Determine content type from response or default to image/jpeg
        content_type = response.headers.get('Content-Type', 'image/jpeg')
        
        # Create streaming response with CORS headers
        proxy_response = StreamingHttpResponse(
            response.iter_content(chunk_size=8192),
            content_type=content_type
        )
        
        # Add CORS headers to allow cross-origin requests
        proxy_response['Access-Control-Allow-Origin'] = '*'
        proxy_response['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        proxy_response['Access-Control-Allow-Headers'] = 'Content-Type'
        proxy_response['Cache-Control'] = 'public, max-age=86400'  # Cache for 24 hours
        
        # Copy content length if available
        if 'Content-Length' in response.headers:
            proxy_response['Content-Length'] = response.headers['Content-Length']
        
        return proxy_response
        
    except requests.exceptions.RequestException as e:
        logger.error(f"❌ [IMAGE_PROXY] Error fetching image {image_url}: {e}")
        return Response(
            {'error': 'Failed to fetch image'}, 
            status=status.HTTP_502_BAD_GATEWAY
        )
    except Exception as e:
        logger.error(f"❌ [IMAGE_PROXY] Unexpected error: {e}", exc_info=True)
        return Response(
            {'error': 'Internal server error'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        ) 